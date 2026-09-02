"""Flow Execution Engine — DAG-walking runtime for Agent Flow workflows.

Walks a flow's node graph in topological order, executing each node type:
- Input: provides initial context
- Samurai: delegates to LLM via agent's routing profile
- Coding: performs governed, programming-memory-aware IDE operations
- Shogun Approval: gate that checks approval policy
- Logic/Decision: evaluates condition to select branches
- Output: formats and stores final result

Supports parallel execution of independent sibling nodes.
"""

from __future__ import annotations

import asyncio
import hashlib
import json
import logging
import re
import time
import uuid
from collections import defaultdict, deque
from collections.abc import Awaitable, Callable
from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import httpx
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from shogun.db.engine import async_session_factory
from shogun.db.models.agent import Agent
from shogun.db.models.agent_flow import AgentFlow, AgentFlowEdge, AgentFlowNode
from shogun.db.models.agent_flow_run import AgentFlowRun, AgentFlowRunEdge
from shogun.db.models.model_definition import ModelDefinition
from shogun.db.models.model_provider import ModelProvider
from shogun.db.models.model_router import ModelRegistryEntry
from shogun.db.models.model_routing import ModelRoutingProfile
from shogun.services.provider_credentials import provider_api_key
from shogun.services.model_reasoning import apply_chat_reasoning
from shogun.services.model_transport import model_chat_completion
from shogun.services.provider_oauth import ensure_provider_access_token
from shogun.services.structured_transformations import (
    deterministic_profile_source_units,
    expected_deterministic_matrix_rows,
    try_deterministic_matrix_transform,
)
from shogun.services.tool_calling_profiles import (
    infer_tool_calling_profile,
    normalize_native_tool_calls,
    normalize_text_tool_calls,
)

log = logging.getLogger("shogun.flow_engine")

# ── Active runs registry (for cancellation) ─────────────────
_active_runs: dict[str, asyncio.Task] = {}
_launch_events: dict[str, asyncio.Event] = {}
_run_state_locks: dict[str, asyncio.Lock] = {}
_run_mado_sessions: dict[str, set[str]] = {}
_child_run_semaphore: asyncio.Semaphore | None = None


def _run_state_lock(run_id: uuid.UUID) -> asyncio.Lock:
    """Serialize read-modify-write updates to one run's JSON state."""
    key = str(run_id)
    lock = _run_state_locks.get(key)
    if lock is None:
        lock = asyncio.Lock()
        _run_state_locks[key] = lock
    return lock


def _flow_mado_session_identity(run_id: uuid.UUID | None, session_name: Any) -> tuple[str, str]:
    """Return an execution-scoped session ID and a reusable profile name."""
    safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", str(session_name or "flow_browser").strip())[:72]
    safe_name = safe_name or "flow_browser"
    profile_name = f"flow_{safe_name}"
    if run_id is None:
        return profile_name, profile_name
    return f"flow_{run_id.hex}_{safe_name}", profile_name


async def _close_run_mado_sessions(run_id: uuid.UUID) -> None:
    """Close and forget every transient browser session owned by a flow run."""
    session_ids = _run_mado_sessions.pop(str(run_id), set())
    if not session_ids:
        return

    from shogun.services import mado_service
    from shogun.services.mado_hardening import runtime_registry

    async def close_one(session_id: str) -> None:
        try:
            result = await mado_service.close_browser(session_id)
            if result.get("status") == "error":
                log.warning("Failed to close Mado session %s for flow run %s: %s", session_id, run_id, result)
        except Exception:
            log.warning("Failed to close Mado session %s for flow run %s", session_id, run_id, exc_info=True)
        finally:
            runtime_registry.discard(session_id)

    await asyncio.gather(*(close_one(session_id) for session_id in session_ids))


@dataclass(slots=True)
class ChildFlowExecutionOptions:
    """Controls a governed child execution without introducing another engine."""

    version_mode: str = "locked"
    flow_version: int | None = None
    timeout_seconds: int | None = None
    execution_mode: str = "sequential"
    on_failure: str = "fail_parent"
    max_retries: int = 0


class ModelCallError(RuntimeError):
    """Actionable terminal error for an exhausted routed model chain."""

    def __init__(
        self,
        *,
        context: str,
        provider: str,
        model: str,
        timeout: int,
        cause: Exception,
        input_characters: int,
    ) -> None:
        timed_out = isinstance(cause, (httpx.TimeoutException, asyncio.TimeoutError))
        cause_message = str(cause).strip()
        if timed_out:
            message = f"{context} timed out after {timeout}s using {provider}/{model}"
        else:
            message = f"{context} failed using {provider}/{model}: {cause_message or type(cause).__name__}"
        super().__init__(message)
        self.provider = provider
        self.model = model
        self.timeout_seconds = timeout
        self.cause_type = type(cause).__name__
        self.input_characters = input_characters
        self.estimated_input_tokens = max(1, input_characters // 4)


class IncompleteMatrixOutputError(ValueError):
    """Raised when a structured extraction is validly shaped but visibly incomplete."""

    def __init__(
        self,
        message: str,
        *,
        candidate_rows: list[list[Any]] | None = None,
        minimum_rows: int = 0,
    ) -> None:
        super().__init__(message)
        self.candidate_rows = [list(row) for row in (candidate_rows or [])]
        self.minimum_rows = max(0, int(minimum_rows))


class MalformedMatrixOutputError(ValueError):
    """Raised when a model response cannot be decoded as the required row matrix."""


class SourceIntelligenceResolutionError(ValueError):
    """Auto profile selection failed with safe, auditable evidence."""

    def __init__(self, message: str, evidence: dict[str, Any]):
        super().__init__(message)
        self.source_intelligence = deepcopy(evidence)


_FLOW_ARTIFACT_MARKER = "__shogun_flow_artifact__"
_MAPPING_PROFILE_CARRIER_MARKER = "__shogun_mapping_profile_contract__"


def _canonical_output_for_model(output: dict[str, Any]) -> str:
    """Render canonical records once and redact declared sensitive fields for LLM context."""

    from copy import deepcopy

    canonical = deepcopy(output.get("canonical") or {})
    privacy = dict(output.get("privacy") or {})
    sensitive_paths = {
        str(path)
        for path in [
            *(privacy.get("pii_fields") or []),
            *(privacy.get("secret_fields") or []),
        ]
        if str(path).strip()
    }
    records = canonical.get("records") if isinstance(canonical, dict) else None
    if isinstance(records, list):
        for record in records:
            if not isinstance(record, dict):
                continue
            for path in sensitive_paths:
                current: Any = record
                parts = [part for part in path.split(".") if part]
                for part in parts[:-1]:
                    if not isinstance(current, dict) or part not in current:
                        current = None
                        break
                    current = current[part]
                if isinstance(current, dict) and parts and parts[-1] in current:
                    current[parts[-1]] = "[REDACTED FOR MODEL CONTEXT]"
    return json.dumps(
        {
            "canonical": canonical,
            "profile": output.get("profile") or {},
            "privacy": {
                "classification": privacy.get("classification", "internal"),
                "sensitive_fields_redacted": sorted(sensitive_paths),
                "retention": privacy.get("retention", "flow_policy"),
            },
        },
        ensure_ascii=False,
        default=str,
    )


def _flow_artifact_descriptor(
    node: AgentFlowNode | Any | None,
    output: Any,
) -> dict[str, Any]:
    """Describe an upstream result without duplicating its potentially large payload.

    AgentFlow model consumers receive the full predecessor content through the
    established context channel.  This manifest gives the model stable type and
    provenance information so it can reason about files as artifacts instead of
    guessing their role from a formatted string.
    """
    config = dict(getattr(node, "config", None) or {})
    node_type = str(getattr(node, "node_type", "") or "unknown")
    label = str(getattr(node, "label", "") or node_type)
    action = str(config.get("action") or "")
    kind = node_type
    role = "input"
    source_path = str(
        config.get("input_path")
        or config.get("file_path")
        or config.get("path")
        or ""
    )
    if isinstance(output, dict) and output.get("__shogun_file_template__"):
        kind = str(output.get("format") or "template")
        role = "template"
        source_path = str(output.get("template_path") or source_path)
    elif node_type == "office":
        kind = {
            "pdf_read": "pdf",
            "excel_read": "xlsx",
            "word_read": "docx",
            "pptx_read": "pptx",
        }.get(action, action or "office")
    elif node_type == "mapping_rpa":
        kind = "mapping"
        role = "deterministic_transform"
    content_length = len(output) if isinstance(output, str) else None
    return {
        _FLOW_ARTIFACT_MARKER: True,
        "node_id": str(getattr(node, "id", "") or ""),
        "label": label,
        "node_type": node_type,
        "kind": kind,
        "role": role,
        "source_path": source_path,
        "content_characters": content_length,
    }


def _downstream_output_contracts(
    node_id: str,
    edge_by_source: dict[str, list[tuple[str, str | None]]],
    node_map: dict[str, AgentFlowNode],
) -> list[dict[str, Any]]:
    """Return direct, non-secret output constraints visible to a Samurai node."""
    contracts: list[dict[str, Any]] = []
    for target_id, _handle in edge_by_source.get(node_id, []):
        target = node_map.get(target_id)
        if not target:
            continue
        config = dict(target.config or {})
        node_type = str(target.node_type or "")
        action = str(config.get("action") or "")
        if node_type not in {"mapping_rpa", "office", "output", "workspace"}:
            continue
        contracts.append(
            {
                "node_id": target_id,
                "label": str(target.label or node_type),
                "node_type": node_type,
                "action": action,
                "format": (
                    str((config.get("output") or {}).get("type") or "mapping")
                    if node_type == "mapping_rpa"
                    else
                    "xlsx"
                    if action in {"excel_create", "excel_write"}
                    else "docx"
                    if action in {"word_create", "word_replace"}
                    else "pptx"
                    if action in {"pptx_create", "pptx_replace"}
                    else str(config.get("format") or "")
                ),
                "sheet_name": str(config.get("sheet_name") or ""),
                "start_range": str(config.get("start_range") or config.get("data_start_cell") or ""),
                "output_path": str(config.get("output_path") or ""),
                "output_filename": str(config.get("output_filename") or ""),
            }
        )
    return contracts


def _error_message(error: Exception) -> str:
    """Return a useful message even for exceptions such as an empty ReadTimeout."""
    message = str(error).strip()
    if message:
        return message
    if isinstance(error, (httpx.TimeoutException, asyncio.TimeoutError)):
        return "Model request timed out before a response was received"
    return type(error).__name__


def _validated_node_result(result: Any) -> Any:
    """Convert legacy string failure sentinels into real node failures."""
    if isinstance(result, str):
        message = result.strip()
        if re.match(r"^\[(?:ERROR|BLOCKED)\]", message, re.IGNORECASE):
            raise RuntimeError(message)
    return result


def _template_contract_without_example(fixed_context: str) -> str:
    """Keep schema/format guidance while removing populated reference rows."""
    return str(fixed_context or "").split("[POPULATED ONE-SHOT EXAMPLE]", 1)[0].strip()


def _representative_document_sample(context: str, max_characters: int = 16_000) -> str:
    """Return bounded samples from across a document for mapping-plan inference."""
    text = str(context or "")
    if len(text) <= max_characters:
        return text
    first = max_characters // 2
    middle = max_characters // 4
    last = max_characters - first - middle
    midpoint = max(first, (len(text) // 2) - (middle // 2))
    return (
        "[SOURCE SAMPLE: BEGINNING]\n"
        + text[:first]
        + "\n\n[SOURCE SAMPLE: MIDDLE]\n"
        + text[midpoint:midpoint + middle]
        + "\n\n[SOURCE SAMPLE: END]\n"
        + text[-last:]
    )


def _samurai_checkpoint_path(config: dict[str, Any]) -> Path | None:
    flow_id = re.sub(r"[^A-Za-z0-9_-]", "", str(config.get("_flow_id") or ""))
    node_id = re.sub(r"[^A-Za-z0-9_-]", "", str(config.get("_node_id") or ""))
    if not flow_id or not node_id:
        return None
    from shogun.config import settings

    return settings.workspace_path.resolve() / ".shogun" / "agentflow-checkpoints" / flow_id / f"{node_id}.json"


def _load_samurai_checkpoint(config: dict[str, Any], fingerprint: str) -> dict[str, Any]:
    target = _samurai_checkpoint_path(config)
    if not target or not target.is_file():
        return {"version": 1, "fingerprint": fingerprint, "outputs": {}}
    try:
        payload = json.loads(target.read_text(encoding="utf-8"))
        if payload.get("version") == 1 and payload.get("fingerprint") == fingerprint:
            payload["outputs"] = dict(payload.get("outputs") or {})
            return payload
    except Exception as exc:
        log.warning("Ignoring unreadable Samurai checkpoint %s: %s", target, exc)
    return {"version": 1, "fingerprint": fingerprint, "outputs": {}}


def _save_samurai_checkpoint(config: dict[str, Any], payload: dict[str, Any]) -> None:
    target = _samurai_checkpoint_path(config)
    if not target:
        return
    try:
        target.parent.mkdir(parents=True, exist_ok=True)
        # Concurrent scheduled runs can share a flow/node fingerprint. Merge
        # completed batches instead of letting the last writer discard work
        # checkpointed by another run.
        if target.is_file():
            try:
                current = json.loads(target.read_text(encoding="utf-8"))
                if current.get("fingerprint") == payload.get("fingerprint"):
                    merged_outputs = dict(current.get("outputs") or {})
                    merged_outputs.update(payload.get("outputs") or {})
                    payload = {**current, **payload, "outputs": merged_outputs}
            except Exception:
                pass
        run_suffix = re.sub(r"[^A-Za-z0-9_-]", "", str(config.get("_run_id") or "runtime"))
        temporary = target.with_name(f"{target.name}.{run_suffix}.tmp")
        temporary.write_text(
            json.dumps(payload, ensure_ascii=False, default=str),
            encoding="utf-8",
        )
        temporary.replace(target)
    except Exception as exc:
        log.warning("Could not persist Samurai checkpoint %s: %s", target, exc)


def _node_failure_action(config: dict[str, Any]) -> str:
    """Normalize legacy and current node failure policies."""
    return config.get("failure_action") or {
        "fail_parent": "stop",
        "continue_with_error": "continue",
        "route_to_error": "continue",
    }.get(config.get("on_failure", "fail_parent"), "stop")


def _failure_action_is_terminal(action: str) -> bool:
    """Only explicit skip/continue policies may proceed after a node error."""
    return action not in {"skip", "continue"}


def _parallel_child_semaphore() -> asyncio.Semaphore:
    global _child_run_semaphore
    if _child_run_semaphore is None:
        from shogun.config import settings

        _child_run_semaphore = asyncio.Semaphore(settings.flow_stacking_max_parallel_children)
    return _child_run_semaphore


# ═══════════════════════════════════════════════════════════════
# PUBLIC API
# ═══════════════════════════════════════════════════════════════


async def start_flow_run(
    flow_id: uuid.UUID,
    trigger_type: str = "manual",
    input_payload: dict[str, Any] | None = None,
    governance_context: dict[str, Any] | None = None,
) -> uuid.UUID:
    """Create a FlowRun record and launch execution as a background task.

    Returns the run ID immediately. Execution proceeds asynchronously.
    """
    run_id = uuid.uuid4()

    async with async_session_factory() as session:
        # Verify flow exists and is not deleted
        result = await session.execute(
            select(AgentFlow).where(
                AgentFlow.id == flow_id,
                AgentFlow.is_deleted.is_(False),
                AgentFlow.flow_type == "standard",
            )
        )
        flow = result.scalar_one_or_none()
        if not flow:
            raise ValueError(f"Agent Flow {flow_id} not found or deleted")

        effective_governance = await _root_governance_context(governance_context or {})
        _apply_flow_generation_settings(effective_governance, flow)
        payload = _json_object(input_payload or {}, "Flow input")
        run = AgentFlowRun(
            id=run_id,
            flow_id=flow_id,
            flow_version=flow.version,
            root_run_id=run_id,
            parent_run_id=None,
            parent_node_id=None,
            run_depth=0,
            status="pending",
            trigger_type=trigger_type,
            node_states={},
            result_summary={},
            input_payload=payload,
            output_payload={},
            artifacts=[],
            governance_context=effective_governance,
        )
        session.add(run)
        await session.commit()

    # Launch as background task
    launch_ready = asyncio.Event()
    _launch_events[str(run_id)] = launch_ready
    task = asyncio.create_task(_execute_flow(run_id, flow_id))
    _active_runs[str(run_id)] = task

    # Do not hand the run ID to an orchestrator until the executor has loaded
    # the committed row. This prevents concurrent SQLite sessions from
    # colliding during the launch hand-off.
    try:
        await asyncio.wait_for(launch_ready.wait(), timeout=0.25)
    except TimeoutError:
        log.warning("Flow run %s executor launch acknowledgement timed out", run_id)

    # Auto-cleanup when done
    def _cleanup(t: asyncio.Task):
        _active_runs.pop(str(run_id), None)
        _launch_events.pop(str(run_id), None)
        _run_state_locks.pop(str(run_id), None)

    task.add_done_callback(_cleanup)

    log.info("Flow run %s started for flow %s (trigger=%s)", run_id, flow_id, trigger_type)
    return run_id


async def cancel_flow_run(run_id: uuid.UUID) -> bool:
    """Cancel a running flow execution."""
    task = _active_runs.get(str(run_id))
    if task and not task.done():
        task.cancel()
        try:
            await task
        except asyncio.CancelledError:
            pass
        async with async_session_factory() as session:
            requested = await session.get(AgentFlowRun, run_id)
            if not requested:
                return False
            root_id = requested.root_run_id or requested.id
            result = await session.execute(
                select(AgentFlowRun).where(
                    AgentFlowRun.root_run_id == root_id,
                    AgentFlowRun.status.in_(["pending", "running", "waiting_for_approval"]),
                )
            )
            now = datetime.now(timezone.utc)
            cancelled_children: list[asyncio.Task] = []
            for run in result.scalars().all():
                child_task = _active_runs.get(str(run.id))
                if child_task and not child_task.done():
                    child_task.cancel()
                    if child_task is not task:
                        cancelled_children.append(child_task)
                run.status = "cancelled"
                run.completed_at = now
                run.error_message = "Cancelled by user or ancestor run"
            await session.commit()
        if cancelled_children:
            await asyncio.gather(*cancelled_children, return_exceptions=True)
        await _sync_run_edge_status(run_id, "cancelled")
        log.info("Flow run %s cancelled", run_id)
        return True
    return False


def request_cancel_all_flow_runs(*, exclude_current: bool = True) -> list[asyncio.Task]:
    """Synchronously signal cancellation to every in-process flow task."""
    current = asyncio.current_task() if exclude_current else None
    tasks = [
        task
        for task in tuple(_active_runs.values())
        if task is not current and not task.done()
    ]
    for task in tasks:
        task.cancel()
    return tasks


async def cancel_all_flow_runs(reason: str = "HARAKIRI activated") -> int:
    """Cancel every in-process flow, including child flows, and await exit."""
    tasks = request_cancel_all_flow_runs()
    if tasks:
        await asyncio.gather(*tasks, return_exceptions=True)
    log.critical("Cancelled %d active flow task(s): %s", len(tasks), reason)
    return len(tasks)


# ═══════════════════════════════════════════════════════════════
# CORE EXECUTION LOOP
# ═══════════════════════════════════════════════════════════════


async def _execute_flow(run_id: uuid.UUID, flow_id: uuid.UUID) -> None:
    """Main execution loop — loads flow, walks DAG, executes nodes."""
    launch_ready = _launch_events.get(str(run_id))
    try:
        async with async_session_factory() as session:
            # ── 1. Load flow with nodes and edges ──────────────────
            result = await session.execute(
                select(AgentFlow)
                .where(AgentFlow.id == flow_id)
                .options(
                    selectinload(AgentFlow.nodes),
                    selectinload(AgentFlow.edges),
                )
            )
            flow = result.scalar_one_or_none()
            if not flow:
                await _fail_run(run_id, "Flow not found")
                return

            nodes = list(flow.nodes)
            edges = list(flow.edges)

            if not nodes:
                await _fail_run(run_id, "Flow has no nodes")
                return

            # ── 2. Mark run as running ─────────────────────────────
            # The run is committed before this background task is launched. A
            # short visibility retry keeps SQLite's single-connection test and
            # desktop modes from treating a concurrent session hand-off as a
            # permanently missing run.
            run = None
            for _ in range(20):
                run = await session.get(AgentFlowRun, run_id, populate_existing=True)
                if run is not None:
                    break
                await asyncio.sleep(0.01)
            if run is None:
                raise LookupError(f"Flow run {run_id} was not visible after launch")
            run_input = dict(run.input_payload or {})
            governance_context = dict(run.governance_context or {})
            run_trigger_type = str(run.trigger_type or "manual")
            run.status = "running"
            run.started_at = datetime.now(timezone.utc)

            # Initialize node states
            node_states: dict[str, dict[str, Any]] = {}
            for node in nodes:
                node_states[str(node.id)] = {
                    "status": "pending",
                    "output": None,
                    "error": None,
                    "started_at": None,
                    "completed_at": None,
                }
            run.node_states = node_states
            await session.commit()
            if launch_ready:
                launch_ready.set()

        # ── 3. Topological sort ────────────────────────────────
        try:
            execution_layers = _topological_sort(nodes, edges)
        except ValueError as e:
            await _fail_run(run_id, str(e))
            return

        # ── 4. Build lookup maps ───────────────────────────────
        node_map: dict[str, AgentFlowNode] = {str(n.id): n for n in nodes}
        edge_list = edges

        # Build predecessor map: node_id → [source_node_ids]
        predecessors: dict[str, list[str]] = defaultdict(list)
        for edge in edge_list:
            predecessors[str(edge.target_node_id)].append(str(edge.source_node_id))

        # Build edge map for logic nodes: source_node_id → [(target_node_id, source_handle)]
        edge_by_source: dict[str, list[tuple[str, str | None]]] = defaultdict(list)
        for edge in edge_list:
            edge_by_source[str(edge.source_node_id)].append((str(edge.target_node_id), edge.source_handle))

        # ── 5. Walk layers ─────────────────────────────────────
        # node_outputs stores the output of each completed node
        node_outputs: dict[str, Any] = {}
        # skipped_nodes tracks nodes that should be skipped (logic branch pruning)
        skipped_nodes: set[str] = set()

        for layer in execution_layers:
            # Filter out skipped nodes
            active_nodes = [nid for nid in layer if nid not in skipped_nodes]

            if not active_nodes:
                continue

            # Execute all nodes in this layer in parallel
            tasks = []
            for node_id in active_nodes:
                node = node_map[node_id]
                # Gather predecessor outputs as context
                pred_outputs = {}
                for pred_id in predecessors.get(node_id, []):
                    if pred_id in node_outputs:
                        pred_outputs[pred_id] = node_outputs[pred_id]
                template_inputs = _collect_upstream_file_templates(
                    node_id,
                    predecessors,
                    node_outputs,
                )

                tasks.append(
                    _execute_single_node(
                        run_id=run_id,
                        node=node,
                        predecessor_outputs=pred_outputs,
                        node_map=node_map,
                        run_input=run_input,
                        governance_context=governance_context,
                        flow_name=flow.name,
                        trigger_type=run_trigger_type,
                        template_inputs=template_inputs,
                        downstream_contracts=_downstream_output_contracts(
                            node_id,
                            edge_by_source,
                            node_map,
                        ),
                    )
                )

            results = await asyncio.gather(*tasks, return_exceptions=True)

            # Process results
            for node_id, result in zip(active_nodes, results):
                node = node_map[node_id]
                if isinstance(result, Exception):
                    node_outputs[node_id] = None
                    failure_event_id = await _record_node_failure_event(run_id, node, result)
                    error_message = _error_message(result)
                    await _update_node_state(
                        run_id,
                        node_id,
                        "failed",
                        error=error_message,
                        failure_event_id=failure_event_id,
                    )
                    # Check failure action
                    config = node.config or {}
                    failure_action = _node_failure_action(config)
                    if _failure_action_is_terminal(failure_action):
                        # Retries are exhausted inside the node executor. A terminal
                        # retry/escalate (or unknown) action must fail the run rather
                        # than silently falling through to downstream nodes.
                        await _fail_run(
                            run_id,
                            f"Node '{node.label}' failed: {error_message}",
                            node_states_override=True,
                        )
                        return
                    elif failure_action == "skip":
                        # Mark downstream nodes as skipped
                        _mark_downstream_skipped(node_id, edge_by_source, skipped_nodes)
                    elif failure_action == "continue":
                        node_outputs[node_id] = {
                            "status": "failed",
                            "output": {},
                            "artifacts": [],
                            "errors": [error_message],
                        }
                else:
                    node_outputs[node_id] = result
                    await _update_node_state(run_id, node_id, "completed", output=result)

                    # ── Logic/Decision branch pruning ──────────
                    if node.node_type == "logic":
                        # result is True/False — prune the non-taken branch
                        taken_handle = None if result else "false"
                        for target_id, handle in edge_by_source.get(node_id, []):
                            if handle != taken_handle:
                                # This branch was NOT taken — skip all downstream
                                skipped_nodes.add(target_id)
                                _mark_downstream_skipped(target_id, edge_by_source, skipped_nodes)
                            else:
                                # Ensure the taken branch is NOT skipped
                                skipped_nodes.discard(target_id)
                    elif node.node_type == "mapping_rpa" and isinstance(result, dict):
                        # Mapping nodes expose deterministic route handles while
                        # preserving unlabelled edges as the normal success path.
                        status = str(result.get("status") or "SUCCESS").lower()
                        for target_id, handle in edge_by_source.get(node_id, []):
                            normalized = str(handle or "").lower()
                            if normalized and normalized != status:
                                skipped_nodes.add(target_id)
                                _mark_downstream_skipped(target_id, edge_by_source, skipped_nodes)
                            else:
                                skipped_nodes.discard(target_id)

        # ── 6. Mark skipped nodes ──────────────────────────────
        for nid in skipped_nodes:
            await _update_node_state(run_id, nid, "skipped")

        # ── 7. Build result summary ────────────────────────────
        # Collect output node results
        output_results = {}
        for node in nodes:
            if node.node_type == "output" and str(node.id) in node_outputs:
                output_results[node.label] = node_outputs[str(node.id)]

        await _complete_run(run_id, output_results or node_outputs)

    except asyncio.CancelledError:
        log.info("Flow run %s was cancelled", run_id)
        await _cancel_run_record(run_id, "Cancelled by user or ancestor run")
        raise
    except Exception as exc:
        log.exception("Flow run %s failed with unexpected error", run_id)
        await _fail_run(run_id, f"Unexpected error: {str(exc)[:500]}")
    finally:
        await _close_run_mado_sessions(run_id)
        if launch_ready and not launch_ready.is_set():
            launch_ready.set()


# ═══════════════════════════════════════════════════════════════
# NODE EXECUTORS
# ═══════════════════════════════════════════════════════════════


async def _execute_single_node(
    run_id: uuid.UUID,
    node: AgentFlowNode,
    predecessor_outputs: dict[str, Any],
    node_map: dict[str, AgentFlowNode],
    run_input: dict[str, Any] | None = None,
    governance_context: dict[str, Any] | None = None,
    flow_name: str = "Agent Flow",
    trigger_type: str = "manual",
    template_inputs: list[dict[str, Any]] | None = None,
    downstream_contracts: list[dict[str, Any]] | None = None,
) -> Any:
    """Execute a single node and return its output."""
    node_id = str(node.id)
    await _update_node_state(run_id, node_id, "running")

    config = dict(node.config or {})
    node_type = node.node_type

    samurai_transformation_mode = "general"
    samurai_transformation_profile: Any | None = None
    samurai_transformation_candidates: list[Any] = []
    if node_type == "samurai":
        from shogun.schemas.agent_flow import (
            SamuraiTransformationConfig,
            normalize_samurai_transformation_config,
        )

        config = normalize_samurai_transformation_config(config)
        samurai_transformation = SamuraiTransformationConfig.model_validate(config)
        samurai_transformation_mode = samurai_transformation.transformation_mode
        samurai_transformation_profile = samurai_transformation.transformation_profile
        samurai_transformation_candidates = list(
            samurai_transformation.transformation_candidates
        )

    # Transformation profiles may arrive through the legacy upstream
    # Mapping/RPA contract or through an explicit Samurai profile-selection
    # mode. Auto mode delegates bounded inspection to Source Intelligence;
    # prompts and filenames alone never become execution authority.
    upstream_transformation_profiles: list[dict[str, Any]] = []
    transformation_profile_evidence: list[dict[str, Any]] = []
    transformation_profile_carrier_ids: set[str] = set()
    if node_type == "samurai":
        from shogun.mapping.schema import MappingConfig

        for predecessor_id in predecessor_outputs:
            predecessor = node_map.get(predecessor_id)
            if predecessor is None or getattr(predecessor, "node_type", None) != "mapping_rpa":
                continue
            predecessor_config = dict(getattr(predecessor, "config", None) or {})
            if predecessor_config.get("transformation_profile") is None:
                continue
            mapping_config = MappingConfig.model_validate(predecessor_config)
            if mapping_config.execution_mode == "transform":
                # A normal Mapping/RPA transform may retain profile metadata in
                # an old template or imported configuration.  Its output is
                # ordinary source data for downstream nodes; only the explicit
                # contract mode is allowed to attach a Samurai execution
                # contract.  Never reinterpret transform-mode metadata as
                # trusted profile authority.
                continue
            profile = mapping_config.transformation_profile
            if profile is None:
                continue
            predecessor_output = predecessor_outputs.get(predecessor_id)
            if mapping_config.execution_mode == "contract":
                resolved_profile = _trusted_contract_profile_from_carrier(
                    profile,
                    predecessor_output,
                    carrier_label=str(getattr(predecessor, "label", predecessor_id)),
                )
                transformation_profile_carrier_ids.add(predecessor_id)
                upstream_transformation_profiles.append(resolved_profile)
                transformation_profile_evidence.append(
                    deepcopy(predecessor_output.get("registry_evidence") or {})
                )
                continue
            elif mapping_config.execution_mode == "profile":
                # Structured enterprise ingress profiles execute inside their
                # Mapping/RPA node. Downstream nodes consume canonical data;
                # they must not reinterpret the same profile as a PDF/Samurai
                # extraction contract.
                if not (
                    isinstance(predecessor_output, dict)
                    and predecessor_output.get("__shogun_canonical_output__")
                    and str(predecessor_output.get("status") or "").upper() in {"SUCCESS", "PARTIAL"}
                ):
                    raise ValueError(
                        f"Mapping/RPA enterprise profile '{getattr(predecessor, 'label', predecessor_id)}' "
                        "did not provide successful canonical output."
                    )
                continue
            elif isinstance(predecessor_output, dict) and str(
                predecessor_output.get("status") or "SUCCESS"
            ).upper() not in {"SUCCESS", "PARTIAL"}:
                raise ValueError(
                    f"Mapping/RPA profile predecessor '{getattr(predecessor, 'label', predecessor_id)}' "
                    "did not complete successfully."
                )
            upstream_transformation_profiles.append(profile.model_dump(mode="json"))

        # Normalize and reject ambiguous contracts before any source context is
        # assembled. A normal Samurai has no profile; a deterministic Samurai
        # has exactly one explicitly connected Mapping/RPA profile.
        active_profile = _active_transformation_profile(
            {"_transformation_profiles": upstream_transformation_profiles}
        )
        upstream_transformation_profiles = [active_profile] if active_profile else []

        if samurai_transformation_mode != "general" and upstream_transformation_profiles:
            raise ValueError(
                "Samurai direct/auto transformation selection conflicts with an upstream "
                "Mapping/RPA transformation-profile contract. Use exactly one selection path."
            )
        if samurai_transformation_mode == "profile":
            if samurai_transformation_profile is None:  # Schema validation is intentionally defensive.
                raise ValueError("Samurai profile mode requires a transformation_profile.")
            resolved_profile, resolved_evidence = await _resolve_direct_samurai_profile(
                samurai_transformation_profile
            )
            upstream_transformation_profiles = [resolved_profile]
            transformation_profile_evidence = [resolved_evidence]

    # Build context string from predecessor outputs
    context_parts: list[str] = []
    chunkable_context_parts: list[str] = []
    fixed_context_parts: list[str] = []
    input_artifacts: list[dict[str, Any]] = []
    transformation_source_contexts: list[dict[str, str]] = []
    transformation_source_inputs: list[dict[str, Any]] = []
    for pred_id, output in predecessor_outputs.items():
        pred_node = node_map.get(pred_id)
        pred_label = pred_node.label if pred_node else pred_id
        if pred_id in transformation_profile_carrier_ids:
            # The carrier is control-plane configuration, not source data. Its
            # tiny success marker must never be parsed as another PDF/record or
            # advertised to the model as a business artifact.
            continue
        if output is not None:
            artifact_descriptor = _flow_artifact_descriptor(pred_node, output)
            input_artifacts.append(artifact_descriptor)
            if isinstance(output, dict) and output.get("__shogun_file_template__"):
                from shogun.services.file_template import format_template_guidance

                output_text = format_template_guidance(output)
                is_fixed_context = True
            elif isinstance(output, dict) and output.get("__shogun_canonical_output__"):
                output_text = _canonical_output_for_model(output)
                is_fixed_context = False
            else:
                output_text = str(output)
                is_fixed_context = False
            # Data-processing and delivery nodes must receive complete results.
            # Samurai handles large inputs with model-aware chunking below.
            # Keep the legacy guard only for executors that do not yet support
            # chunking and could otherwise overrun a single model request.
            if node_type in {"coding", "mado_browser"}:
                output_text = _truncate(output_text, 4000)
            if is_fixed_context:
                labelled_output = f"[Reference-only template from '{pred_label}']:\n{output_text}"
            else:
                labelled_output = f"[Output from '{pred_label}']:\n{output_text}"
            context_parts.append(labelled_output)
            (fixed_context_parts if is_fixed_context else chunkable_context_parts).append(labelled_output)
            if node_type == "samurai" and not is_fixed_context:
                transformation_source_contexts.append(
                    {
                        "node_id": str(pred_id),
                        "label": str(pred_label),
                        "content": output_text,
                    }
                )
                predecessor_config = dict(getattr(pred_node, "config", None) or {})
                action = str(predecessor_config.get("action") or "").lower()
                inferred_transport = (
                    predecessor_config.get("source_transport")
                    or predecessor_config.get("transport")
                    or (
                        predecessor_config.get("input_type")
                        if predecessor_config.get("input_type") in {"api", "event", "nexus", "subflow"}
                        else None
                    )
                    or {
                        "pdf_read": "pdf",
                        "excel_read": "excel",
                        "word_read": "word",
                        "pptx_read": "powerpoint",
                    }.get(action)
                )
                source_path = str(artifact_descriptor.get("source_path") or "")
                transformation_source_inputs.append(
                    {
                        "source_id": str(pred_id),
                        "label": str(pred_label),
                        "raw_output": output,
                        "text_output": output_text,
                        "context": {
                            "transport": inferred_transport,
                            "object": predecessor_config.get("source_object")
                            or predecessor_config.get("object"),
                            "record_shape": predecessor_config.get("record_shape"),
                            "record_path": predecessor_config.get("record_path"),
                            "content_type": predecessor_config.get("content_type")
                            or artifact_descriptor.get("kind"),
                            "file_name": Path(source_path).name if source_path else None,
                            "connector": predecessor_config.get("connector"),
                            "platform_hint": predecessor_config.get("platform_hint"),
                        },
                    }
                )
    context_str = "\n\n".join(context_parts) if context_parts else ""
    chunkable_context_str = "\n\n".join(chunkable_context_parts)
    fixed_context_str = "\n\n".join(fixed_context_parts)

    # Additional context injection from config
    if config.get("context_injection"):
        injected = f"[Additional Context]:\n{config['context_injection']}"
        context_str = f"{context_str}\n\n{injected}" if context_str else injected
        fixed_context_str = f"{fixed_context_str}\n\n{injected}" if fixed_context_str else injected

    # A Samurai instruction attachment is the node's prompt, not extra
    # context. Resolve it before skill selection so routing, governed native
    # reads, and the model all see the same authoritative instruction.
    if node_type == "samurai" and config.get("instruction_file"):
        config = {
            **config,
            "task_description": await _resolve_samurai_task_description(config),
            "_instruction_file_resolved": True,
        }

    if node_type == "samurai" and samurai_transformation_mode == "auto":
        resolved_profile, resolved_evidence = await _resolve_auto_samurai_profile(
            source_inputs=transformation_source_inputs,
            private_profiles=samurai_transformation_candidates,
            config={
                **config,
                "_flow_id": str(node.flow_id),
                "_node_id": node_id,
                "_run_id": str(run_id),
            },
            governance_context=governance_context or {},
        )
        upstream_transformation_profiles = [resolved_profile]
        transformation_profile_evidence = [resolved_evidence]

    if node_type == "samurai" and upstream_transformation_profiles:
        transformation_profile_evidence = [
            _require_runtime_transformation_profile_evidence(
                upstream_transformation_profiles[0],
                transformation_profile_evidence,
            )
        ]

    # Order 9: node-level skill activation. Compact briefs influence model
    # execution but never extend the node's tool or posture permissions. Keep
    # private instructions separate from predecessor content so deterministic
    # delivery nodes cannot publish them.
    execution_context_str = context_str
    active_skill_run_ids: list[str] = []
    # An explicit deterministic profile returns before model routing. Avoid
    # skill retrieval/activation as well so this path is genuinely model-free
    # and its template contract cannot be polluted by unrelated skill text.
    deterministic_samurai = node_type == "samurai" and bool(upstream_transformation_profiles)
    if _node_uses_active_skill_context(node_type, config) and not deterministic_samurai:
        try:
            from shogun.schemas.skills import SkillActivationRequest
            from shogun.services.active_skill_service import SkillActivationService
            from shogun.services.posture_guard import get_posture_permissions

            posture = governance_context or await get_posture_permissions()
            node_objective = (
                config.get("task_description")
                or config.get("prompt")
                or config.get("description")
                or node_type
            )
            async with async_session_factory() as skill_session:
                activation = await SkillActivationService(skill_session).activate(
                    SkillActivationRequest(
                        run_id=str(run_id),
                        flow_id=str(node.flow_id),
                        node_id=node_id,
                        agent_id="shogun",
                        objective=f"{node.label}: {node_objective}",
                        context=context_str[-4000:],
                        posture=posture.get("active_tier", posture.get("posture", "guarded")),
                        available_tools=list(
                            set((config.get("required_tools") or []) + (posture.get("allowed_tools") or []))
                        ),
                        max_skills=3,
                        usage_location="agent_flow",
                        ide_enabled=bool(posture.get("ide_enabled", False)),
                    )
                )
                await skill_session.commit()
            active_skill_run_ids = [str(item["active_skill_run_id"]) for item in activation["active_skills"]]
            if activation["context_block"]:
                execution_context_str += f"\n\n{activation['context_block']}"
                fixed_context_str += f"\n\n{activation['context_block']}"
        except Exception as exc:
            logging.getLogger("shogun.flow").warning("Active skill selection skipped: %s", exc)

    try:
        if node_type == "input":
            result = await _exec_input(config, context_str, run_input or {})
        elif node_type == "file_template":
            result = await _exec_file_template(config)
        elif node_type == "samurai":
            samurai_runtime_config = {
                **config,
                "_flow_id": str(node.flow_id),
                "_node_id": node_id,
                "_run_id": str(run_id),
                "_input_artifacts": input_artifacts,
                "_output_contracts": list(downstream_contracts or []),
                "_transformation_profiles": upstream_transformation_profiles,
                "_transformation_profile_evidence": transformation_profile_evidence,
                "_transformation_source_contexts": transformation_source_contexts,
            }
            active_samurai_profile = (
                upstream_transformation_profiles[0]
                if len(upstream_transformation_profiles) == 1
                else None
            )
            if (
                active_samurai_profile
                and active_samurai_profile.get("adapter") == "canonical_entity_map_v1"
            ):
                result = await _exec_samurai_enterprise_profile(
                    samurai_runtime_config,
                    active_samurai_profile,
                    transformation_profile_evidence[0],
                    transformation_source_inputs,
                )
            else:
                result = await _exec_samurai(
                    samurai_runtime_config,
                    chunkable_context_str,
                    governance_context or {},
                    fixed_context_str=fixed_context_str,
                    progress_callback=lambda completed, total: _update_node_progress(
                        run_id,
                        node_id,
                        completed,
                        total,
                    ),
                )
        elif node_type == "email_read":
            result = await _exec_email_read(config)
        elif node_type == "calendar_read":
            result = await _exec_calendar_read(config)
        elif node_type == "coding":
            result = await _exec_coding(config, execution_context_str, governance_context or {})
        elif node_type == "shogun_approval":
            result = await _exec_approval(config, predecessor_outputs, governance_context or {})
        elif node_type == "logic":
            result = await _exec_logic(config, predecessor_outputs, governance_context or {})
        elif node_type == "output":
            result = await _exec_output(
                config,
                context_str,
                predecessor_outputs,
                run_id,
                node.label,
                node_id,
            )
            memory_config = config.get("memory_infusion") or {}
            if memory_config.get("enabled"):
                from shogun.services.flow_memory_infusion import infuse_flow_output_memory

                async with async_session_factory() as memory_session:
                    await infuse_flow_output_memory(
                        session=memory_session,
                        raw_config=memory_config,
                        flow_id=node.flow_id,
                        flow_name=flow_name,
                        run_id=run_id,
                        node_id=node.id,
                        node_label=node.label,
                        output=result,
                        predecessor_outputs=predecessor_outputs,
                    )
                    await memory_session.commit()
        elif node_type == "mado_browser":
            result = await _exec_mado_browser(config, context_str, run_id, node_id, governance_context or {})
        elif node_type == "email_send":
            result = await _exec_email_send(config, context_str)
        elif node_type == "channel_send":
            result = await _exec_channel_send(config, context_str)
        elif node_type == "workspace":
            result = await _exec_workspace(config, context_str, run_id, trigger_type)
        elif node_type == "mapping_rpa":
            result = await _exec_mapping_rpa(
                config,
                predecessor_outputs,
                flow_id=str(node.flow_id),
                node_id=node_id,
            )
        elif node_type == "office":
            result = await _exec_office(
                config,
                context_str,
                run_id,
                node_id,
                trigger_type,
                template_inputs=template_inputs,
                predecessor_outputs=predecessor_outputs,
            )
        elif node_type == "subflow":
            result = await _exec_subflow(
                run_id,
                node,
                predecessor_outputs,
                run_input or {},
                governance_context or {},
            )
        elif node_type == "stack_orchestrator":
            raise ValueError("Flow Stacking is not available in Yellow Label.")
        else:
            raise ValueError(f"Unknown node type: {node_type}")
        result = _validated_node_result(result)
    except Exception as exc:
        await _finalize_node_skills(active_skill_run_ids, "failed", str(exc))
        raise
    await _finalize_node_skills(active_skill_run_ids, "success", f"Node '{node.label}' completed")
    return result


def _node_uses_active_skill_context(node_type: str, config: dict[str, Any]) -> bool:
    """Return whether a node actually sends its context to an LLM."""
    return node_type == "samurai" or (node_type == "coding" and config.get("action", "analyze") == "analyze")


async def _finalize_node_skills(active_skill_run_ids: list[str], outcome: str, summary: str) -> None:
    if not active_skill_run_ids:
        return
    try:
        from shogun.services.active_skill_service import SkillActivationService
        from shogun.services.skill_trajectory_service import SkillTrajectoryService

        async with async_session_factory() as session:
            await SkillTrajectoryService(session).link_output(
                active_skill_run_ids,
                output_summary=summary,
                output_type="agent_flow_node",
                metadata={"outcome": outcome},
            )
            service = SkillActivationService(session)
            for active_id in active_skill_run_ids:
                await service.outcome(uuid.UUID(active_id), outcome, summary)
            await session.commit()
    except Exception as exc:
        logging.getLogger("shogun.flow").warning("Active skill outcome capture skipped: %s", exc)


async def _exec_stack_orchestrator(
    parent_flow_run_id: uuid.UUID,
    config: dict[str, Any],
    run_input: dict[str, Any],
) -> dict[str, Any]:
    """Start the backend control service represented by a Katana control node."""
    from shogun.config import settings
    from shogun.db.models.stack_orchestrator import StackRun
    from shogun.schemas.stack_orchestrator import StackOrchestratorCreate
    from shogun.services.stack_orchestrator import StackOrchestratorService

    selected_stack_id = config.get("selected_stack_id")
    async with async_session_factory() as session:
        parent = await session.get(AgentFlowRun, parent_flow_run_id)
        if parent and selected_stack_id and str(parent.flow_id) == str(selected_stack_id):
            raise ValueError("Stack Orchestrator cannot select the Agent Flow that contains itself.")
        body = StackOrchestratorCreate(
            mode=config.get("mode", "selected_stack"),
            stack_template_id=config.get("stack_template_id"),
            selected_stack_id=selected_stack_id,
            objective=config.get("objective") or str(run_input.get("objective") or "Execute Agent Stack"),
            success_criteria=config.get("success_criteria") or [],
            allowed_tools=config.get("allowed_tools") or [],
            model_routing_profile=config.get("model_routing_profile", "balanced"),
            max_runtime_minutes=int(config.get("max_runtime_minutes", 60)),
            max_iterations=int(config.get("max_iterations", 50)),
            max_retry_attempts_per_step=int(config.get("max_retry_attempts_per_step", 2)),
            checkpoint_frequency=config.get("checkpoint_frequency", "after_each_step"),
            context_compaction=(
                "enabled" if config.get("context_compaction", True) in {True, "enabled"} else "disabled"
            ),
            verification_required=bool(config.get("verification_required", True)),
            approval_policy=config.get("approval_policy", "inherited"),
            artifact_policy=config.get("artifact_policy", "retain_all"),
            output_publication=config.get("output_publication", "summary_and_final"),
            failure_policy=config.get("failure_policy", "pause"),
            input_payload=run_input,
        )
        service = StackOrchestratorService(session)
        stack = await service.create(body)
        stack_run_id = stack.id
        if stack.status == "created":
            await service.start(stack.id)

    while True:
        await asyncio.sleep(settings.stack_orchestrator_poll_interval_seconds)
        async with async_session_factory() as session:
            stack = await session.get(StackRun, stack_run_id)
            if not stack:
                raise ValueError("Stack Orchestrator run record disappeared.")
            if stack.status not in {"created", "running"}:
                return {
                    "stack_run_id": str(stack.id),
                    "status": stack.status,
                    "objective": stack.objective,
                    "current_step_id": stack.current_step_id,
                    "completed_steps": stack.completed_steps,
                    "failed_steps": stack.failed_steps,
                    "final_summary": stack.final_summary,
                    "published_output": stack.published_output,
                    "requires_review": stack.status == "waiting_approval",
                }


async def _exec_input(config: dict, context_str: str, run_input: dict[str, Any] | None = None) -> Any:
    """Input node — returns its configuration as initial context.

    Handles multiple input types:
    - manual: uses manual_input text or description
    - document: reads uploaded file content from disk
    - scheduled/api/event/nexus: uses description as context
    """
    import logging
    from pathlib import Path

    log = logging.getLogger("shogun.flow")
    description = config.get("description", "")
    input_type = config.get("input_type", "manual")

    output_parts = []

    if run_input:
        # API, stack, and parent-flow payloads remain structured for downstream
        # Subflow mapping. Existing text-only flows retain their prior behavior.
        if not description and input_type in {"api", "event", "nexus", "subflow"}:
            return run_input
        output_parts.append(json.dumps(run_input, ensure_ascii=False, default=str))

    # Always include description if present
    if description:
        output_parts.append(description)

    # Type-specific context
    if input_type == "manual":
        manual_input = config.get("manual_input", "")
        if manual_input:
            output_parts.append(manual_input)

    elif input_type == "document":
        uploaded = config.get("uploaded_file")
        workspace_path = str(config.get("workspace_path") or "").strip()
        attachment_file_id = str(config.get("attachment_file_id") or config.get("file_id") or "").strip()
        source = str(config.get("document_source") or "").strip().lower()
        if not source:
            source = (
                "upload"
                if uploaded and uploaded.get("path")
                else "workspace"
                if workspace_path
                else "attachment"
                if attachment_file_id
                else "upload"
            )

        try:
            from shogun.config import settings
            from shogun.services.file_formats import FileFormatService

            if source == "workspace":
                if not workspace_path:
                    raise ValueError("No workspace file was selected")
                workspace_root = settings.workspace_path.resolve()
                requested = Path(workspace_path)
                file_path = requested.resolve() if requested.is_absolute() else (workspace_root / requested).resolve()
                try:
                    file_path.relative_to(workspace_root)
                except ValueError as exc:
                    raise ValueError("Workspace document must remain inside the configured workspace") from exc
                if not file_path.is_file():
                    raise FileNotFoundError(f"Workspace document not found: {workspace_path}")
                payload = await FileFormatService(allowed_roots=[workspace_root]).read(
                    path=str(file_path),
                    max_chars=settings.agent_flow_document_max_chars,
                )
            elif source == "attachment":
                if not attachment_file_id:
                    raise ValueError("No chat attachment reference was bound to this node")
                try:
                    file_id = uuid.UUID(attachment_file_id)
                except ValueError as exc:
                    raise ValueError("The chat attachment reference is invalid") from exc
                async with async_session_factory() as session:
                    payload = await FileFormatService(session).read(
                        file_id=file_id,
                        max_chars=settings.agent_flow_document_max_chars,
                    )
            elif source == "upload":
                if not uploaded or not uploaded.get("path"):
                    if uploaded and uploaded.get("filename"):
                        raise ValueError(
                            f"The upload for '{uploaded['filename']}' did not complete successfully "
                            "(server path is missing). Please remove and re-upload the document."
                        )
                    raise ValueError("No document was uploaded to this node")
                file_path = Path(uploaded["path"])
                if not file_path.is_file():
                    raise FileNotFoundError(f"Uploaded document not found: {uploaded.get('filename', 'unknown')}")
                payload = await FileFormatService(allowed_roots=[file_path.parent]).read(
                    path=str(file_path),
                    max_chars=settings.agent_flow_document_max_chars,
                )
            else:
                raise ValueError(f"Unsupported document source: {source}")

            filename = str(payload.get("filename") or config.get("attachment_filename") or "document")
            content = str(payload.get("content") or "")
            if not content.strip():
                raise ValueError(f"Document '{filename}' contained no readable content")
            if payload.get("truncated"):
                raise ValueError(
                    f"Document '{filename}' exceeds the AgentFlow extraction safety limit "
                    f"of {settings.agent_flow_document_max_chars:,} characters. Split the document "
                    "or increase SHOGUN_AGENT_FLOW_DOCUMENT_MAX_CHARS."
                )
            output_parts.append(f"[Document: {filename}]\n{content}")
            log.info("[Flow] Input: read %s document %s (%d chars)", source, filename, len(content))
        except Exception as exc:
            raise ValueError(f"Could not read {source} document: {exc}") from exc

    # Add any context from upstream nodes
    if context_str:
        output_parts.append(context_str)

    if not output_parts:
        output_parts.append(f"Workflow triggered ({input_type})")

    return "\n\n".join(output_parts)


async def _exec_subflow(
    parent_run_id: uuid.UUID,
    node: AgentFlowNode,
    predecessor_outputs: dict[str, Any],
    run_input: dict[str, Any],
    governance_context: dict[str, Any],
) -> dict[str, Any]:
    """Resolve a Subflow node and execute it through the existing DAG engine."""
    config = node.config or {}
    child_flow_id = config.get("child_flow_id")
    if not child_flow_id:
        raise ValueError("Subflow execution failed: no child flow is selected.")
    try:
        child_uuid = uuid.UUID(str(child_flow_id))
    except (TypeError, ValueError) as exc:
        raise ValueError("Subflow execution failed: the selected child flow ID is invalid.") from exc

    mapping_context = {
        "input": run_input.get("input", run_input),
        "node": {node_id: {"output": output} for node_id, output in predecessor_outputs.items()},
        "artifacts": run_input.get("artifacts", {}),
        "context": {**run_input.get("context", {}), **governance_context},
        "governance": governance_context,
    }
    input_mapping = config.get("input_mapping") or {}
    if input_mapping:
        child_input = resolve_flow_mapping(input_mapping, mapping_context)
    elif predecessor_outputs:
        last_output = list(predecessor_outputs.values())[-1]
        if isinstance(last_output, dict) and "output" in last_output:
            child_input = _json_object(last_output["output"], "Child flow input")
        elif isinstance(last_output, dict):
            child_input = last_output
        else:
            child_input = {"input": last_output}
    else:
        child_input = run_input

    options = ChildFlowExecutionOptions(
        version_mode=config.get("child_flow_version_mode", "locked"),
        flow_version=config.get("child_flow_version"),
        timeout_seconds=config.get("timeout_seconds"),
        execution_mode=config.get("execution_mode", "sequential"),
        on_failure=config.get("on_failure", "fail_parent"),
        max_retries=max(0, int(config.get("max_retries", 0))),
    )
    result = await execute_child_flow(
        parent_run_id=parent_run_id,
        parent_node_id=node.id,
        child_flow_id=child_uuid,
        child_input=child_input,
        options=options,
        governance_context=governance_context,
    )
    output_mapping = config.get("output_mapping") or {}
    if output_mapping:
        result["raw_output"] = result.get("output", {})
        result["output"] = resolve_flow_mapping(
            output_mapping,
            {"child": result, "output": result.get("output", {})},
        )
    return result


async def execute_child_flow(
    parent_run_id: uuid.UUID,
    parent_node_id: uuid.UUID,
    child_flow_id: uuid.UUID,
    child_input: dict[str, Any],
    options: ChildFlowExecutionOptions,
    governance_context: dict[str, Any],
) -> dict[str, Any]:
    """Create, govern, await, and audit one child run using `_execute_flow`."""
    from shogun.config import settings
    from shogun.edition import feature_available
    from shogun.services.event_logger import EventLogger

    if not feature_available("flow_stack") or not settings.flow_stacking_enabled:
        raise ValueError("Flow Stacking is not available in Yellow Label.")

    child_input = _json_object(child_input, "Child flow input")
    child_run_id = uuid.uuid4()
    async with _parallel_child_semaphore():
        async with async_session_factory() as session:
            parent = await session.get(AgentFlowRun, parent_run_id)
            child_flow = await session.get(AgentFlow, child_flow_id)
            if not parent:
                raise ValueError("Subflow execution failed: parent run no longer exists.")
            if not child_flow or child_flow.is_deleted:
                raise ValueError("Subflow execution failed: the selected child flow no longer exists.")
            if not child_flow.allow_as_subflow:
                raise ValueError("Subflow blocked: the selected flow does not allow child execution.")
            await _validate_child_safety(session, parent, child_flow)
            _validate_child_permissions(governance_context, child_flow.required_tools or [])

            version_mode = options.version_mode or "locked"
            if version_mode not in {"locked", "latest"}:
                raise ValueError("Subflow blocked: version mode must be 'locked' or 'latest'.")
            if version_mode == "latest" and not settings.flow_stacking_allow_latest_version:
                raise ValueError("Subflow blocked: latest-version references are disabled.")
            if (
                version_mode == "locked"
                and options.flow_version is not None
                and options.flow_version != child_flow.version
            ):
                raise ValueError(
                    f"Subflow blocked: locked version {options.flow_version} is unavailable; "
                    f"current version is {child_flow.version}."
                )

            root_run_id = parent.root_run_id or parent.id
            child_governance = _inherit_governance(governance_context, child_flow)
            child_run = AgentFlowRun(
                id=child_run_id,
                flow_id=child_flow.id,
                flow_version=child_flow.version,
                root_run_id=root_run_id,
                parent_run_id=parent.id,
                parent_node_id=parent_node_id,
                run_depth=parent.run_depth + 1,
                status="pending",
                trigger_type="subflow",
                node_states={},
                result_summary={},
                input_payload=child_input,
                output_payload={},
                artifacts=[],
                governance_context=child_governance,
            )
            edge = AgentFlowRunEdge(
                root_run_id=root_run_id,
                parent_run_id=parent.id,
                child_run_id=child_run.id,
                parent_node_id=parent_node_id,
                child_flow_id=child_flow.id,
                execution_mode=options.execution_mode,
                status="created",
            )
            session.add_all([child_run, edge])
            await session.commit()
            child_name = child_flow.name
            child_risk = child_flow.risk_tier

        await EventLogger.emit_governance_event(
            "flow.subflow.started",
            f"Child flow '{child_name}' started",
            trace_id=str(parent_run_id),
            session_id=str(child_run_id),
            risk_score=child_risk,
            detail={
                "root_run_id": str(root_run_id),
                "parent_run_id": str(parent_run_id),
                "child_run_id": str(child_run_id),
                "child_flow_id": str(child_flow_id),
                "parent_node_id": str(parent_node_id),
                "run_depth": parent.run_depth + 1,
                "flow_version": child_flow.version,
            },
            governance_flags=child_governance,
        )

        retries = options.max_retries + 1
        for attempt in range(retries):
            task = asyncio.create_task(_execute_flow(child_run_id, child_flow_id))
            _active_runs[str(child_run_id)] = task
            try:
                timeout = (
                    options.timeout_seconds
                    or child_flow.default_timeout_seconds
                    or settings.flow_stacking_default_timeout_seconds
                )
                await asyncio.wait_for(task, timeout=timeout)
                break
            except TimeoutError as exc:
                task.cancel()
                await _cancel_run_record(child_run_id, f"Subflow timed out after {timeout} seconds")
                if attempt + 1 >= retries:
                    await EventLogger.emit_governance_event(
                        "flow.subflow.failed",
                        f"Child flow '{child_name}' timed out",
                        result="failure",
                        severity="error",
                        trace_id=str(parent_run_id),
                        session_id=str(child_run_id),
                        risk_score=child_risk,
                        detail={"child_run_id": str(child_run_id), "timeout_seconds": timeout},
                    )
                    raise ValueError(
                        f"Subflow timed out after {timeout} seconds. The child run was cancelled."
                    ) from exc
            finally:
                _active_runs.pop(str(child_run_id), None)

        async with async_session_factory() as session:
            completed = await session.get(AgentFlowRun, child_run_id)
            if not completed:
                raise ValueError("Subflow execution failed: child run record disappeared.")
            status = completed.status
            output = completed.output_payload or completed.result_summary or {}
            result = {
                "child_run_id": str(completed.id),
                "child_flow_id": str(completed.flow_id),
                "flow_version": completed.flow_version,
                "status": status,
                "output": output,
                "artifacts": completed.artifacts or [],
                "summary": completed.result_summary or {},
                "errors": [completed.error_message] if completed.error_message else [],
            }

        event_type = "flow.subflow.completed" if status == "completed" else "flow.subflow.failed"
        await EventLogger.emit_governance_event(
            event_type,
            f"Child flow '{child_name}' {status}",
            result="success" if status == "completed" else "failure",
            severity="info" if status == "completed" else "error",
            trace_id=str(parent_run_id),
            session_id=str(child_run_id),
            risk_score=child_risk,
            detail=result,
            governance_flags=child_governance,
        )
        if status != "completed" and options.on_failure == "fail_parent":
            raise ValueError(completed.error_message or f"Child flow failed with status {status}.")
        return result


async def _validate_child_safety(session: AsyncSession, parent: AgentFlowRun, child_flow: AgentFlow) -> None:
    from shogun.config import settings

    if parent.flow_id == child_flow.id:
        raise ValueError("Subflow blocked: a flow cannot call itself.")
    next_depth = parent.run_depth + 1
    max_depth = min(settings.flow_stacking_max_depth, settings.flow_stacking_hard_max_depth)
    if next_depth > max_depth:
        raise ValueError(f"Subflow blocked: maximum hierarchy depth of {max_depth} would be exceeded.")

    ancestor_flow_ids = {parent.flow_id}
    ancestor_id = parent.parent_run_id
    while ancestor_id:
        ancestor = await session.get(AgentFlowRun, ancestor_id)
        if not ancestor:
            break
        ancestor_flow_ids.add(ancestor.flow_id)
        ancestor_id = ancestor.parent_run_id
    if child_flow.id in ancestor_flow_ids:
        raise ValueError("Subflow blocked: cycle detected. The child flow is already an ancestor.")

    child_count = await session.scalar(
        select(func.count()).select_from(AgentFlowRun).where(AgentFlowRun.parent_run_id == parent.id)
    )
    if int(child_count or 0) >= settings.flow_stacking_max_child_runs_per_parent:
        raise ValueError("Subflow blocked: maximum child runs for this parent has been reached.")
    root_count = await session.scalar(
        select(func.count())
        .select_from(AgentFlowRun)
        .where(AgentFlowRun.root_run_id == (parent.root_run_id or parent.id))
    )
    if int(root_count or 0) >= settings.flow_stacking_max_total_runs_per_root:
        raise ValueError("Subflow blocked: maximum total runs for this execution tree has been reached.")


def _validate_child_permissions(governance: dict[str, Any], required_tools: list[str]) -> None:
    denied = [tool for tool in required_tools if not _tool_allowed_by_governance(tool, governance)]
    if denied:
        raise ValueError(
            "Subflow blocked: the child flow requires tools that are not allowed by the parent "
            f"governance context: {', '.join(sorted(denied))}."
        )


def _tool_allowed_by_governance(tool: str, governance: dict[str, Any]) -> bool:
    permissions = governance.get("permissions", governance)
    explicit = governance.get("allowed_tools")
    if isinstance(explicit, list) and tool not in explicit:
        return False
    checks = {
        "mado_browser": "mado_enabled",
        "browse_web": "mado_enabled",
        "take_screenshot": "mado_enabled",
        "email_send": "comms_send_email",
        "send_email": "comms_send_email",
        "workspace": "workspace_enabled",
        "office": "office_enabled",
        "shell": "shell_enabled",
    }
    permission = checks.get(tool)
    return True if permission is None else bool(permissions.get(permission, False))


async def _root_governance_context(requested: dict[str, Any]) -> dict[str, Any]:
    from shogun.services.posture_guard import get_posture_permissions

    permissions = await get_posture_permissions()
    context = {
        "posture_level": permissions.get("active_tier", "tactical"),
        "approval_mode": requested.get("approval_mode", "inherit"),
        "permissions": permissions,
        "audit_context": requested.get("audit_context", {}),
        "workspace_boundaries": requested.get("workspace_boundaries", []),
        "customer_context": requested.get("customer_context", {}),
        "risk_tier": requested.get("risk_tier", "low"),
    }
    if isinstance(requested.get("allowed_tools"), list):
        context["allowed_tools"] = list(dict.fromkeys(requested["allowed_tools"]))
    return context


def _inherit_governance(parent: dict[str, Any], child_flow: AgentFlow) -> dict[str, Any]:
    inherited = json.loads(json.dumps(parent, default=str))
    risk_rank = {"low": 0, "medium": 1, "high": 2, "critical": 3}
    parent_risk = inherited.get("risk_tier", "low")
    inherited["risk_tier"] = max(
        (parent_risk, child_flow.risk_tier),
        key=lambda value: risk_rank.get(str(value), 0),
    )
    inherited["inherited"] = True
    inherited["child_flow_id"] = str(child_flow.id)
    inherited["child_required_tools"] = list(child_flow.required_tools or [])
    _apply_flow_generation_settings(inherited, child_flow)
    return inherited


def _apply_flow_generation_settings(context: dict[str, Any], flow: AgentFlow) -> None:
    """Replace inherited generation controls with the settings of this flow."""
    context.pop("flow_seed", None)
    context.pop("flow_seed_model_id", None)
    if flow.seed is not None:
        context["flow_seed"] = int(flow.seed)
        if flow.seed_model_id:
            context["flow_seed_model_id"] = str(flow.seed_model_id)


def _with_flow_generation_settings(
    routing_context: dict[str, Any] | None,
    governance_context: dict[str, Any] | None,
) -> dict[str, Any]:
    """Add the current flow's seed controls to a model routing decision."""
    combined = dict(routing_context or {})
    if (governance_context or {}).get("flow_seed") is not None:
        combined["flow_seed"] = int(governance_context["flow_seed"])
        if (governance_context or {}).get("flow_seed_model_id"):
            combined["flow_seed_model_id"] = str(governance_context["flow_seed_model_id"])
    return combined


_FLOW_TOKEN = re.compile(r"{{\s*([^{}]+?)\s*}}")


def resolve_flow_mapping(value: Any, context: dict[str, Any]) -> Any:
    """Resolve the MVP Flow Stacking mapping syntax recursively."""
    if isinstance(value, dict):
        return {key: resolve_flow_mapping(item, context) for key, item in value.items()}
    if isinstance(value, list):
        return [resolve_flow_mapping(item, context) for item in value]
    if not isinstance(value, str):
        return value
    exact = _FLOW_TOKEN.fullmatch(value)
    if exact:
        return _lookup_flow_token(context, exact.group(1))
    return _FLOW_TOKEN.sub(lambda match: str(_lookup_flow_token(context, match.group(1))), value)


def _lookup_flow_token(context: dict[str, Any], path: str) -> Any:
    current: Any = context
    for part in path.split("."):
        if isinstance(current, dict) and part in current:
            current = current[part]
        elif isinstance(current, list) and part.isdigit() and int(part) < len(current):
            current = current[int(part)]
        else:
            raise ValueError(f"Flow mapping could not resolve '{{{{{path}}}}}'.")
    return current


def _json_object(value: Any, label: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        value = {"input": value}
    try:
        serialized = json.dumps(value, default=str)
        result = json.loads(serialized)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label} must be JSON-serializable.") from exc
    return result


async def _infer_template_mapping_plan(
    *,
    task_description: str,
    fixed_context: str,
    source_context: str,
    agent_persona: str,
    model_chain: list[tuple[ModelProvider, str, str, dict]],
    routing_context: dict[str, Any] | None,
    expected_width: int | None,
    timeout: int,
    max_tokens: int | None,
) -> list[list[Any]] | None:
    """Infer destination-column logic once instead of rediscovering it per chunk.

    The returned matrix is executable guidance, not business data. Each row is
    ``[destination index, destination column, source evidence, rule]``. A
    malformed or incomplete plan is non-fatal; the established direct mapping
    path remains available as a compatibility fallback.
    """
    if "[POPULATED ONE-SHOT EXAMPLE]" not in fixed_context:
        return None
    sample = _representative_document_sample(source_context)
    prompt = (
        "Create a reusable mapping plan for this document-to-template transformation. "
        "The populated template is reference-only: learn its layout and logic, but never copy its "
        "business values. Runtime source data is the sole source of output records.\n\n"
        f"--- TRANSFORMATION INSTRUCTIONS ---\n{task_description}\n\n"
        f"--- TEMPLATE CONTRACT AND REFERENCE EXAMPLE ---\n{fixed_context}\n\n"
        f"--- REPRESENTATIVE RUNTIME SOURCE SAMPLES ---\n{sample}\n\n"
        "Submit a two-dimensional planning matrix only. Each planning row must contain exactly four values: "
        "[zero-based destination column index, destination column name, source field/evidence, mapping or "
        "normalization rule]. Include one planning row for every destination column, in destination order. "
        "Do not submit any source business record or example value as an output row."
    )
    try:
        rows = await _call_llm_chain_rows(
            [
                {"role": "system", "content": agent_persona},
                {"role": "user", "content": prompt},
            ],
            model_chain,
            timeout=timeout,
            retry_count=0,
            context="AgentFlow Samurai transformation planning",
            expected_width=4,
            max_tokens=min(int(max_tokens or 4096), 4096),
            routing_context=routing_context,
        )
        if expected_width is not None and len(rows) != expected_width:
            raise ValueError(
                f"mapping plan returned {len(rows)} destination columns; template requires {expected_width}"
            )
        return rows
    except Exception as exc:
        log.warning("Transformation planning was unavailable; using direct chunk mapping: %s", exc)
        return None


async def _exec_samurai(
    config: dict,
    context_str: str,
    governance_context: dict[str, Any] | None = None,
    progress_callback: Callable[[int, int], Awaitable[None]] | None = None,
    fixed_context_str: str = "",
) -> str:
    """Samurai node — delegates task to LLM using agent's routing profile."""
    task_description = await _resolve_samurai_task_description(config)
    expected_output = config.get("expected_output", "")
    agent_id = config.get("agent_id")
    routing_profile_id = (governance_context or {}).get("model_profile") or config.get("routing_profile_id")
    timeout = config.get("timeout", 300)
    retry_count = config.get("retry_count", 0)

    if not task_description:
        raise ValueError("Samurai node has no task description")

    # Preserve the original one-node workflow contract for the two governed,
    # read-only personal-data tools. The flow runtime executes them through the
    # same posture and ToolGate checks as dedicated Email/Calendar Read nodes,
    # then gives their structured results to the Samurai for synthesis.
    native_read_context: dict[str, Any] = {}
    for tool_name in _samurai_native_tool_references(task_description, config):
        if tool_name == "fetch_inbox":
            native_read_context[tool_name] = await _exec_email_read(
                {
                    "folder": config.get("email_folder", "INBOX"),
                    "page": config.get("email_page", 1),
                    "per_page": config.get("email_per_page", 10),
                    "unread_only": config.get("email_unread_only", False),
                }
            )
        elif tool_name == "list_calendar_events":
            native_read_context[tool_name] = await _exec_calendar_read(
                {
                    "start_date": config.get("calendar_start_date"),
                    "end_date": config.get("calendar_end_date"),
                    "days_ahead": config.get("calendar_days_ahead", 1),
                }
            )

    # Build the prompt
    combined_context = "\n\n".join(part for part in (fixed_context_str, context_str) if part)
    user_message = task_description
    if combined_context:
        user_message = f"{task_description}\n\n--- CONTEXT FROM PREVIOUS STEPS ---\n{combined_context}"
    if native_read_context:
        user_message += (
            "\n\n--- GOVERNED NATIVE READ RESULTS ---\n"
            + json.dumps(native_read_context, default=str, ensure_ascii=False)
            + "\nUse these results to complete the task. Do not claim that the reads failed or request "
            "another tool call; the tools have already been executed by the flow runtime."
        )
    input_artifacts = [
        item
        for item in (config.get("_input_artifacts") or [])
        if isinstance(item, dict) and item.get(_FLOW_ARTIFACT_MARKER)
    ]
    output_contracts = [item for item in (config.get("_output_contracts") or []) if isinstance(item, dict)]
    if input_artifacts:
        user_message += (
            "\n\n--- FLOW INPUT ARTIFACT MANIFEST ---\n"
            + json.dumps(input_artifacts, ensure_ascii=False, default=str)
            + "\nThe complete content of these artifacts is supplied in the preceding context. "
            "Treat templates as reference-only and runtime inputs as the sole source of business data."
        )
    if output_contracts:
        user_message += (
            "\n\n--- GOVERNED DOWNSTREAM OUTPUT CONTRACT ---\n"
            + json.dumps(output_contracts, ensure_ascii=False, default=str)
            + "\nReturn the typed content required by this downstream node. Do not write, rename, or "
            "save the destination file yourself; the downstream node exclusively owns that side effect."
        )
    if expected_output:
        user_message += f"\n\n--- EXPECTED OUTPUT FORMAT ---\n{expected_output}"

    # Decide whether this is a template-backed bulk transformation before
    # routing. Routing the entire 300-page source first can incorrectly remove
    # otherwise useful fallback models whose context is ample for an
    # individual chunk. The chunk executor performs its own bounded routing.
    downstream_excel_output = any(
        item.get("action") in {"excel_create", "excel_write"}
        or item.get("format") == "xlsx"
        for item in output_contracts
    )
    requires_matrix_output = downstream_excel_output or _requires_structured_matrix_output(
        task_description,
        str(config.get("expected_output") or ""),
        fixed_context_str,
    )
    matrix_width_match = re.search(r'"logical_columns"\s*:\s*(\d+)', fixed_context_str)
    if not matrix_width_match:
        matrix_width_match = re.search(
            r"exactly\s+(\d+)\s+values",
            f"{task_description}\n{expected_output}",
            re.IGNORECASE,
        )
    expected_matrix_width = int(matrix_width_match.group(1)) if matrix_width_match else None

    transformation_profile = _active_transformation_profile(config)
    transformation_profile_evidence = _active_transformation_profile_evidence(
        config,
        transformation_profile,
    )
    if transformation_profile and not requires_matrix_output:
        raise ValueError("A document transformation profile requires a matrix output contract.")
    if transformation_profile:
        if not context_str or not fixed_context_str:
            raise ValueError(
                "A document transformation profile requires runtime source data and a file template."
            )
        try:
            source_contexts = config.get("_transformation_source_contexts") or [
                {"label": "combined runtime source", "content": context_str}
            ]
            if not isinstance(source_contexts, list) or any(
                not isinstance(source, dict) for source in source_contexts
            ):
                raise ValueError("Samurai transformation source contexts must be a list of objects.")
            _validate_transformation_profile_sources(transformation_profile, source_contexts)
            deterministic = try_deterministic_matrix_transform(
                profile=transformation_profile,
                source_context=context_str,
                fixed_context=fixed_context_str,
            )
            if expected_matrix_width is not None and any(
                len(row) != expected_matrix_width for row in deterministic.rows
            ):
                raise ValueError(
                    f"Transformation profile {deterministic.profile_id} returned rows outside the "
                    f"{expected_matrix_width}-column template contract."
                )
            _validate_matrix_coverage(
                deterministic.rows,
                context_str,
                task_description,
                config,
                label=f"Transformation profile {deterministic.profile_id}",
            )
        except Exception as deterministic_error:
            if not transformation_profile.get("model_fallback", False):
                raise
            log.warning(
                "Transformation profile %s failed closed validation; explicit model fallback is enabled "
                "(error_type=%s)",
                transformation_profile.get("id"),
                type(deterministic_error).__name__,
            )
            # Once the operator opts into a model fallback, use the completely
            # generic extraction path. Profile-aware source splitting and row
            # validation would otherwise re-run the same mismatched contract.
            config = {**config, "_transformation_profiles": []}
            transformation_profile = None
        else:
            if progress_callback:
                progress_total = max(1, len(context_str))
                try:
                    await progress_callback(progress_total, progress_total)
                except Exception as progress_error:
                    log.warning("Could not persist AgentFlow node progress: %s", progress_error)
            log.info(
                "AgentFlow Samurai used transformation profile %s with adapter %s for %d validated row(s) "
                "(source=%s, version=%s, hash=%s)",
                deterministic.profile_id,
                deterministic.adapter_id,
                len(deterministic.rows),
                (transformation_profile_evidence or {}).get("source")
                or (transformation_profile_evidence or {}).get("profile_source")
                or "legacy",
                (transformation_profile_evidence or {}).get("version") or "unrecorded",
                (transformation_profile_evidence or {}).get("content_hash") or "unrecorded",
            )
            if transformation_profile_evidence:
                await _record_samurai_transformation_profile_event(
                    config,
                    deterministic.profile_id,
                    deterministic.adapter_id,
                    len(deterministic.rows),
                    transformation_profile_evidence,
                )
            return json.dumps(deterministic.rows, ensure_ascii=False, default=str)

    # Resolve agent persona
    agent_persona = "You are a Samurai agent executing a task in an automated workflow."
    if "[POPULATED ONE-SHOT EXAMPLE]" in fixed_context_str:
        agent_persona += (
            " Populated file-template examples are reference-only demonstrations of structure and formatting."
            " Never treat their records or values as factual input. Generate business data only from the"
            " non-template runtime inputs supplied for this execution."
        )
    async with async_session_factory() as session:
        from shogun.services.model_router import NoEligibleModelError, configured_max_input_tokens

        if agent_id:
            agent_result = await session.execute(select(Agent).where(Agent.id == uuid.UUID(agent_id)))
            agent = agent_result.scalar_one_or_none()
            if agent:
                agent_persona = (
                    f"You are {agent.name}, a Samurai agent. "
                    f"{agent.description or ''}\n"
                    "You are executing a task as part of an automated Agent Flow workflow. "
                    "Respond with the requested output directly. Do not ask clarifying questions."
                )
                # Use agent's routing profile if not overridden
                if not routing_profile_id and agent.model_routing_profile_id:
                    routing_profile_id = str(agent.model_routing_profile_id)

        # Resolve a task-aware primary + fallback chain. Older databases fall
        # back to the original profile resolver inside this helper.
        route_options = {
            "task_type": config.get("task_type") or "stack_step_execution",
            # Samurai nodes synthesize predecessor/runtime-fetched context and
            # do not expose native tools to the model. Legacy/generated
            # ``requires_tools`` metadata must therefore not disqualify an
            # otherwise valid chat model.
            "required_capabilities": ["chat"],
            "routing_profile_id": routing_profile_id,
            "stack_run_id": (governance_context or {}).get("stack_run_id"),
            "step_id": (governance_context or {}).get("stack_step_id"),
            "retry_count": retry_count,
            "risk_level": config.get("risk_level", "low"),
        }
        # Prefer one bounded model turn when the selected route can hold the
        # complete request. This is the same interaction model users expect
        # from a chat with attachments. Chunking is a capacity/performance
        # fallback, not the default merely because an Excel template exists.
        chunk_required = False
        try:
            model_chain, _routing = await _resolve_task_llm_chain(
                session,
                prompt=user_message,
                context_size_estimate=max(1, len(user_message) // 4),
                **route_options,
            )
        except NoEligibleModelError as exc:
            # A document can be larger than every model's single-request
            # context while still being perfectly processable in batches. Keep
            # every other routing/policy failure authoritative. The router can
            # report this either as exhausted capacity or as a missing
            # ``long_context`` capability, depending on whether another
            # registry model survives its capacity filter before the active
            # profile is applied.
            routing_error = str(exc).lower()
            if not any(
                marker in routing_error
                for marker in ("context capacity", "enough context", "long_context")
            ):
                raise
            chunk_required = True
            try:
                model_chain, _routing = await _resolve_task_llm_chain(
                    session,
                    prompt=task_description,
                    context_size_estimate=0,
                    **route_options,
                )
            except NoEligibleModelError as chunk_route_error:
                # The first error proves at least one enabled, connected model
                # existed before the single-request context filter. Upgraded
                # registries can nevertheless fail the second capability pass
                # because of stale row metadata. Preserve the configured/legacy
                # connected chain for chunk execution instead of resurrecting a
                # misleading "Required capabilities: chat" failure.
                model_chain = await _resolve_llm_chain(session, routing_profile_id)
                if not model_chain:
                    raise chunk_route_error
                provider, model_name, *_ = model_chain[0]
                registry_entry = None
                try:
                    registry_entry = await session.scalar(
                        select(ModelRegistryEntry).where(
                            ModelRegistryEntry.provider_id == provider.id,
                            ModelRegistryEntry.model_id == model_name,
                        )
                    )
                except Exception:
                    pass
                if registry_entry:
                    selected_context = int(registry_entry.context_window)
                    selected_output = int(registry_entry.max_output_tokens)
                    selected_input = configured_max_input_tokens(registry_entry)
                else:
                    selected_context = int((provider.config or {}).get("context_window") or 8192)
                    selected_output = min(4096, max(128, selected_context // 4))
                    selected_input = max(1024, selected_context - selected_output)
                _routing = {
                    "active_profile": "connected_chunk_compatibility",
                    "selected_model": model_name,
                    "selected_provider": provider.provider_type,
                    "selected_context_window": selected_context,
                    "selected_max_input_tokens": selected_input,
                    "selected_max_output_tokens": selected_output,
                    "fallback_reason": str(chunk_route_error),
                }
                log.warning(
                    "Chunk routing bypassed stale registry eligibility for connected model %s/%s",
                    provider.name,
                    model_name,
                )

    if not model_chain:
        raise ValueError("No active LLM provider available for Samurai execution")

    _routing = _with_flow_generation_settings(_routing, governance_context)

    max_output_tokens = _routing.get("selected_max_output_tokens") if _routing else None
    max_input_tokens = int((_routing or {}).get("selected_max_input_tokens") or 0)
    if not max_input_tokens:
        context_window = int((_routing or {}).get("selected_context_window") or 8192)
        max_input_tokens = max(1024, context_window - int(max_output_tokens or 2048))

    primary_chain_item = model_chain[0]
    primary_provider = (
        primary_chain_item[0]
        if isinstance(primary_chain_item, (list, tuple)) and primary_chain_item
        else primary_chain_item
    )
    is_local_model = bool(getattr(primary_provider, "is_local", False)) or getattr(
        primary_provider, "provider_type", ""
    ) in {"ollama", "lmstudio", "local"}
    local_batch_limit = max(
        1024,
        min(
            int(
                (
                    config.get("local_matrix_chunk_tokens")
                    if requires_matrix_output
                    else config.get("local_document_chunk_tokens")
                )
                or 8192
            ),
            32_768,
        ),
    )
    local_chunk_required = bool(
        is_local_model and context_str and len(context_str) // 4 > local_batch_limit
    )
    structural_source_units = len(_model_source_units_for_config(context_str, config)) if context_str else 0
    exhaustive_matrix_extraction = bool(
        requires_matrix_output
        and context_str
        and structural_source_units > 1
        and _exhaustive_matrix_task(task_description)
    )

    if context_str and (
        chunk_required
        or local_chunk_required
        or exhaustive_matrix_extraction
        or len(user_message) // 4 > max_input_tokens
    ):
        checkpoint_fingerprint = hashlib.sha256(
            json.dumps(
                {
                    "task": task_description,
                    "expected_output": expected_output,
                    "template": fixed_context_str,
                    "source": context_str,
                    "routing_profile_id": routing_profile_id,
                    "local_matrix_chunk_tokens": config.get("local_matrix_chunk_tokens"),
                    "matrix_extraction_strategy": 2,
                    "matrix_chunk_tokens": config.get("matrix_chunk_tokens"),
                    "matrix_chunk_max_units": config.get("matrix_chunk_max_units"),
                    "minimum_matrix_rows": config.get("minimum_matrix_rows"),
                    "minimum_source_coverage_ratio": config.get("minimum_source_coverage_ratio"),
                    "allow_sparse_matrix_output": config.get("allow_sparse_matrix_output"),
                },
                ensure_ascii=False,
                sort_keys=True,
                default=str,
            ).encode("utf-8")
        ).hexdigest()
        checkpoint = _load_samurai_checkpoint(config, checkpoint_fingerprint)
        mapping_plan = checkpoint.get("mapping_plan")
        if "[POPULATED ONE-SHOT EXAMPLE]" in fixed_context_str and not isinstance(mapping_plan, list):
            mapping_plan = await _infer_template_mapping_plan(
                task_description=task_description,
                fixed_context=fixed_context_str,
                source_context=context_str,
                agent_persona=agent_persona,
                model_chain=model_chain,
                routing_context=_routing,
                expected_width=expected_matrix_width,
                timeout=max(30, min(timeout, int(config.get("planning_timeout") or 180))),
                max_tokens=min(int(max_output_tokens or 2048), 2048),
            )
            if mapping_plan:
                checkpoint["mapping_plan"] = mapping_plan
                checkpoint["updated_at"] = datetime.now(timezone.utc).isoformat()
                _save_samurai_checkpoint(config, checkpoint)
        fixed_message = task_description
        compact_template_context = _template_contract_without_example(fixed_context_str)
        if compact_template_context:
            fixed_message += (
                "\n\n--- FIXED TEMPLATE OUTPUT CONTRACT ---\n"
                f"{compact_template_context}"
            )
        if mapping_plan:
            fixed_message += (
                "\n\n--- APPROVED TRANSFORMATION PLAN ---\n"
                + json.dumps(mapping_plan, ensure_ascii=False, default=str)
                + "\nApply this plan to runtime source records. The plan describes mapping logic only; "
                "it contains no authoritative business rows."
            )
        if expected_output:
            fixed_message += f"\n\n--- EXPECTED OUTPUT FORMAT ---\n{expected_output}"
        if requires_matrix_output:
            width_rule = (
                f" Every row must contain exactly {expected_matrix_width} values."
                if expected_matrix_width is not None
                else " Every row must match the Excel template's complete column count."
            )
            fixed_message += (
                "\n\n--- RUNTIME EXCEL OUTPUT CONTRACT ---\n"
                "Return only one valid JSON two-dimensional array. Each inner array is one Excel row;"
                f" each value is one cell.{width_rule} Do not return Markdown, headings, commentary,"
                " summaries, objects, or code fences."
            )
        fixed_tokens = max(1, len(fixed_message) // 4) + 256
        chunk_token_budget = max_input_tokens - fixed_tokens
        if chunk_token_budget < 512:
            raise ValueError(
                "The selected model's input allocation is too small for this Samurai task. "
                "Increase Max input or reduce Max output in Katana."
            )
        # A model's theoretical context window is not a practical batch size,
        # especially for CPU/GPU-constrained local inference. Bound local
        # chunks so a 12B model is less likely to spend the entire node timeout
        # on one request. Timed-out chunks are bisected again below.
        if is_local_model:
            if requires_matrix_output:
                configured_batch_tokens = max(
                    1024,
                    min(int(config.get("local_matrix_chunk_tokens") or 8192), 32_768),
                )
                # Preserve every configured fallback that can accept the
                # compact fixed prompt plus a source batch. The router payload
                # carries these capacities so a large primary model does not
                # accidentally create chunks that exclude a smaller fallback.
                fallback_inputs = [
                    int(item.get("max_input_tokens") or 0)
                    for item in ((_routing or {}).get("fallback_models") or [])
                    if int(item.get("max_input_tokens") or 0) > 0
                ]
                if fallback_inputs:
                    configured_batch_tokens = min(
                        configured_batch_tokens,
                        max(1024, min(fallback_inputs) - fixed_tokens),
                    )
                chunk_token_budget = min(chunk_token_budget, configured_batch_tokens)
            else:
                chunk_token_budget = min(chunk_token_budget, 8192)
        elif requires_matrix_output:
            # A large remote context window makes the request fit, but it does
            # not make a single response a reliable container for hundreds of
            # spreadsheet rows. Keep source batches and result matrices
            # bounded independently of the advertised model context.
            remote_matrix_chunk_tokens = max(
                2048,
                min(int(config.get("matrix_chunk_tokens") or 24_576), 65_536),
            )
            chunk_token_budget = min(chunk_token_budget, remote_matrix_chunk_tokens)
        chunk_max_output_tokens = max_output_tokens
        if is_local_model and requires_matrix_output:
            chunk_max_output_tokens = min(int(max_output_tokens or 8192), 8192)
        elif requires_matrix_output:
            chunk_max_output_tokens = min(int(max_output_tokens or 32_768), 32_768)
        requested_local_chunk_timeout = int(config.get("local_chunk_timeout") or 600)
        local_chunk_timeout = max(60, min(requested_local_chunk_timeout, 1800))
        chunk_call_timeout = max(timeout, local_chunk_timeout) if is_local_model else timeout
        matrix_chunk_max_units = (
            max(1, min(int(config.get("matrix_chunk_max_units") or 10), 100))
            if requires_matrix_output
            else None
        )
        chunks = _split_model_context(
            context_str,
            chunk_token_budget * 4,
            max_units=matrix_chunk_max_units,
            profile=transformation_profile,
        )
        if checkpoint.get("total_chunks") not in {None, len(chunks)}:
            checkpoint["outputs"] = {}
        checkpoint["total_chunks"] = len(chunks)
        total_characters = max(1, sum(len(chunk) for chunk in chunks))
        completed_characters = 0
        initial_progress_characters = max(1, (total_characters + 99) // 100)
        outputs: list[str] = []
        progress_lock = asyncio.Lock()

        async def report_progress(completed: int) -> None:
            if not progress_callback:
                return
            try:
                visible_completed = (
                    total_characters
                    if completed >= total_characters
                    else max(completed, initial_progress_characters)
                )
                await progress_callback(visible_completed, total_characters)
            except Exception as exc:
                log.warning("Could not persist AgentFlow node progress: %s", exc)

        async def mark_chunk_completed(character_count: int) -> None:
            nonlocal completed_characters
            async with progress_lock:
                completed_characters = min(
                    total_characters,
                    completed_characters + character_count,
                )
                await report_progress(completed_characters)

        await report_progress(0)

        async def process_chunk(
            chunk: str,
            label: str,
            split_depth: int = 0,
            validation_retry: int = 0,
            validation_feedback: str = "",
            retained_rows: list[list[Any]] | None = None,
        ) -> list[str]:
            nonlocal completed_characters
            accepted_rows = [list(row) for row in (retained_rows or [])]
            minimum_rows = 0
            source_evidence = 0
            evidence_label = "source unit(s)"
            if requires_matrix_output:
                minimum_rows, source_evidence, evidence_label = _minimum_matrix_rows_for_source(
                    chunk,
                    task_description,
                    config,
                )
            chunk_message = (
                f"{fixed_message}\n\n"
                f"--- CONTEXT FROM PREVIOUS STEPS ({label}) ---\n{chunk}\n\n"
                "Process every relevant record in this chunk. Return only the requested structured "
                "data for this chunk; do not summarize, sample, or omit repeated records."
            )
            if minimum_rows:
                chunk_message += (
                    f"\n\n--- COMPLETENESS REQUIREMENT ---\nThis chunk contains "
                    f"{source_evidence} {evidence_label} and requires at least {minimum_rows} output row(s). "
                    "An empty or shorter matrix is invalid. Verify every required row before submitting."
                )
            if validation_feedback:
                chunk_message += (
                    "\n\n--- CORRECTIVE RETRY ---\nThe previous response failed deterministic validation: "
                    f"{validation_feedback}\nRe-read this complete source chunk and submit a corrected matrix."
                )
            if accepted_rows:
                missing_rows = max(0, minimum_rows - len(accepted_rows))
                chunk_message += (
                    "\n\n--- RETAINED VALID ROWS ---\n"
                    f"Shogun has safely retained {len(accepted_rows)} valid row(s) from the previous attempt. "
                    f"At least {missing_rows} additional row(s) are still required. Return only the missing "
                    "row(s); do not repeat any retained row. The retained rows are:\n"
                    + json.dumps(accepted_rows, ensure_ascii=False, default=str)
                )
            if native_read_context:
                chunk_message += (
                    "\n\n--- GOVERNED NATIVE READ RESULTS ---\n"
                    + json.dumps(native_read_context, default=str, ensure_ascii=False)
                )
            try:
                messages = [
                    {"role": "system", "content": agent_persona},
                    {"role": "user", "content": chunk_message},
                ]
                chunk_chain = model_chain
                chunk_routing = _routing
                # Eligibility must be evaluated against this actual batch.
                # The previous implementation froze a chain selected for the
                # complete document, which silently removed smaller-context
                # fallback models even after the source had been chunked.
                try:
                    async with async_session_factory() as chunk_session:
                        chunk_chain, chunk_routing = await _resolve_task_llm_chain(
                            chunk_session,
                            prompt=chunk_message,
                            context_size_estimate=max(1, len(chunk_message) // 4),
                            **route_options,
                        )
                    chunk_routing = _with_flow_generation_settings(
                        chunk_routing,
                        governance_context,
                    )
                except Exception as route_error:
                    log.warning(
                        "%s could not refresh chunk-level routing; using the established chain: %s",
                        label,
                        route_error,
                )
                if requires_matrix_output:
                    rows = await _call_llm_chain_rows_with_fallback(
                        messages,
                        chunk_chain,
                        timeout=chunk_call_timeout,
                        retry_count=retry_count,
                        context=f"AgentFlow Samurai node {label.lower()}",
                        expected_width=expected_matrix_width,
                        max_tokens=chunk_max_output_tokens,
                        routing_context=chunk_routing,
                        governance_context=governance_context,
                        row_validator=lambda candidate_rows: _validate_matrix_coverage(
                            _merge_matrix_attempt_rows(accepted_rows, candidate_rows),
                            chunk,
                            task_description,
                            config,
                            label=label,
                        ),
                    )
                    rows = _merge_matrix_attempt_rows(accepted_rows, rows)
                    _validate_matrix_coverage(
                        rows,
                        chunk,
                        task_description,
                        config,
                        label=label,
                    )
                    output = json.dumps(rows, ensure_ascii=False, default=str)
                else:
                    output = await _call_llm_chain(
                        messages,
                        chunk_chain,
                        timeout=chunk_call_timeout,
                        retry_count=retry_count,
                        context=f"AgentFlow Samurai node {label.lower()}",
                        max_tokens=chunk_max_output_tokens,
                        routing_context=chunk_routing,
                    )
                await mark_chunk_completed(len(chunk))
                return [output]
            except (
                ModelCallError,
                IncompleteMatrixOutputError,
                MalformedMatrixOutputError,
            ) as exc:
                incomplete = isinstance(exc, IncompleteMatrixOutputError)
                malformed = isinstance(exc, MalformedMatrixOutputError)
                timed_out = isinstance(exc, ModelCallError) and "timeout" in exc.cause_type.lower()
                error_text = str(exc).lower()
                context_rejected = any(
                    marker in error_text
                    for marker in (
                        "context length",
                        "context window",
                        "maximum context",
                        "prompt too long",
                        "too many tokens",
                    )
                )
                source_unit_count = len(_model_source_units_for_config(chunk, config))
                max_split_depth = 3 if incomplete or malformed else 2
                recoverable = incomplete or malformed or timed_out or context_rejected
                if not recoverable:
                    raise
                can_split = (
                    source_unit_count > 1 and split_depth < 8
                ) or (
                    split_depth < max_split_depth and len(chunk) > 2000
                )
                if can_split:
                    subchunks = _split_model_context(
                        chunk,
                        max(1000, len(chunk) // 2),
                        max_units=max(1, source_unit_count // 2) if source_unit_count > 1 else None,
                        profile=transformation_profile,
                    )
                    if len(subchunks) >= 2:
                        log.warning(
                            "%s was not safely complete (%s); retrying it as %d smaller parts",
                            label,
                            type(exc).__name__,
                            len(subchunks),
                        )
                        recovered: list[str] = []
                        for part_index, subchunk in enumerate(subchunks, start=1):
                            recovered.extend(
                                await process_chunk(
                                    subchunk,
                                    f"{label}, part {part_index}/{len(subchunks)}",
                                    split_depth + 1,
                                )
                            )
                        return recovered

                retry_rows = accepted_rows
                if incomplete:
                    retry_rows = _merge_matrix_attempt_rows(
                        accepted_rows,
                        getattr(exc, "candidate_rows", []),
                    )
                max_leaf_retries = max(
                    0,
                    min(int(config.get("matrix_leaf_retries") or 3), 3),
                )
                if (incomplete or malformed) and validation_retry < max_leaf_retries:
                    log.warning(
                        "%s cannot be subdivided further; issuing corrective matrix retry %d/%d",
                        label,
                        validation_retry + 1,
                        max_leaf_retries,
                    )
                    return await process_chunk(
                        chunk,
                        label,
                        split_depth,
                        validation_retry + 1,
                        str(exc),
                        retry_rows,
                    )
                raise

        checkpoint_outputs = dict(checkpoint.get("outputs") or {})
        matrix_concurrency = (
            max(1, min(int(config.get("matrix_chunk_concurrency") or 4), 8))
            if requires_matrix_output and not is_local_model
            else 1
        )
        chunk_semaphore = asyncio.Semaphore(matrix_concurrency)
        checkpoint_lock = asyncio.Lock()

        async def run_top_level_chunk(index: int, chunk: str) -> tuple[int, list[str]]:
            checkpoint_key = str(index)
            cached_output = checkpoint_outputs.get(checkpoint_key)
            if isinstance(cached_output, list) and all(
                isinstance(item, str) for item in cached_output
            ):
                cache_is_valid = True
                if requires_matrix_output:
                    try:
                        from shogun.services.file_template import parse_excel_rows

                        cached_rows: list[list[Any]] = []
                        for cached_part in cached_output:
                            cached_rows.extend(
                                parse_excel_rows(cached_part, require_structured_json=True)
                            )
                        cached_rows = _validate_agentflow_rows(
                            {"rows": cached_rows},
                            expected_matrix_width,
                        )
                        _validate_matrix_coverage(
                            cached_rows,
                            chunk,
                            task_description,
                            config,
                            label=f"checkpoint chunk {index}/{len(chunks)}",
                        )
                    except Exception as cache_error:
                        cache_is_valid = False
                        log.warning(
                            "Discarding incomplete AgentFlow checkpoint chunk %s/%s: %s",
                            index,
                            len(chunks),
                            cache_error,
                        )
                if cache_is_valid:
                    await mark_chunk_completed(len(chunk))
                    return index, cached_output
                async with checkpoint_lock:
                    checkpoint_outputs.pop(checkpoint_key, None)
                    checkpoint["outputs"] = checkpoint_outputs
                    checkpoint["updated_at"] = datetime.now(timezone.utc).isoformat()
                    _save_samurai_checkpoint(config, checkpoint)
            async with chunk_semaphore:
                chunk_outputs = await process_chunk(chunk, f"chunk {index}/{len(chunks)}")
            async with checkpoint_lock:
                checkpoint_outputs[checkpoint_key] = chunk_outputs
                checkpoint["outputs"] = checkpoint_outputs
                checkpoint["updated_at"] = datetime.now(timezone.utc).isoformat()
                _save_samurai_checkpoint(config, checkpoint)
            return index, chunk_outputs

        chunk_tasks = [
            asyncio.create_task(run_top_level_chunk(index, chunk))
            for index, chunk in enumerate(chunks, start=1)
        ]
        try:
            chunk_results = await asyncio.gather(*chunk_tasks)
        except BaseException:
            for task in chunk_tasks:
                if not task.done():
                    task.cancel()
            await asyncio.gather(*chunk_tasks, return_exceptions=True)
            raise
        for _index, chunk_outputs in sorted(chunk_results, key=lambda item: item[0]):
            outputs.extend(chunk_outputs)
        await report_progress(total_characters)
        merged = _merge_structured_chunk_matrices(
            outputs,
            task_description,
            fixed_context_str,
            config,
            force_matrix_output=requires_matrix_output,
        )
        if merged is not None:
            return merged
        return "\n".join(outputs)

    direct_progress_total = max(100, len(context_str or user_message))
    if progress_callback:
        try:
            await progress_callback(max(1, (direct_progress_total + 99) // 100), direct_progress_total)
        except Exception as progress_error:
            log.warning("Could not persist AgentFlow node progress: %s", progress_error)
    messages = [
        {"role": "system", "content": agent_persona},
        {"role": "user", "content": user_message},
    ]
    if requires_matrix_output:
        direct_matrix_max_tokens = (
            min(int(max_output_tokens or 8192), 8192)
            if is_local_model
            else max_output_tokens
        )
        rows = await _call_llm_chain_rows_with_fallback(
            messages,
            model_chain,
            timeout=timeout,
            retry_count=retry_count,
            context="AgentFlow Samurai node",
            expected_width=expected_matrix_width,
            max_tokens=direct_matrix_max_tokens,
            routing_context=_routing,
            governance_context=governance_context,
            row_validator=lambda candidate_rows: _validate_matrix_coverage(
                candidate_rows,
                context_str,
                task_description,
                config,
                label="AgentFlow Samurai node",
            ),
        )
        _validate_matrix_coverage(
            rows,
            context_str,
            task_description,
            config,
            label="AgentFlow Samurai node",
        )
        output = json.dumps(rows, ensure_ascii=False, default=str)
    else:
        output = await _call_llm_chain(
            messages,
            model_chain,
            timeout=timeout,
            retry_count=retry_count,
            context="AgentFlow Samurai node",
            max_tokens=max_output_tokens,
            routing_context=_routing,
        )
    if progress_callback:
        try:
            await progress_callback(direct_progress_total, direct_progress_total)
        except Exception as progress_error:
            log.warning("Could not persist AgentFlow node progress: %s", progress_error)
    merged = _merge_structured_chunk_matrices(
        [output],
        task_description,
        fixed_context_str,
        config,
        force_matrix_output=requires_matrix_output,
    )
    return merged if merged is not None else output


async def _resolve_samurai_task_description(config: dict[str, Any]) -> str:
    """Return the Samurai prompt, replacing typed text with an attached file.

    Instruction attachments are server-created AgentFlow uploads. Keeping the
    allowed root fixed prevents a manually edited flow configuration from
    turning this convenience feature into an arbitrary filesystem read.
    """
    typed_prompt = str(config.get("task_description") or "")
    instruction_file = config.get("instruction_file")
    if not instruction_file:
        return typed_prompt
    if config.get("_instruction_file_resolved"):
        return typed_prompt
    if not isinstance(instruction_file, dict):
        raise ValueError("Samurai instruction file configuration is invalid")

    from pathlib import Path

    from shogun.config import settings
    from shogun.services.file_formats import FileFormatService

    filename = str(instruction_file.get("filename") or "instruction file").strip()
    raw_path = str(instruction_file.get("path") or "").strip()
    if not raw_path:
        if instruction_file.get("error"):
            raise ValueError(
                f"The Samurai instruction file '{filename}' did not upload successfully. "
                "Remove it and upload the file again."
            )
        raise ValueError("Samurai node has no uploaded instruction file")

    file_path = Path(raw_path).resolve()
    upload_root = (Path(settings.uploads_path) / "agent_flows").resolve()
    try:
        file_path.relative_to(upload_root)
    except ValueError as exc:
        raise ValueError("Samurai instruction file must be an AgentFlow upload") from exc

    allowed_extensions = {".pdf", ".docx", ".md"}
    if file_path.suffix.lower() not in allowed_extensions:
        raise ValueError("Samurai instruction files must be PDF, Word (.docx), or Markdown (.md)")
    if not file_path.is_file():
        raise FileNotFoundError(f"Samurai instruction file not found: {filename}")

    payload = await FileFormatService(allowed_roots=[upload_root]).read(
        path=str(file_path),
        max_chars=settings.agent_flow_document_max_chars,
    )
    content = str(payload.get("content") or "")
    if not content.strip():
        raise ValueError(f"Samurai instruction file '{filename}' contained no readable text")
    if payload.get("truncated"):
        raise ValueError(
            f"Samurai instruction file '{filename}' exceeds the AgentFlow extraction safety limit "
            f"of {settings.agent_flow_document_max_chars:,} characters. Split the instructions "
            "or increase SHOGUN_AGENT_FLOW_DOCUMENT_MAX_CHARS."
        )
    return content


def _active_transformation_profile(config: dict[str, Any]) -> dict[str, Any] | None:
    from shogun.mapping.schema import MappingTransformationProfile

    raw_profiles = config.get("_transformation_profiles") or []
    if isinstance(raw_profiles, dict):
        raw_profiles = [raw_profiles]
    if not isinstance(raw_profiles, list):
        raise ValueError("Samurai transformation profiles must be supplied as a list of objects.")
    if any(not isinstance(profile, dict) for profile in raw_profiles):
        raise ValueError("Every Samurai transformation profile must be an object.")
    if len(raw_profiles) > 1:
        profile_ids = ", ".join(str(profile.get("id") or "unnamed") for profile in raw_profiles)
        raise ValueError(
            "Samurai received multiple transformation profiles "
            f"({profile_ids}). Select or connect exactly one profile."
        )
    if not raw_profiles:
        return None
    # Validate shape without rewriting the trusted immutable registry
    # definition. In particular, do not mix flow-config defaults or caller
    # fields back into the profile that deterministic execution consumes.
    MappingTransformationProfile.model_validate(raw_profiles[0])
    return deepcopy(raw_profiles[0])


def _active_transformation_profile_evidence(
    config: dict[str, Any],
    profile: dict[str, Any] | None,
) -> dict[str, Any] | None:
    """Return one trusted runtime evidence envelope without exposing mechanics."""

    raw_evidence = config.get("_transformation_profile_evidence") or []
    if isinstance(raw_evidence, dict):
        raw_evidence = [raw_evidence]
    if not isinstance(raw_evidence, list) or any(
        not isinstance(item, dict) for item in raw_evidence
    ):
        raise ValueError("Samurai transformation profile evidence must be a list of objects.")
    if len(raw_evidence) > 1:
        raise ValueError("Samurai received multiple transformation profile evidence envelopes.")
    if not raw_evidence:
        return None
    if profile is None:
        raise ValueError("Samurai received transformation profile evidence without a profile.")
    evidence = deepcopy(raw_evidence[0])
    if (
        str(evidence.get("profile_id") or "") != str(profile.get("id") or "")
        or str(evidence.get("adapter_id") or "") != str(profile.get("adapter") or "")
    ):
        raise ValueError("Samurai transformation profile evidence does not match its active profile.")
    return evidence


def _require_runtime_transformation_profile_evidence(
    profile: dict[str, Any],
    raw_evidence: list[dict[str, Any]],
) -> dict[str, Any]:
    """Require one immutable evidence envelope before runtime activation."""

    from shogun.services.transformation_profile_registry import profile_content_hash

    if (
        not isinstance(raw_evidence, list)
        or len(raw_evidence) != 1
        or not isinstance(raw_evidence[0], dict)
        or not raw_evidence[0]
    ):
        raise ValueError(
            "Samurai transformation profile activation requires trusted registry or private-file evidence."
        )
    evidence = _active_transformation_profile_evidence(
        {"_transformation_profile_evidence": raw_evidence},
        profile,
    )
    if evidence is None:
        raise ValueError(
            "Samurai transformation profile activation requires trusted registry or private-file evidence."
        )
    version = evidence.get("version")
    content_hash = str(evidence.get("content_hash") or "").lower()
    if isinstance(version, bool) or not isinstance(version, int) or version < 1:
        raise ValueError("Samurai transformation profile evidence has no valid version pin.")
    if not re.fullmatch(r"[a-f0-9]{64}", content_hash):
        raise ValueError("Samurai transformation profile evidence has no valid content-hash pin.")
    if content_hash != profile_content_hash(profile):
        raise ValueError(
            "Samurai transformation profile evidence does not match the resolved profile definition."
        )
    return evidence


async def _record_samurai_transformation_profile_event(
    config: dict[str, Any],
    profile_id: str,
    adapter_id: str,
    row_count: int,
    evidence: dict[str, Any],
) -> None:
    """Record profile provenance without placing its definition in the audit log."""

    try:
        from shogun.services.event_logger import EventLogger

        detail = {
            "flow_id": str(config.get("_flow_id") or ""),
            "flow_run_id": str(config.get("_run_id") or ""),
            "node_id": str(config.get("_node_id") or ""),
            "profile_id": profile_id,
            "adapter_id": adapter_id,
            "profile_version": evidence.get("version"),
            "content_hash": evidence.get("content_hash"),
            "profile_source": evidence.get("source") or evidence.get("profile_source"),
            "selection_mode": evidence.get("selection_mode") or "upstream_contract",
            "rows_validated": row_count,
        }
        source_resolution = evidence.get("source_intelligence")
        if isinstance(source_resolution, dict):
            detail["source_intelligence"] = deepcopy(source_resolution)
        await EventLogger.emit(
            category="decision",
            event_type="agent_flow.samurai.transformation_profile.executed",
            action=f"Samurai executed transformation profile '{profile_id}'",
            result="success",
            trace_id=str(config.get("_run_id") or "") or None,
            detail=detail,
        )
    except Exception:
        # Audit storage must not corrupt an otherwise deterministic output; the
        # structured engine log above retains the same immutable profile pins.
        log.exception("Could not write Samurai transformation-profile execution audit event")


def _validate_transformation_profile_sources(
    profile: dict[str, Any],
    sources: list[dict[str, Any]],
) -> None:
    """Require every direct runtime source to satisfy the explicit contract.

    The deterministic adapter ultimately receives one combined context so it
    can aggregate records across files. Validating only that combined string
    could let one matching PDF hide an unrelated or wrong-version PDF. These
    checks use only regexes declared by the profile; no filename or SAP/domain
    inference is involved.
    """
    parameters = profile.get("parameters") or {}
    raw_required = parameters.get("required_source_patterns") or []
    if not isinstance(raw_required, list):
        raise ValueError(
            f"Transformation profile '{profile.get('id')}' requires required_source_patterns to be a list."
        )
    patterns = [str(pattern) for pattern in raw_required]
    if profile.get("adapter") == "sectioned_record_matrix_v1":
        section_pattern = parameters.get("section_pattern")
        if section_pattern not in (None, ""):
            patterns.append(str(section_pattern))

    compiled: list[re.Pattern[str]] = []
    for pattern in dict.fromkeys(patterns):
        try:
            compiled.append(re.compile(pattern))
        except re.error as exc:
            raise ValueError(
                f"Transformation profile '{profile.get('id')}' has an invalid source regex: {exc}"
            ) from exc

    for index, source in enumerate(sources, start=1):
        label = str(source.get("label") or source.get("node_id") or f"source {index}")
        content = str(source.get("content") or "")
        if not content.strip() or any(pattern.search(content) is None for pattern in compiled):
            raise ValueError(
                f"Runtime source '{label}' does not match transformation profile '{profile.get('id')}'."
            )


def _model_source_units(text: str) -> list[str]:
    """Return stable document units suitable for bounded extraction and coverage checks."""
    if not text:
        return []
    marker = re.compile(r"(?m)(?=^--- (?:Page \d+|Excel row \d+|Slide \d+) ---\s*$)")
    marked_units = [unit for unit in marker.split(text) if unit]
    return marked_units if len(marked_units) > 1 else [text]


def _model_source_units_for_config(text: str, config: dict[str, Any]) -> list[str]:
    profile = _active_transformation_profile(config)
    if profile:
        return deterministic_profile_source_units(profile, text)
    return _model_source_units(text)


def _exhaustive_matrix_task(task_description: str) -> bool:
    """Identify contracts that promise exhaustive records rather than a sparse report."""
    task = str(task_description or "")
    return bool(
        re.search(
            r"\b(?:every|each)\s+(?:relevant\s+)?(?:record|row|order|item|material|entry|line)\b"
            r"|\ball\s+(?:relevant\s+)?(?:records|rows|orders|items|materials|entries|lines)\b"
            r"|\b(?:read|process|extract|convert)\s+(?:the\s+)?complete\b"
            r"|\bdo\s+not\s+(?:sample|omit|skip)\b",
            task,
            re.IGNORECASE,
        )
    )


def _minimum_matrix_rows_for_source(
    source_context: str,
    task_description: str,
    config: dict[str, Any],
) -> tuple[int, int, str]:
    """Return a conservative minimum row count and its source evidence."""
    units = len(_model_source_units_for_config(source_context, config))
    explicit = config.get("minimum_matrix_rows")
    if explicit not in (None, ""):
        return max(0, int(explicit)), units, "source unit(s)"
    if config.get("allow_sparse_matrix_output") or not _exhaustive_matrix_task(task_description):
        return 0, units, "source unit(s)"
    if not source_context.strip():
        return 0, units, "source unit(s)"

    transformation_profile = _active_transformation_profile(config)
    if transformation_profile is not None:
        expected_rows = expected_deterministic_matrix_rows(
            transformation_profile,
            source_context,
        )
        try:
            ratio = float(config.get("minimum_source_coverage_ratio", 1.0))
        except (TypeError, ValueError):
            ratio = 1.0
        ratio = max(0.05, min(ratio, 1.0))
        minimum = int((expected_rows * ratio) + 0.999999)
        return minimum, expected_rows, "profile-required row(s)"

    if units <= 1:
        return 1, units, "source unit(s)"
    try:
        ratio = float(config.get("minimum_source_coverage_ratio", 0.25))
    except (TypeError, ValueError):
        ratio = 0.25
    ratio = max(0.05, min(ratio, 1.0))
    return max(1, int((units * ratio) + 0.999999)), units, "source unit(s)"


def _merge_matrix_attempt_rows(
    retained_rows: list[list[Any]],
    candidate_rows: list[list[Any]],
) -> list[list[Any]]:
    """Merge extraction attempts without multiplying rows repeated by a retry.

    A corrective request may return only missing rows or repeat the complete
    partial matrix. Treat each attempt as a multiset and retain the maximum
    occurrence count for each exact row across attempts. This preserves
    legitimate duplicate source rows while preventing the same partial answer
    from being appended on every retry.
    """
    merged = [list(row) for row in retained_rows]
    retained_counts: dict[str, int] = defaultdict(int)
    candidate_counts: dict[str, int] = defaultdict(int)
    for row in merged:
        retained_counts[json.dumps(row, ensure_ascii=False, default=str)] += 1
    for row in candidate_rows:
        normalized_row = list(row)
        key = json.dumps(normalized_row, ensure_ascii=False, default=str)
        candidate_counts[key] += 1
        if candidate_counts[key] > retained_counts[key]:
            merged.append(normalized_row)
    return merged


def _validate_matrix_coverage(
    rows: list[list[Any]],
    source_context: str,
    task_description: str,
    config: dict[str, Any],
    *,
    label: str,
) -> None:
    """Reject shaped-but-obviously-incomplete exhaustive extraction results."""
    minimum_rows, source_evidence, evidence_label = _minimum_matrix_rows_for_source(
        source_context,
        task_description,
        config,
    )
    if len(rows) >= minimum_rows:
        return
    raise IncompleteMatrixOutputError(
        f"{label} returned only {len(rows)} row(s) for {source_evidence} {evidence_label}; "
        f"this exhaustive extraction requires at least {minimum_rows} row(s) before it can be accepted.",
        candidate_rows=rows,
        minimum_rows=minimum_rows,
    )


def _split_model_context(
    text: str,
    max_chars: int,
    max_units: int | None = None,
    profile: dict[str, Any] | None = None,
) -> list[str]:
    """Split long context at source boundaries without dropping or duplicating text."""
    limit = max(1000, int(max_chars))
    unit_limit = max(1, int(max_units)) if max_units not in (None, 0) else None
    source_units = (
        deterministic_profile_source_units(profile, text)
        if profile
        else _model_source_units(text)
    )
    if len(text) <= limit and (unit_limit is None or len(source_units) <= unit_limit):
        return [text]
    if len(source_units) > 1:
        chunks: list[str] = []
        current_units: list[str] = []
        current_length = 0
        for unit in source_units:
            if len(unit) > limit:
                if current_units:
                    chunks.append("".join(current_units))
                    current_units = []
                    current_length = 0
                if unit != text:
                    chunks.extend(_split_model_context(unit, limit, unit_limit, profile=profile))
                else:
                    chunks.extend(unit[index:index + limit] for index in range(0, len(unit), limit))
                continue
            if current_units and (
                current_length + len(unit) > limit
                or (unit_limit is not None and len(current_units) >= unit_limit)
            ):
                chunks.append("".join(current_units))
                current_units = []
                current_length = 0
            current_units.append(unit)
            current_length += len(unit)
        if current_units:
            chunks.append("".join(current_units))
        return chunks
    chunks: list[str] = []
    current = ""
    for part in re.split(r"(\n\n+)", text):
        if not part:
            continue
        while len(part) > limit:
            if current:
                chunks.append(current)
                current = ""
            chunks.append(part[:limit])
            part = part[limit:]
        if len(current) + len(part) > limit and current:
            chunks.append(current)
            current = ""
        current += part
    if current:
        chunks.append(current)
    return chunks


def _merge_structured_chunk_matrices(
    outputs: list[str],
    task_description: str,
    fixed_context: str,
    config: dict[str, Any],
    *,
    force_matrix_output: bool = False,
) -> str | None:
    """Combine independently generated 2D-array chunks into one validated matrix."""
    if not force_matrix_output and not _requires_structured_matrix_output(
        task_description,
        str(config.get("expected_output") or ""),
        fixed_context,
    ):
        return None

    from shogun.services.file_template import parse_excel_rows

    contract_text = f"{task_description}\n{config.get('expected_output') or ''}\n{fixed_context}"
    width_match = re.search(r'"logical_columns"\s*:\s*(\d+)', fixed_context)
    if not width_match:
        width_match = re.search(r"exactly\s+(\d+)\s+values", contract_text, re.IGNORECASE)
    expected_width = int(width_match.group(1)) if width_match else None
    rows: list[list[Any]] = []
    for output_index, output in enumerate(outputs, 1):
        parsed = parse_excel_rows(output, require_structured_json=True)
        if expected_width is not None:
            invalid = [(index + 1, len(row)) for index, row in enumerate(parsed) if len(row) != expected_width]
            if invalid:
                details = ", ".join(f"row {row} has {width}" for row, width in invalid[:6])
                raise ValueError(
                    f"Samurai chunk {output_index} violated the {expected_width}-column template contract: {details}."
                )
        rows.extend(parsed)

    # Exact-looking source rows can be legitimate repeated occurrences. Only
    # an explicit node configuration may collapse them; prose such as "do not
    # create duplicates" is too ambiguous to prove parser overlap.
    deduplicate = bool(config.get("deduplicate_rows"))
    if deduplicate:
        unique_rows: list[list[Any]] = []
        seen: set[str] = set()
        for row in rows:
            signature = json.dumps(row, ensure_ascii=False, sort_keys=True, default=str)
            if signature not in seen:
                seen.add(signature)
                unique_rows.append(row)
        rows = unique_rows

    log.info(
        "Merged %d structured Samurai chunk output(s) into %d validated row(s), width=%s, deduplicated=%s",
        len(outputs),
        len(rows),
        expected_width or "variable",
        deduplicate,
    )
    return json.dumps(rows, ensure_ascii=False, default=str)


def _requires_structured_matrix_output(
    task_description: str,
    expected_output: str,
    fixed_context: str,
) -> bool:
    """Require a row matrix for explicit contracts or upstream Excel templates."""
    contract_text = f"{task_description}\n{expected_output}"
    if re.search(r"two[- ]dimensional array|2d array|array of arrays", contract_text, re.IGNORECASE):
        return True
    if "[FILE TEMPLATE CONTRACT]" not in fixed_context:
        return False
    return bool(
        re.search(r"^Format:\s*xlsx\s*$", fixed_context, re.IGNORECASE | re.MULTILINE)
        or re.search(r'"kind"\s*:\s*"excel"', fixed_context, re.IGNORECASE)
    )


_SAMURAI_NATIVE_READ_TOOLS = ("fetch_inbox", "list_calendar_events")


def _samurai_native_tool_references(task_description: str, config: dict[str, Any]) -> list[str]:
    """Identify governed native reads requested by a backward-compatible Samurai node."""
    declared = {
        str(name)
        for name in [*(config.get("required_tools") or []), *(config.get("allowed_tools") or [])]
    }
    prompt = str(task_description or "")
    return [
        name
        for name in _SAMURAI_NATIVE_READ_TOOLS
        if name in declared or re.search(rf"\b{re.escape(name)}\b", prompt, re.IGNORECASE)
    ]


async def _exec_email_read(config: dict[str, Any]) -> dict[str, Any]:
    args = {
        "folder": str(config.get("folder") or "INBOX"),
        "page": max(1, int(config.get("page") or 1)),
        "per_page": max(1, min(int(config.get("per_page") or 10), 50)),
    }
    result = await _exec_governed_native_read("fetch_inbox", args)
    if config.get("unread_only", True):
        messages = [message for message in result.get("messages", []) if not message.get("is_read", False)]
        result = {**result, "messages": messages, "returned": len(messages), "unread_only": True}
    return result


def _collect_upstream_file_templates(
    node_id: str,
    predecessors: dict[str, list[str]],
    node_outputs: dict[str, Any],
) -> list[dict[str, Any]]:
    """Collect completed File Template payloads from all ancestors of a node."""
    found: list[dict[str, Any]] = []
    visited: set[str] = set()
    pending = list(predecessors.get(node_id, []))
    while pending:
        ancestor_id = pending.pop()
        if ancestor_id in visited:
            continue
        visited.add(ancestor_id)
        output = node_outputs.get(ancestor_id)
        if isinstance(output, dict) and output.get("__shogun_file_template__"):
            found.append(output)
        pending.extend(predecessors.get(ancestor_id, []))
    return found


async def _exec_file_template(config: dict[str, Any]) -> dict[str, Any]:
    """Extract a bounded model contract from a workspace template file."""
    from shogun.config import settings
    from shogun.services.file_template import extract_file_template

    payload = extract_file_template(
        template_path=str(config.get("template_path") or ""),
        workspace_root=settings.workspace_path,
        guidance_mode=str(config.get("guidance_mode") or "structure_only"),
        example_handling=str(config.get("example_handling") or "replace"),
        max_chars=min(settings.agent_flow_document_max_chars, 12_000),
        merge_key_columns=config.get("merge_key_columns"),
        merge_preserve_columns=config.get("merge_preserve_columns"),
    )
    payload["render_mode"] = str(config.get("render_mode") or "adaptive")
    payload["data_start_cell"] = str(config.get("data_start_cell") or "").strip()
    return payload


async def _exec_calendar_read(config: dict[str, Any]) -> dict[str, Any]:
    start_date = config.get("start_date")
    end_date = config.get("end_date")
    if not start_date or not end_date:
        start = datetime.now().astimezone().replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=max(1, min(int(config.get("days_ahead") or 1), 31)))
        start_date = start.isoformat()
        end_date = end.isoformat()
    return await _exec_governed_native_read(
        "list_calendar_events",
        {"start_date": str(start_date), "end_date": str(end_date)},
    )


async def _exec_governed_native_read(tool_name: str, args: dict[str, Any]) -> dict[str, Any]:
    """Execute one explicitly supported read tool under current posture and ToolGate policy."""
    from shogun.services.campaign_presets import get_preset
    from shogun.services.native_skills import NATIVE_TOOLS, execute_native_tool
    from shogun.services.posture_guard import filter_tools_by_posture, get_posture_tool_filter
    from shogun.services.tool_gate import GateAction, check_tool_access, get_toolgate_scope

    definition = next(
        (item for item in NATIVE_TOOLS if item.get("function", {}).get("name") == tool_name),
        None,
    )
    if definition is None:
        raise ValueError(f"AgentFlow native read tool '{tool_name}' is not registered.")

    posture = await get_posture_tool_filter()
    if posture.get("kill_switch_active"):
        raise PermissionError("AgentFlow native reads are blocked while the security kill switch is active.")
    allowed, _denied = filter_tools_by_posture([definition], posture)
    if not allowed:
        raise PermissionError(f"Security posture does not allow AgentFlow tool '{tool_name}'.")

    scope = get_toolgate_scope(posture)
    tier = scope["base_tier"]
    mode = "ronin_desktop" if tier == "ronin" else "campaign" if tier == "campaign" else "standard"
    preset_key = posture.get("active_campaign_preset")
    decision = await check_tool_access(
        mode,
        tool_name,
        args,
        get_preset(preset_key) if preset_key else None,
        local_scope=scope["key"],
    )
    if decision.action == GateAction.CONFIRM:
        raise PermissionError(
            f"AgentFlow tool '{tool_name}' requires confirmation and cannot run unattended: {decision.reason}"
        )
    if decision.action == GateAction.BLOCK:
        raise PermissionError(f"AgentFlow tool '{tool_name}' was blocked by ToolGate: {decision.reason}")

    async with async_session_factory() as session:
        raw_result = await execute_native_tool(tool_name, args, session)
    try:
        result = json.loads(raw_result)
    except (TypeError, json.JSONDecodeError) as exc:
        raise ValueError(f"AgentFlow tool '{tool_name}' returned invalid structured output.") from exc
    if not isinstance(result, dict):
        raise ValueError(f"AgentFlow tool '{tool_name}' returned a non-object result.")
    if str(result.get("status", "success")).lower() in {"error", "failed"}:
        raise ValueError(str(result.get("message") or result.get("error") or f"{tool_name} failed"))
    return result


async def _exec_coding(
    config: dict[str, Any],
    context_str: str,
    governance_context: dict[str, Any] | None = None,
) -> Any:
    """Execute a governed, memory-aware coding operation through IDE Mode."""
    from shogun.services.ide_service import ide_service

    action = str(config.get("action") or "analyze")
    workspace_id = str(config.get("workspace_id") or "").strip()
    task = str(config.get("task_description") or "").strip()
    memory: list[dict[str, Any]] = []

    if workspace_id and config.get("recall_memory", True):
        memory = await ide_service.search_programming_memory(
            workspace_id,
            str(config.get("query") or task or context_str or "coding task")[:500],
            limit=int(config.get("memory_limit", 5)),
            include_global=bool(config.get("include_global_memory", False)),
        )

    memory_context = ""
    if memory:
        memory_context = "\n\n--- VERIFIED PROGRAMMING MEMORY ---\n" + "\n\n".join(
            f"{item['title']} [{item['validation_status']}]\n"
            f"Problem: {item['problem']}\nSolution: {item['solution']}"
            for item in memory
        )

    if action == "analyze":
        coding_config = {
            **config,
            "task_description": task or "Analyze the supplied coding task and produce a safe implementation plan.",
            "task_type": "coding_edit",
            "expected_output": config.get("expected_output")
            or "A repository-aware plan with affected files, risks, tests, and verification steps.",
            "requires_tools": False,
        }
        return await _exec_samurai(
            coding_config,
            f"{context_str}{memory_context}",
            governance_context,
        )

    if not workspace_id:
        raise ValueError("Coding node requires an approved IDE workspace for this action.")

    if action == "list_files":
        result: Any = await ide_service.list_files(workspace_id, str(config.get("file_glob") or "*"))
    elif action == "search":
        result = await ide_service.search(
            workspace_id,
            str(config.get("query") or task),
            str(config.get("file_glob") or "*"),
        )
    elif action == "read_file":
        result = await ide_service.read_file(workspace_id, str(config.get("path") or ""))
    elif action == "apply_patch":
        content = str(config.get("content_template") or "").replace("{{context}}", context_str)
        result = await ide_service.write(
            workspace_id,
            str(config.get("path") or ""),
            content,
            approval=bool(config.get("approval", False)),
        )
    elif action == "run_task":
        result = await ide_service.run_command(
            workspace_id,
            str(config.get("command") or ""),
            approval=bool(config.get("approval", False)),
            timeout=int(config.get("timeout", 300)),
        )
        if result.get("exit_code") != 0:
            raise ValueError(
                f"Coding task failed with exit code {result.get('exit_code')}: "
                f"{str(result.get('output') or '')[-2000:]}"
            )
        if config.get("remember_on_success"):
            await ide_service.remember_programming_solution(
                workspace_id,
                {
                    "title": task or f"Verified coding task: {config.get('command')}",
                    "problem": task or "Verify the current implementation.",
                    "solution": str(result.get("output") or "The configured verification command passed."),
                    "evidence": f"{config.get('command')} exited with code 0",
                    "validation_status": "tests_passed",
                    "kind": "solution",
                    "files": [config["path"]] if config.get("path") else [],
                    "tags": ["agentflow-coding-node", "verified"],
                },
            )
    else:
        raise ValueError(f"Unknown Coding node action: {action}")

    return {
        "action": action,
        "task": task,
        "result": result,
        "recalled_memory_ids": [item["id"] for item in memory],
    }


async def _exec_channel_send(config: dict, context_str: str) -> str:
    """Send an AgentFlow message to Telegram."""
    from shogun.services.notification_service import send_channel_message

    channel = "telegram"
    template = config.get("message_template") or "{{context}}"
    message = template.replace("{{context}}", context_str).strip()
    if not message:
        raise ValueError("Channel Send node produced an empty message")

    telegram_chat_ids = config.get("telegram_chat_ids") or None
    if telegram_chat_ids is None and config.get("chat_id") is not None:
        telegram_chat_ids = [str(config["chat_id"])]

    results = await send_channel_message(
        message,
        channel=channel,
        telegram_chat_ids=telegram_chat_ids,
        telegram_message_thread_id=config.get("message_thread_id"),
    )
    selected = [channel]
    failures = [name for name in selected if not results.get(name, {}).get("ok")]
    if failures:
        detail = "; ".join(
            f"{name}: {results.get(name, {}).get('error') or results.get(name, {}).get('errors')}" for name in failures
        )
        raise ValueError(f"Channel delivery failed ({detail})")
    return f"Message delivered via {', '.join(selected)}"


async def _exec_approval(
    config: dict,
    predecessor_outputs: dict[str, Any],
    governance_context: dict[str, Any] | None = None,
) -> str:
    """Shogun Approval node — gate that checks approval policy."""
    approval_mode = config.get("approval_mode", "manual")
    confidence_threshold = config.get("confidence_threshold", 85)

    # Aggregate predecessor output for review
    review_content = "\n\n".join(str(v) for v in predecessor_outputs.values() if v is not None)

    if approval_mode == "manual":
        # In Phase 2, manual approval auto-approves with a note
        # Full human-in-the-loop requires WebSocket (future phase)
        return f"[AUTO-APPROVED — manual approval mode]\n{review_content}"

    elif approval_mode == "ai_assisted":
        # Use LLM to evaluate if the output is acceptable
        async with async_session_factory() as session:
            model_chain, routing_context = await _resolve_task_llm_chain(
                session,
                prompt=review_content[:3000],
                task_type="final_review",
                required_capabilities=["chat"],
                routing_profile_id=(governance_context or {}).get("model_profile"),
            )
        routing_context = _with_flow_generation_settings(routing_context, governance_context)

        if model_chain:
            judge_messages = [
                {
                    "role": "system",
                    "content": (
                        "You are a quality assurance judge reviewing the output of an AI agent. "
                        "Evaluate the following output and respond with APPROVED if it meets quality standards, "
                        "or REJECTED with a brief reason if it does not."
                    ),
                },
                {"role": "user", "content": f"Review this output:\n\n{review_content[:3000]}"},
            ]
            verdict = await _call_llm_chain(
                judge_messages,
                model_chain,
                timeout=60,
                retry_count=0,
                context="AgentFlow approval node",
                routing_context=routing_context,
            )
            if "REJECTED" in verdict.upper():
                raise ValueError(f"AI review rejected: {verdict[:500]}")
            return f"[AI-APPROVED]\n{review_content}"

        return f"[AUTO-APPROVED — no LLM available for AI review]\n{review_content}"

    elif approval_mode == "confidence_threshold":
        # Auto-approve — threshold is informational in Phase 2
        return f"[AUTO-APPROVED — confidence threshold {confidence_threshold}%]\n{review_content}"

    elif approval_mode == "policy_based":
        # Check Torii posture for agentflow_execute permission
        try:
            from shogun.services.posture_guard import get_posture_permissions

            perms = await get_posture_permissions()
            if not perms.get("agentflow_execute", False):
                raise ValueError("Policy-based approval denied: agentflow_execute not permitted at current tier")
        except ImportError:
            pass
        return f"[POLICY-APPROVED]\n{review_content}"

    return f"[APPROVED]\n{review_content}"


async def _exec_logic(
    config: dict,
    predecessor_outputs: dict[str, Any],
    governance_context: dict[str, Any] | None = None,
) -> bool:
    """Logic/Decision node — evaluates condition and returns True (right) or False (bottom)."""
    condition = config.get("condition_expression", "")

    if not condition:
        # No condition — always take the TRUE branch
        return True

    # Build context for evaluation
    context = "\n\n".join(str(v) for v in predecessor_outputs.values() if v is not None)

    # Use LLM to evaluate the condition
    async with async_session_factory() as session:
        model_chain, routing_context = await _resolve_task_llm_chain(
            session,
            prompt=f"{condition}\n\n{context[:3000]}",
            task_type="classification",
            required_capabilities=["chat"],
            routing_profile_id=(governance_context or {}).get("model_profile"),
        )
    routing_context = _with_flow_generation_settings(routing_context, governance_context)

    if not model_chain:
        # No LLM — default to True
        log.warning("Logic node: no LLM available, defaulting to TRUE branch")
        return True

    eval_messages = [
        {
            "role": "system",
            "content": (
                "You are a logic evaluator. Given a condition and context, "
                "evaluate whether the condition is TRUE or FALSE. "
                "Respond with exactly one word: TRUE or FALSE."
            ),
        },
        {
            "role": "user",
            "content": f"Condition: {condition}\n\nContext:\n{context[:3000]}",
        },
    ]

    try:
        result = await _call_llm_chain(
            eval_messages,
            model_chain,
            timeout=30,
            retry_count=0,
            context="AgentFlow logic node",
            routing_context=routing_context,
        )
        return "TRUE" in result.upper()
    except Exception:
        log.warning("Logic node evaluation failed, defaulting to TRUE")
        return True


async def _exec_output(
    config: dict,
    context_str: str,
    predecessor_outputs: dict[str, Any],
    run_id: uuid.UUID | None = None,
    node_label: str = "output",
    node_id: str | None = None,
) -> str:
    """Output node — formats and returns the final result, saving it to workspace."""
    output_type = config.get("output_type", "artifact")
    fmt = config.get("format", "markdown")

    # Collect all predecessor outputs
    content = context_str or "\n\n".join(str(v) for v in predecessor_outputs.values() if v is not None)

    final_content = content
    if fmt == "json":
        import json

        try:
            # Try to parse as JSON, otherwise wrap as JSON
            result = json.loads(content)
            final_content = json.dumps(result, indent=2)
        except (json.JSONDecodeError, TypeError):
            final_content = json.dumps(
                {
                    "output_type": output_type,
                    "content": content,
                },
                indent=2,
            )
    elif fmt == "plain":
        # Strip markdown formatting
        import re

        final_content = re.sub(r"[#*_`~\[\]]", "", content)

    # Never create a misleading zero-byte "successful" report. An empty
    # predecessor result is still a result that the user needs to understand.
    if not final_content or not str(final_content).strip():
        final_content = (
            "No output was produced by the preceding node. "
            "Open the run details to check that node's response and configuration."
        )

    # Save to workspace automatically
    if run_id:
        from shogun.config import settings
        from datetime import datetime
        import logging

        log = logging.getLogger("shogun.flow")
        workspace_dir = settings.workspace_path.resolve() / "output"
        workspace_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_ext = "md" if fmt == "markdown" else "json" if fmt == "json" else "html" if fmt == "html" else "txt"

        # Sanitize node label for filename
        import re

        safe_label = re.sub(r"[^a-zA-Z0-9_-]", "_", node_label).strip("_").lower()
        if not safe_label:
            safe_label = "output"

        short_run_id = str(run_id)[:8]
        filename = f"report_{safe_label}_{timestamp}_{short_run_id}.{file_ext}"

        try:
            target_path = workspace_dir / filename
            temporary_path = target_path.with_suffix(f"{target_path.suffix}.tmp")
            temporary_path.write_text(final_content, encoding="utf-8")
            temporary_path.replace(target_path)
            if node_id:
                artifact_path = str(target_path.relative_to(settings.workspace_path.resolve())).replace("\\", "/")
                await _record_node_artifact(run_id, node_id, artifact_path)
            log.info("Saved flow output report to workspace: %s", target_path)
        except Exception as e:
            log.error("Failed to save flow output report to workspace: %s", e)

    return final_content


async def _exec_mado_browser(
    config: dict,
    context_str: str,
    run_id: uuid.UUID | None = None,
    node_id: str | None = None,
    governance_context: dict[str, Any] | None = None,
) -> str:
    """Mado Browser node — executes browser automation actions.

    Supports: navigate, extract_content, screenshot, fill_form,
              click, execute_js, wait_for
    """
    from fastapi import HTTPException

    from shogun.services import mado_service
    from shogun.services.mado_hardening import governed_action, permission_guard, runtime_registry
    from shogun.services.posture_guard import (
        check_mado_access,
        check_mado_browser_mode,
        check_mado_session_limit,
        get_posture_tool_filter,
    )

    action = config.get("action", "navigate")
    url = config.get("url", "")
    selector = config.get("selector")
    session_name = config.get("session_name", "flow_browser")
    browser_mode = config.get("browser_mode", "headless")

    # Reuse a named browser inside one run, but never inherit its runtime clock
    # or live page state in a later/parallel run. The persistent profile remains
    # stable so sequential executions can retain cookies and authenticated state.
    flow_session_id, profile_name = _flow_mado_session_identity(run_id, session_name)
    governance_context = governance_context or {}

    # Enforce the same local Torii, Harakiri, Gensui, browser-mode, and
    # session-limit rules used by chat and the Mado API.
    try:
        await check_mado_access()
        posture = await get_posture_tool_filter()
        check_mado_browser_mode(browser_mode, posture)
        if flow_session_id not in mado_service._active_contexts:
            await check_mado_session_limit()
    except HTTPException as exc:
        return f"[BLOCKED] {exc.detail}"

    # Ensure browser is launched
    launch_result = await mado_service.launch_browser(
        session_id=flow_session_id,
        profile_name=profile_name,
        mode=browser_mode,
    )

    if launch_result.get("status") == "error":
        return f"[ERROR] Failed to launch browser: {launch_result.get('error', 'Unknown')}"

    if run_id is not None:
        _run_mado_sessions.setdefault(str(run_id), set()).add(flow_session_id)

    runtime_registry.register(
        flow_session_id,
        profile_id=profile_name,
        posture=posture.get("active_tier"),
        mode=browser_mode,
        stack_run_id=governance_context.get("stack_run_id"),
        step_run_id=governance_context.get("step_run_id"),
        agent_id=governance_context.get("agent_id"),
    )

    async def run_governed(action_type: str, operation, verification: dict | None = None):
        result = await governed_action(
            flow_session_id,
            action_type,
            operation,
            detail={"flow_run_id": str(run_id) if run_id else None},
            verification=verification,
        )
        artifact_path = result.get("verification", {}).get("artifact", {}).get("path") or result.get("path")
        if run_id and node_id and artifact_path:
            await _record_node_artifact(run_id, node_id, artifact_path)
        return result

    try:
        if action == "navigate":
            # Use URL from config first; fall back to context string
            target_url = url or (context_str.strip().split("\n")[0] if context_str else "")
            if not target_url:
                return "[ERROR] No URL specified for navigation"

            await permission_guard.check("mado.navigation.open_url", url=target_url)

            log.info("[Mado/Flow] navigate → %s (session=%s)", target_url, flow_session_id)
            result = await run_governed(
                "mado.navigation.open_url",
                lambda: mado_service.navigate(session_id=flow_session_id, url=target_url),
                {"verification_type": "no_error_banner"},
            )
            if result.get("status") == "error":
                log.error("[Mado/Flow] navigate FAILED: %s", result.get("error"))
                return f"[ERROR] Navigation failed: {result.get('error', 'Unknown')}"
            if result.get("status") == "blocked":
                return f"[BLOCKED] {result.get('reason', 'Domain not allowed')}"
            log.info("[Mado/Flow] navigate OK → %s", result.get("title", "N/A"))
            return f"Navigated to: {result.get('url', target_url)}\nTitle: {result.get('title', 'N/A')}"

        elif action == "extract_content":
            extract_type = config.get("extract_type", "text")
            # A source node can be self-contained. Previously an extract node's
            # configured URL was ignored, causing fresh sessions to scrape
            # about:blank and report an empty result.
            if url:
                await permission_guard.check("mado.navigation.open_url", url=url)
                navigation = await run_governed(
                    "mado.navigation.open_url",
                    lambda: mado_service.navigate(session_id=flow_session_id, url=url),
                    {"verification_type": "no_error_banner"},
                )
                if navigation.get("status") == "error":
                    return f"[ERROR] Navigation failed before extraction: {navigation.get('error', 'Unknown')}"
                if navigation.get("status") == "blocked":
                    return f"[BLOCKED] {navigation.get('reason', 'Domain not allowed')}"
            log.info("[Mado/Flow] extract '%s' (type=%s, session=%s)", selector, extract_type, flow_session_id)
            result = await run_governed(
                "mado.page.extract_text",
                lambda: mado_service.extract_content(
                    session_id=flow_session_id, selector=selector, extract_type=extract_type
                ),
            )
            content = result.get("content", "")
            status = result.get("status", "unknown")
            log.info("[Mado/Flow] extract result: status=%s, length=%d chars", status, len(content))
            if not content:
                return "[No content extracted — the page may not have matching elements]"
            return content

        elif action == "screenshot":
            full_page = config.get("full_page", False)
            result = await run_governed(
                "mado.page.screenshot",
                lambda: mado_service.screenshot(session_id=flow_session_id, full_page=full_page, selector=selector),
            )
            return f"Screenshot saved: {result.get('filename', 'unknown')}\nPath: {result.get('path', 'N/A')}"

        elif action == "fill_form":
            fields = config.get("fields", [])
            if not fields:
                return "[ERROR] No form fields specified"
            result = await run_governed(
                "mado.form.fill",
                lambda: mado_service.fill_form(session_id=flow_session_id, fields=fields),
            )
            return f"Filled {result.get('filled', 0)}/{result.get('total', 0)} fields"

        elif action == "click":
            if not selector:
                return "[ERROR] No selector specified for click"
            result = await run_governed(
                "mado.action.click",
                lambda: mado_service.click_element(session_id=flow_session_id, selector=selector),
                config.get("verification"),
            )
            return f"Clicked: {selector}\nURL after click: {result.get('url', 'N/A')}"

        elif action == "execute_js":
            script = config.get("script", "")
            if not script:
                return "[ERROR] No JavaScript specified"
            result = await run_governed(
                "mado.action.execute_js",
                lambda: mado_service.execute_js(session_id=flow_session_id, script=script),
            )
            return f"JS result: {result.get('result', 'undefined')}"

        elif action == "wait_for":
            if not selector:
                return "[ERROR] No selector specified for wait"
            timeout = config.get("timeout", 10000)
            result = await run_governed(
                "mado.action.wait",
                lambda: mado_service.wait_for_selector(session_id=flow_session_id, selector=selector, timeout=timeout),
                {"verification_type": "element_exists", "expected": selector},
            )
            if result.get("status") == "timeout":
                return f"[TIMEOUT] Selector '{selector}' not found within {timeout}ms"
            return f"Element found: {selector}"

        else:
            return f"[ERROR] Unknown Mado action: {action}"

    except Exception as exc:
        return f"[ERROR] Browser action '{action}' failed: {str(exc)[:500]}"


async def _exec_email_send(config: dict, context_str: str) -> str:
    """Email Send node — sends an email via the configured SMTP account.

    Uses the existing EmailService infrastructure.  If no body_template is
    provided the full predecessor output is used as the email body.
    """
    from shogun.services.email_service import EmailService
    from shogun.schemas.channels import EmailComposeRequest

    to_address = config.get("to_address", "")
    if not to_address:
        raise ValueError("Email Send node has no recipient (to_address)")

    subject = config.get("subject", "Shogun Agent Flow — Email")
    body_template = config.get("body_template", "")
    cc_address = config.get("cc_address") or None
    bcc_address = config.get("bcc_address") or None

    # Build the email body
    if body_template:
        # Allow a simple {{context}} placeholder for predecessor output
        body = body_template.replace("{{context}}", context_str)
    else:
        # No template — use the raw predecessor output as the body
        body = context_str or "(No content from previous steps)"

    # Send via EmailService
    async with async_session_factory() as session:
        svc = EmailService(session)
        acc = await svc.get_account()
        if not acc:
            raise ValueError("No email account configured. Set up an account in the Mail page first.")

        compose = EmailComposeRequest(
            to_address=to_address,
            cc_address=cc_address,
            bcc_address=bcc_address,
            subject=subject,
            body=body,
        )
        result = await svc.send_email(compose)

    status = result.get("message", "Sent")
    log.info("Email sent to %s — %s", to_address, status)
    return f"Email sent to {to_address}\nSubject: {subject}\nStatus: {status}"


def _scheduled_output_path(
    target: Path,
    trigger_type: str,
    run_id: uuid.UUID | None,
) -> Path:
    """Give each scheduled run its own timestamped, traceable output file."""
    if trigger_type != "scheduled" or not run_id:
        return target
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    version = f"v{timestamp}_{str(run_id)[:8]}"
    return target.with_name(f"{target.stem}_{version}{target.suffix}")


async def _exec_workspace(
    config: dict,
    context_str: str,
    run_id: uuid.UUID | None = None,
    trigger_type: str = "manual",
) -> str:
    """Workspace node — performs file operations inside the agent workspace.

    Actions: read_file, write_file, list_files, mkdir, delete, copy
    """
    import shutil

    from shogun.config import settings
    from shogun.services.posture_guard import get_posture_permissions

    action = config.get("action", "read_file")
    file_path = config.get("path", "").strip()
    content_template = config.get("content_template", "")

    # Check posture
    perms = await get_posture_permissions()
    if not perms.get("workspace_enabled", True):
        return "[BLOCKED] Workspace access is disabled at current security posture"

    root = settings.workspace_path.resolve()
    root.mkdir(parents=True, exist_ok=True)

    def _safe(rel: str) -> Path:
        """Resolve and validate a relative path within workspace."""
        if not rel:
            return root
        if ".." in rel:
            raise ValueError(f"Path traversal blocked: {rel}")
        target = (root / rel).resolve()
        try:
            target.relative_to(root)
        except ValueError:
            raise ValueError(f"Path escape blocked: {rel}")
        return target

    try:
        if action == "read_file":
            if not file_path:
                return "[ERROR] No path specified for read_file"
            target = _safe(file_path)
            if not target.is_file():
                return f"[ERROR] File not found: {file_path}"
            try:
                content = target.read_text(encoding="utf-8", errors="replace")
            except UnicodeDecodeError:
                return f"[Binary file — {target.stat().st_size} bytes] Cannot read as text: {file_path}"
            log.info("[Flow/Workspace] read_file: %s (%d chars)", file_path, len(content))
            return content

        elif action == "write_file":
            if not file_path:
                return "[ERROR] No path specified for write_file"
            target = _scheduled_output_path(_safe(file_path), trigger_type, run_id)
            target.parent.mkdir(parents=True, exist_ok=True)
            # Content: template with {{context}} or raw predecessor output
            if content_template:
                body = content_template.replace("{{context}}", context_str)
            else:
                body = context_str or ""
            target.write_text(body, encoding="utf-8")
            saved_path = str(target.relative_to(root)).replace("\\", "/")
            log.info("[Flow/Workspace] write_file: %s (%d chars)", saved_path, len(body))
            return f"Written {len(body)} characters to: {saved_path}"

        elif action == "list_files":
            target = _safe(file_path) if file_path else root
            if not target.is_dir():
                return f"[ERROR] Not a directory: {file_path}"
            entries = []
            for item in sorted(target.iterdir()):
                rel = str(item.relative_to(root)).replace("\\", "/")
                if item.is_dir():
                    entries.append(f"📁 {rel}/")
                else:
                    size = item.stat().st_size
                    entries.append(f"📄 {rel} ({size:,} bytes)")
            log.info("[Flow/Workspace] list_files: %s (%d entries)", file_path or ".", len(entries))
            return "\n".join(entries) if entries else "(empty directory)"

        elif action == "mkdir":
            if not file_path:
                return "[ERROR] No path specified for mkdir"
            target = _safe(file_path)
            target.mkdir(parents=True, exist_ok=True)
            log.info("[Flow/Workspace] mkdir: %s", file_path)
            return f"Directory created: {file_path}"

        elif action == "delete":
            if not file_path:
                return "[ERROR] No path specified for delete"
            target = _safe(file_path)
            if not target.exists():
                return f"[ERROR] Not found: {file_path}"
            if target.is_dir():
                shutil.rmtree(target)
            else:
                target.unlink()
            log.info("[Flow/Workspace] delete: %s", file_path)
            return f"Deleted: {file_path}"

        elif action == "copy":
            source_path = config.get("source_path", "").strip()
            dest_path = config.get("dest_path", "").strip()
            if not source_path or not dest_path:
                return "[ERROR] Both source_path and dest_path are required for copy"
            src = _safe(source_path)
            dst = _safe(dest_path)
            if not src.exists():
                return f"[ERROR] Source not found: {source_path}"
            dst.parent.mkdir(parents=True, exist_ok=True)
            if src.is_dir():
                shutil.copytree(src, dst)
            else:
                shutil.copy2(src, dst)
            log.info("[Flow/Workspace] copy: %s → %s", source_path, dest_path)
            return f"Copied: {source_path} → {dest_path}"

        else:
            return f"[ERROR] Unknown workspace action: {action}"

    except Exception as exc:
        return f"[ERROR] Workspace '{action}' failed: {str(exc)[:500]}"


def _excel_rows_from_context(context: str) -> list[list[Any]]:
    """Normalize model output into typed worksheet rows."""
    from shogun.services.file_template import parse_excel_rows

    return parse_excel_rows(context)


async def _resolve_registered_enterprise_profile(profile: Any) -> tuple[dict[str, Any], dict[str, Any]]:
    """Resolve one trusted active registry version for Mapping/RPA execution."""

    from shogun.db.engine import async_session_factory
    from shogun.mapping.errors import MappingSchemaError
    from shogun.services.transformation_profile_registry import (
        TransformationProfileRegistryError,
        TransformationProfileRegistryService,
    )

    try:
        async with async_session_factory() as session:
            service = TransformationProfileRegistryService(session)
            resolved = await service.resolve_active_definition(
                profile.id,
                expected_version=profile.registry_version,
                expected_hash=profile.content_hash.lower() if profile.content_hash else None,
            )
            definition = resolved["definition"]
            evidence = resolved["registry_evidence"]
            if profile.adapter != evidence["adapter_id"]:
                raise MappingSchemaError(
                    f"Transformation profile '{profile.id}' registry adapter is '{evidence['adapter_id']}', "
                    f"not '{profile.adapter}'"
                )
            return definition, evidence
    except MappingSchemaError:
        raise
    except TransformationProfileRegistryError as exc:
        raise MappingSchemaError(str(exc), field="transformation_profile") from exc


async def _resolve_direct_samurai_profile(
    raw_profile: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Resolve and revalidate a profile attached directly to one Samurai.

    Registry profiles must carry an immutable version/hash pin. Portable
    private files carry their complete definition and hash and are validated
    locally. No caller-supplied profile mechanics are executed directly.
    """

    from shogun.mapping.errors import MappingSchemaError
    from shogun.mapping.schema import MappingTransformationProfile
    from shogun.services.enterprise_transformations import CANONICAL_ENTITY_ADAPTER
    from shogun.services.structured_transformations import SUPPORTED_ADAPTER

    profile = MappingTransformationProfile.model_validate(raw_profile)
    if profile.adapter not in {SUPPORTED_ADAPTER, CANONICAL_ENTITY_ADAPTER}:
        raise MappingSchemaError(
            f"Transformation profile adapter '{profile.adapter}' cannot execute on a Samurai node",
            field="transformation_profile.adapter",
        )
    if profile.is_private_file:
        execution_mode = "contract" if profile.adapter == SUPPORTED_ADAPTER else "profile"
        definition, evidence = _resolve_private_transformation_profile(
            profile,
            execution_mode=execution_mode,
        )
        if profile.adapter == SUPPORTED_ADAPTER:
            definition, evidence = _validate_private_contract_resolution(
                profile,
                definition,
                evidence,
            )
    else:
        if not profile.is_registry_pinned:
            raise MappingSchemaError(
                "A Samurai registry profile requires an immutable registry_version/content_hash pin",
                field="transformation_profile",
            )
        definition, evidence = await _resolve_registered_enterprise_profile(profile)
        if profile.adapter == SUPPORTED_ADAPTER:
            definition, evidence = _validate_contract_registry_resolution(
                profile,
                definition,
                evidence,
            )
        else:
            # The registry resolver already enforced the immutable pin. The
            # canonical adapter repeats manifest/hash/evidence validation with
            # the actual payload immediately before producing output.
            if (
                str(evidence.get("content_hash") or "").lower() != profile.content_hash
                or evidence.get("version") != profile.registry_version
            ):
                raise MappingSchemaError(
                    "Resolved enterprise transformation profile does not match its Samurai pin",
                    field="transformation_profile",
                )
    return deepcopy(definition), {
        **deepcopy(evidence),
        "selection_mode": "profile",
        "profile_source": "private" if profile.is_private_file else "registry",
    }


def _bounded_source_intelligence_text(value: str, *, limit: int = 2_097_152) -> str:
    """Keep representative beginning/middle/end source text within the resolver bound."""

    if len(value) <= limit:
        return value
    marker = "\n\n[... SOURCE INTELLIGENCE SAMPLE BOUNDARY ...]\n\n"
    available = max(3, limit - (2 * len(marker)))
    segment = available // 3
    midpoint = len(value) // 2
    middle_start = max(0, midpoint - (segment // 2))
    return (
        value[:segment]
        + marker
        + value[middle_start : middle_start + segment]
        + marker
        + value[-segment:]
    )[:limit]


def _source_intelligence_artifacts(source_inputs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Convert raw predecessor results into bounded Source Intelligence inputs."""

    artifacts: list[dict[str, Any]] = []
    for source in source_inputs:
        raw_output = source.get("raw_output")
        artifact: dict[str, Any] = {
            "source_id": str(source.get("source_id") or ""),
            "label": str(source.get("label") or "") or None,
            "context": {
                key: value
                for key, value in dict(source.get("context") or {}).items()
                if value not in (None, "")
            },
        }
        if isinstance(raw_output, (dict, list)):
            artifact["payload"] = deepcopy(raw_output)
        else:
            artifact["text"] = _bounded_source_intelligence_text(
                str(source.get("text_output") or raw_output or "")
            )
        artifacts.append(artifact)
    return artifacts


def _safe_source_intelligence_evidence(result: Any) -> dict[str, Any]:
    """Retain match decisions without persisting source excerpts or profile mechanics."""

    candidates = []
    for candidate in list(getattr(result, "candidates", None) or []):
        candidates.append(
            {
                "profile_id": candidate.profile_id,
                "profile_source": candidate.profile_source,
                "platform": candidate.platform,
                "domain": candidate.domain,
                "adapter_id": candidate.adapter_id,
                "version": candidate.version,
                "content_hash": candidate.content_hash,
                "score": candidate.score,
                "exact": candidate.exact,
                "specialist_skill": candidate.specialist_skill,
                "matched": list(candidate.evidence.matched),
                "missing": list(candidate.evidence.missing),
                "negative_matches": list(candidate.evidence.negative_matches),
            }
        )
    classifier_request = getattr(result, "classifier_request", None)
    return {
        "outcome": str(getattr(result, "outcome", "unknown")),
        "execution_allowed": bool(getattr(result, "execution_allowed", False)),
        "specialist_skill": str(
            getattr(result, "specialist_skill", "enterprise-transformation-architect")
        ),
        "candidates": candidates,
        "semantic_candidate_profile_ids": list(
            getattr(classifier_request, "allowed_profile_ids", None) or []
        ),
    }


def _json_object_from_model_response(value: str) -> dict[str, Any]:
    """Decode one strict JSON object, tolerating only an outer Markdown fence."""

    text_value = str(value or "").strip()
    fenced = re.fullmatch(r"```(?:json)?\s*(.*?)\s*```", text_value, re.DOTALL | re.IGNORECASE)
    if fenced:
        text_value = fenced.group(1).strip()
    try:
        parsed = json.loads(text_value)
    except json.JSONDecodeError as exc:
        raise ValueError("Source classifier returned invalid JSON.") from exc
    if not isinstance(parsed, dict):
        raise ValueError("Source classifier must return one JSON object.")
    return parsed


async def _run_source_semantic_classifier(
    classifier_request: Any,
    *,
    config: dict[str, Any],
    governance_context: dict[str, Any],
) -> Any:
    """Run one governed, bounded advisory classifier call."""

    from shogun.schemas.source_intelligence import SemanticClassifierResponse

    request_payload = classifier_request.model_dump(mode="json")
    response_schema = SemanticClassifierResponse.model_json_schema()
    prompt = (
        "Classify this bounded enterprise source summary. You are advisory only: nominate only profile "
        "IDs and specialist skills in the supplied allow-lists. Never invent a profile, parser, mapping, "
        "or executable rule. If the evidence is insufficient, return classification='unknown'. Return "
        "exactly one JSON object matching the response schema and no Markdown.\n\n"
        "SOURCE CLASSIFIER REQUEST:\n"
        + json.dumps(request_payload, ensure_ascii=False, separators=(",", ":"))
        + "\n\nRESPONSE JSON SCHEMA:\n"
        + json.dumps(response_schema, ensure_ascii=False, separators=(",", ":"))
    )
    async with async_session_factory() as session:
        model_chain, routing = await _resolve_task_llm_chain(
            session,
            prompt=prompt,
            task_type="classification",
            required_capabilities=["chat"],
            routing_profile_id=(governance_context or {}).get("model_profile")
            or config.get("routing_profile_id"),
            run_id=config.get("_run_id"),
            risk_level=str((governance_context or {}).get("risk_tier") or "low"),
            context_size_estimate=max(1, len(prompt) // 4),
        )
        if not model_chain:
            raise ValueError("No governed model is available for Source Intelligence classification.")
        routing = _with_flow_generation_settings(routing, governance_context)
        response_text = await _call_llm_chain(
            [
                {
                    "role": "system",
                    "content": (
                        "You classify enterprise input structures. Output strict JSON only. "
                        "Your nomination is advisory and cannot create execution authority."
                    ),
                },
                {"role": "user", "content": prompt},
            ],
            model_chain,
            timeout=max(30, min(int(config.get("source_classifier_timeout") or 90), 180)),
            retry_count=0,
            context="AgentFlow Samurai Source Intelligence classifier",
            max_tokens=2048,
            routing_context=routing,
            usage_session=session,
        )
        await session.commit()
    return SemanticClassifierResponse.model_validate(
        _json_object_from_model_response(response_text)
    )


async def _resolve_auto_samurai_profile(
    *,
    source_inputs: list[dict[str, Any]],
    private_profiles: list[Any],
    config: dict[str, Any],
    governance_context: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Resolve an exact known profile or gate one bounded semantic nomination."""

    from shogun.services.source_intelligence import (
        SourceIntelligenceError,
        SourceIntelligenceService,
        SourceProfileAmbiguousError,
        SourceProfileUnknownError,
    )

    artifacts = _source_intelligence_artifacts(source_inputs)
    if not artifacts:
        raise SourceIntelligenceResolutionError(
            "Samurai auto-detect requires at least one non-template source input.",
            {
                "outcome": "unknown",
                "execution_allowed": False,
                "specialist_skill": "enterprise-transformation-architect",
                "candidates": [],
            },
        )
    try:
        async with async_session_factory() as session:
            service = SourceIntelligenceService(session)
            try:
                resolved = await service.resolve_executable(
                    artifacts,
                    private_profiles=private_profiles,
                )
            except (SourceProfileUnknownError, SourceProfileAmbiguousError) as initial_error:
                result = initial_error.result
                safe_evidence = _safe_source_intelligence_evidence(result)
                classifier_request = result.classifier_request
                if classifier_request is None:
                    candidate_ids = ", ".join(
                        candidate["profile_id"] for candidate in safe_evidence["candidates"][:5]
                    ) or "none"
                    raise SourceIntelligenceResolutionError(
                        f"Source Intelligence outcome is {result.outcome}; no profile may execute. "
                        f"Suggested specialist: {result.specialist_skill}. Candidates: {candidate_ids}.",
                        safe_evidence,
                    ) from initial_error

                classification = await _run_source_semantic_classifier(
                    classifier_request,
                    config=config,
                    governance_context=governance_context,
                )
                semantic_evidence = {
                    "classification": classification.classification,
                    "platform_family": classification.platform_family,
                    "product": classification.product,
                    "business_object": classification.business_object,
                    "candidate_profile_ids": list(classification.candidate_profile_ids),
                    "specialist_skill": classification.specialist_skill,
                    "confidence": classification.confidence,
                    "unknowns": list(classification.unknowns),
                }
                if (
                    classification.classification != "classified"
                    or len(classification.candidate_profile_ids) != 1
                ):
                    raise SourceIntelligenceResolutionError(
                        "Source Intelligence classified the source family but no single installed "
                        f"profile is authorized to execute. Suggested specialist: "
                        f"{classification.specialist_skill}.",
                        {
                            **safe_evidence,
                            "semantic_classification": semantic_evidence,
                        },
                    )
                nomination = await service.resolve_semantic_nomination(
                    classifier_request,
                    classification,
                    private_profiles=private_profiles,
                )
                return deepcopy(nomination.definition), {
                    **deepcopy(nomination.evidence),
                    "selection_mode": "auto_semantic_nomination",
                    "profile_source": "registry",
                    "source_intelligence": {
                        **safe_evidence,
                        "semantic_classification": semantic_evidence,
                        "requires_deterministic_validation": True,
                    },
                }

            return deepcopy(resolved.definition), {
                **deepcopy(resolved.evidence),
                "selection_mode": "auto_exact",
                "profile_source": (
                    resolved.resolution.selected_profile.profile_source
                    if resolved.resolution.selected_profile
                    else "registry"
                ),
                "source_intelligence": _safe_source_intelligence_evidence(
                    resolved.resolution
                ),
            }
    except SourceIntelligenceResolutionError:
        raise
    except (SourceIntelligenceError, ValueError) as exc:
        evidence = deepcopy(getattr(exc, "source_intelligence", None) or {})
        if not evidence:
            result = getattr(exc, "result", None)
            evidence = (
                _safe_source_intelligence_evidence(result)
                if result is not None
                else {
                    "outcome": "unknown",
                    "execution_allowed": False,
                    "specialist_skill": "enterprise-transformation-architect",
                    "candidates": [],
                }
            )
        raise SourceIntelligenceResolutionError(
            f"Source Intelligence could not authorize a transformation profile: {exc}",
            evidence,
        ) from exc


async def _exec_samurai_enterprise_profile(
    config: dict[str, Any],
    profile: dict[str, Any],
    evidence: dict[str, Any],
    source_inputs: list[dict[str, Any]],
) -> dict[str, Any]:
    """Execute a canonical ingress profile selected directly or by auto mode."""

    from shogun.services.enterprise_transformations import execute_enterprise_profile

    if len(source_inputs) != 1:
        raise ValueError(
            "A canonical enterprise transformation profile requires exactly one structured source input."
        )
    raw_payload = source_inputs[0].get("raw_output")
    if not isinstance(raw_payload, (dict, list, str, bytes)):
        raise ValueError(
            "A canonical enterprise transformation profile requires a JSON-compatible structured payload."
        )
    result = execute_enterprise_profile(
        profile,
        raw_payload,
        context={
            "flow_id": str(config.get("_flow_id") or ""),
            "node_id": str(config.get("_node_id") or ""),
            "source_node_id": str(source_inputs[0].get("source_id") or ""),
        },
        registry_evidence=evidence,
    )
    await _record_samurai_transformation_profile_event(
        config,
        str(profile.get("id") or ""),
        str(profile.get("adapter") or ""),
        int(result.get("records_written") or 0),
        evidence,
    )
    return result


def _resolve_private_transformation_profile(
    profile: Any,
    *,
    execution_mode: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Revalidate one flow-local private file and issue execution evidence."""

    from shogun.mapping.errors import MappingSchemaError
    from shogun.services.private_transformation_profiles import (
        PrivateTransformationProfileError,
        PrivateTransformationProfileService,
    )

    if execution_mode not in {"contract", "profile"}:
        raise MappingSchemaError(
            f"Unsupported private transformation execution mode '{execution_mode}'",
            field="transformation_profile",
        )
    try:
        return PrivateTransformationProfileService().resolve_reference(
            profile,
            execution_mode=execution_mode,
        )
    except PrivateTransformationProfileError as exc:
        raise MappingSchemaError(str(exc), field="transformation_profile") from exc


def _normalized_contract_snapshot(value: Any) -> dict[str, Any]:
    """Normalize legacy inline mechanics for comparison with a registry definition."""

    from shogun.mapping.schema import MappingTransformationProfile

    raw = value.model_dump(mode="json") if hasattr(value, "model_dump") else value
    return MappingTransformationProfile.model_validate(raw).model_dump(
        mode="json",
        exclude={"registry_version", "content_hash"},
        exclude_none=True,
    )


def _validate_contract_registry_resolution(
    configured_profile: Any,
    definition: Any,
    evidence: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Validate trusted registry output before it can become a Samurai contract."""

    from shogun.mapping.errors import MappingSchemaError
    from shogun.services.structured_transformations import SUPPORTED_ADAPTER
    from shogun.services.transformation_profile_registry import profile_content_hash

    if not isinstance(definition, dict) or not isinstance(evidence, dict):
        raise MappingSchemaError(
            "Transformation profile registry returned an invalid contract envelope",
            field="transformation_profile",
        )

    profile_id = str(evidence.get("profile_id") or "")
    adapter_id = str(evidence.get("adapter_id") or "")
    status = str(evidence.get("status") or "").lower()
    adapter_status = str(evidence.get("adapter_status") or "").lower()
    version = evidence.get("version")
    content_hash = str(evidence.get("content_hash") or "").lower()
    if profile_id != configured_profile.id or adapter_id != configured_profile.adapter:
        raise MappingSchemaError(
            "Transformation profile registry evidence does not match the AgentFlow contract reference",
            field="transformation_profile",
        )
    if status != "active" or adapter_status != "available":
        raise MappingSchemaError(
            "Transformation profile contract is not active with an available adapter",
            field="transformation_profile",
        )
    if isinstance(version, bool) or not isinstance(version, int) or version < 1:
        raise MappingSchemaError(
            "Transformation profile registry version is invalid",
            field="transformation_profile.registry_version",
        )
    if definition.get("id") != profile_id or definition.get("adapter") != adapter_id:
        raise MappingSchemaError(
            "Resolved transformation profile definition does not match its registry evidence",
            field="transformation_profile",
        )
    try:
        actual_hash = profile_content_hash(definition)
    except (TypeError, ValueError) as exc:
        raise MappingSchemaError(
            "Resolved transformation profile definition is not canonical JSON",
            field="transformation_profile",
        ) from exc
    if actual_hash != content_hash:
        raise MappingSchemaError(
            "Resolved transformation profile definition failed its content-hash check",
            field="transformation_profile.content_hash",
        )
    if configured_profile.is_registry_pinned:
        if version != configured_profile.registry_version or content_hash != configured_profile.content_hash:
            raise MappingSchemaError(
                "Resolved transformation profile does not match the AgentFlow version/hash pin",
                field="transformation_profile",
            )
    elif _normalized_contract_snapshot(configured_profile) != _normalized_contract_snapshot(definition):
        # Backward compatibility is limited to exact snapshots of the active
        # registry definition. Caller lifecycle/status fields cannot create
        # trust and caller mechanics are never passed to Samurai.
        raise MappingSchemaError(
            "Unpinned inline transformation contract does not exactly match the active registry definition",
            field="transformation_profile",
        )
    if adapter_id != SUPPORTED_ADAPTER:
        raise MappingSchemaError(
            f"Transformation profile adapter '{adapter_id}' cannot execute as a Samurai contract",
            field="transformation_profile.adapter",
        )
    return deepcopy(definition), deepcopy(evidence)


def _validate_private_contract_resolution(
    configured_profile: Any,
    definition: Any,
    evidence: Any,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Validate a private contract carrier without consulting the registry."""

    from shogun.mapping.errors import MappingSchemaError
    from shogun.services.private_transformation_profiles import (
        PRIVATE_PROFILE_EVIDENCE_STATUS,
    )
    from shogun.services.structured_transformations import SUPPORTED_ADAPTER
    from shogun.services.transformation_profile_registry import profile_content_hash

    private_file = getattr(configured_profile, "private_file", None)
    if private_file is None or not isinstance(definition, dict) or not isinstance(evidence, dict):
        raise MappingSchemaError(
            "Private transformation profile carrier is invalid",
            field="transformation_profile",
        )
    profile_id = str(evidence.get("profile_id") or "")
    adapter_id = str(evidence.get("adapter_id") or "")
    status = str(evidence.get("status") or "").lower()
    adapter_status = str(evidence.get("adapter_status") or "").lower()
    source = str(evidence.get("source") or "").lower()
    version = evidence.get("version")
    content_hash = str(evidence.get("content_hash") or "").lower()
    if profile_id != configured_profile.id or adapter_id != configured_profile.adapter:
        raise MappingSchemaError(
            "Private transformation profile evidence does not match the AgentFlow reference",
            field="transformation_profile",
        )
    if (
        status != PRIVATE_PROFILE_EVIDENCE_STATUS
        or adapter_status != "available"
        or source != "private_file"
        or evidence.get("server_validated") is not True
    ):
        raise MappingSchemaError(
            "Private transformation profile lacks server-validated execution evidence",
            field="transformation_profile",
        )
    if isinstance(version, bool) or not isinstance(version, int) or version < 1:
        raise MappingSchemaError(
            "Private transformation profile version is invalid",
            field="transformation_profile.private_file.format_version",
        )
    if definition.get("id") != profile_id or definition.get("adapter") != adapter_id:
        raise MappingSchemaError(
            "Resolved private transformation profile does not match its evidence",
            field="transformation_profile",
        )
    try:
        actual_hash = profile_content_hash(definition)
    except (TypeError, ValueError) as exc:
        raise MappingSchemaError(
            "Resolved private transformation profile is not canonical JSON",
            field="transformation_profile",
        ) from exc
    if (
        actual_hash != content_hash
        or content_hash != private_file.content_hash
    ):
        raise MappingSchemaError(
            "Resolved private transformation profile failed its content-hash pin",
            field="transformation_profile.private_file.content_hash",
        )
    if adapter_id != SUPPORTED_ADAPTER:
        raise MappingSchemaError(
            f"Private transformation profile adapter '{adapter_id}' cannot execute as a Samurai contract",
            field="transformation_profile.adapter",
        )
    return deepcopy(definition), deepcopy(evidence)


def _trusted_contract_profile_from_carrier(
    configured_profile: Any,
    carrier: Any,
    *,
    carrier_label: str,
) -> dict[str, Any]:
    """Extract only a registry-resolved profile from an internal contract carrier."""

    if not (
        isinstance(carrier, dict)
        and carrier.get(_MAPPING_PROFILE_CARRIER_MARKER) is True
        and carrier.get("status") == "SUCCESS"
        and carrier.get("type") == "transformation_profile"
    ):
        raise ValueError(
            f"Mapping/RPA contract carrier '{carrier_label}' did not provide a successful "
            "transformation-profile contract."
        )
    if getattr(configured_profile, "is_private_file", False):
        definition, evidence = _validate_private_contract_resolution(
            configured_profile,
            carrier.get("resolved_definition"),
            carrier.get("registry_evidence"),
        )
    else:
        definition, evidence = _validate_contract_registry_resolution(
            configured_profile,
            carrier.get("resolved_definition"),
            carrier.get("registry_evidence"),
        )
    if (
        carrier.get("profile_id") != evidence["profile_id"]
        or carrier.get("adapter") != evidence["adapter_id"]
        or carrier.get("registry_version") != evidence["version"]
        or str(carrier.get("content_hash") or "").lower() != evidence["content_hash"]
    ):
        raise ValueError(
            f"Mapping/RPA contract carrier '{carrier_label}' metadata does not match its registry evidence."
        )
    return definition


async def _exec_mapping_rpa(
    config: dict[str, Any],
    predecessor_outputs: dict[str, Any],
    *,
    flow_id: str,
    node_id: str,
) -> dict[str, Any]:
    """Execute the deterministic mapping node against one selected predecessor."""
    from shogun.mapping.engine import execute_mapping
    from shogun.mapping.errors import MappingError, MappingInputError
    from shogun.mapping.schema import MappingConfig
    from shogun.services.enterprise_transformations import execute_enterprise_profile

    mapping_config = MappingConfig.model_validate(config)
    if mapping_config.execution_mode == "contract":
        profile = mapping_config.transformation_profile
        if profile is None:  # Kept defensive for direct, non-Pydantic callers.
            raise MappingInputError("Mapping / RPA contract mode requires a transformation profile")
        if profile.is_private_file:
            definition, registry_evidence = _resolve_private_transformation_profile(
                profile,
                execution_mode="contract",
            )
            definition, registry_evidence = _validate_private_contract_resolution(
                profile,
                definition,
                registry_evidence,
            )
        else:
            definition, registry_evidence = await _resolve_registered_enterprise_profile(profile)
            definition, registry_evidence = _validate_contract_registry_resolution(
                profile,
                definition,
                registry_evidence,
            )
        return {
            _MAPPING_PROFILE_CARRIER_MARKER: True,
            "status": "SUCCESS",
            "type": "transformation_profile",
            "profile_id": registry_evidence["profile_id"],
            "adapter": registry_evidence["adapter_id"],
            "registry_version": registry_evidence["version"],
            "content_hash": registry_evidence["content_hash"],
            "resolved_definition": definition,
            "registry_evidence": registry_evidence,
        }

    selected_id = mapping_config.input_source_node_id
    if selected_id:
        if selected_id not in predecessor_outputs:
            raise MappingInputError(
                f'Configured input source node "{selected_id}" is not connected',
                source=selected_id,
            )
        payload = predecessor_outputs[selected_id]
    else:
        available = [(key, value) for key, value in predecessor_outputs.items() if value is not None]
        if len(available) != 1:
            raise MappingInputError(
                "Mapping / RPA requires exactly one predecessor unless input_source_node_id is configured",
                received={"predecessor_count": len(available)},
            )
        selected_id, payload = available[0]
    execution_context = {"flow_id": flow_id, "node_id": node_id, "source_node_id": selected_id}
    try:
        if mapping_config.execution_mode == "profile":
            profile = mapping_config.transformation_profile
            if profile is None:  # Kept defensive for direct, non-Pydantic callers.
                raise MappingInputError("Mapping / RPA profile mode requires a transformation profile")
            if profile.is_private_file:
                definition, registry_evidence = _resolve_private_transformation_profile(
                    profile,
                    execution_mode="profile",
                )
            else:
                definition, registry_evidence = await _resolve_registered_enterprise_profile(profile)
            result = execute_enterprise_profile(
                definition,
                payload,
                context=execution_context,
                registry_evidence=registry_evidence,
            )
            result.update(
                {
                    "type": mapping_config.output.type,
                    "start_cell": mapping_config.output.start_cell,
                    "sheet": mapping_config.output.sheet,
                    "include_headers": mapping_config.output.include_headers,
                }
            )
            return result
        return execute_mapping(payload, mapping_config, context=execution_context)
    except MappingError as exc:
        if not mapping_config.route_failures:
            raise
        log.warning("mapping_validation_failed node_id=%s error=%s", node_id, exc)
        return {
            "__shogun_mapping_output__": True,
            "status": exc.code,
            "type": mapping_config.output.type,
            "records_received": 0,
            "records_written": 0,
            "records_failed": 1,
            "errors": [exc.as_dict()],
            "mapping": {
                "name": mapping_config.name,
                "version": mapping_config.version,
                "mode": mapping_config.mode,
                "execution_mode": mapping_config.execution_mode,
                "profile_id": (
                    mapping_config.transformation_profile.id
                    if mapping_config.transformation_profile is not None
                    else None
                ),
            },
        }


def _mapping_payload_from_predecessors(predecessor_outputs: dict[str, Any] | None) -> dict[str, Any] | None:
    payloads = [
        value for value in (predecessor_outputs or {}).values()
        if isinstance(value, dict) and value.get("__shogun_mapping_output__")
    ]
    if len(payloads) > 1:
        raise ValueError("Files node received multiple Mapping / RPA payloads; connect exactly one")
    if not payloads:
        return None
    payload = payloads[0]
    if payload.get("status") not in {"SUCCESS", "PARTIAL"}:
        first_error = (payload.get("errors") or [{}])[0]
        raise ValueError(
            f"Mapping output is not writable ({payload.get('status')}): "
            f"{first_error.get('message', 'validation failed')}"
        )
    return payload


def _safe_spreadsheet_value(value: Any) -> Any:
    """Reject unsupported cells and neutralize spreadsheet formula injection."""

    from datetime import date, datetime
    from decimal import Decimal

    if value is None or isinstance(value, (bool, int, float, Decimal, date, datetime)):
        return value
    if not isinstance(value, str):
        raise ValueError(
            f"Spreadsheet cells must be scalar values; received {type(value).__name__}"
        )
    if re.search(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", value):
        raise ValueError("Spreadsheet text contains an illegal control character")
    if len(value) > 32_767:
        raise ValueError("Spreadsheet text exceeds Excel's 32,767-character cell limit")
    if value.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _mapping_table_rows(payload: dict[str, Any]) -> list[list[Any]]:
    rows = [list(row) for row in (payload.get("rows") or [])]
    if payload.get("include_headers"):
        headers = payload.get("headers") or []
        if not isinstance(headers, list) or not headers:
            raise ValueError("Mapping output requested headers but did not provide a header list")
        rows.insert(0, list(headers))
    return [[_safe_spreadsheet_value(value) for value in row] for row in rows]


def _write_openpyxl_mapping(ws: Any, payload: dict[str, Any], fallback_start: str = "A1") -> int:
    """Write a typed mapping envelope to an openpyxl worksheet."""
    from openpyxl.utils.cell import coordinate_to_tuple

    if payload.get("type") == "cells":
        for cell, value in (payload.get("cells") or {}).items():
            ws[str(cell).upper()] = _safe_spreadsheet_value(value)
        return len(payload.get("cells") or {})
    data_rows = payload.get("rows") or []
    rows = _mapping_table_rows(payload)
    start = str(payload.get("start_cell") or fallback_start or "A1").upper()
    start_row, start_column = coordinate_to_tuple(start)
    for row_offset, row in enumerate(rows):
        for column_offset, value in enumerate(row):
            ws.cell(row=start_row + row_offset, column=start_column + column_offset, value=value)
    return len(data_rows)


async def _exec_office(
    config: dict,
    context_str: str,
    run_id: uuid.UUID | None = None,
    node_id: str | None = None,
    trigger_type: str = "manual",
    template_inputs: list[dict[str, Any]] | None = None,
    predecessor_outputs: dict[str, Any] | None = None,
) -> str:
    """Files node — reads PDFs and performs Office document operations.

    Actions: pdf_read, excel_read, excel_write, excel_create, word_read, word_replace,
             word_create, pptx_read, pptx_replace
    """
    from shogun.config import settings
    from shogun.office.config import load_office_config

    action = config.get("action", "word_read")
    input_path = config.get("input_path", "").strip()
    output_path = config.get("output_path", "").strip()
    sheet_name = config.get("sheet_name", "Sheet1")

    root = settings.workspace_path.resolve()
    root.mkdir(parents=True, exist_ok=True)
    templates = [item for item in (template_inputs or []) if item.get("__shogun_file_template__")]
    mapping_payload = _mapping_payload_from_predecessors(predecessor_outputs)

    def _create_template(expected_format: str) -> dict[str, Any] | None:
        if len(templates) > 1:
            raise ValueError(
                "This Files create node has multiple upstream File Template nodes. "
                "Connect exactly one template to the Samurai that feeds this node."
            )
        if not templates:
            return None
        template = templates[0]
        if template.get("format") != expected_format:
            raise ValueError(
                f"The upstream {template.get('format', 'unknown')} template does not match "
                f"this {expected_format} create action."
            )
        return template

    def _resolve(rel: str) -> str:
        """Resolve relative workspace path to absolute string."""
        if not rel:
            raise ValueError("Path is required")
        if ".." in rel:
            raise ValueError(f"Path traversal blocked: {rel}")
        target = (root / rel).resolve()
        try:
            target.relative_to(root)
        except ValueError:
            raise ValueError(f"Path escape blocked: {rel}")
        return str(target)

    def _resolve_output(rel: str, suffix: str, fallback_name: str) -> str:
        """Resolve the UI's destination-folder plus filename contract."""
        target = Path(_resolve(rel))
        if target.suffix.lower() != suffix:
            configured_name = str(config.get("output_filename") or fallback_name).strip()
            filename = Path(configured_name).name
            if not filename.lower().endswith(suffix):
                filename = f"{filename}{suffix}"
            target = target / filename
        target = _scheduled_output_path(target, trigger_type, run_id)
        return str(target)

    async def _record_output(abs_path: str) -> str:
        target = Path(abs_path)
        relative = str(target.relative_to(root)).replace("\\", "/")
        if run_id and node_id:
            await _record_node_artifact(run_id, node_id, relative)
        return relative

    try:
        # PDF reading is a native file operation. It deliberately does not
        # depend on Office App Mode or an interactive desktop application, so
        # it remains available to scheduled AgentFlows.
        if action == "pdf_read":
            from shogun.services.file_formats import FileFormatService

            abs_path = _resolve(input_path)
            start_page = max(1, int(config.get("start_page") or 1))
            configured_end = config.get("end_page")
            end_page = max(start_page, int(configured_end)) if configured_end else None
            payload = await FileFormatService(allowed_roots=[root]).read(
                path=abs_path,
                start=start_page,
                end=end_page,
                max_chars=settings.agent_flow_document_max_chars,
            )
            filename = str(payload.get("filename") or Path(abs_path).name)
            content = str(payload.get("content") or "")
            if not content.strip():
                return f"[ERROR] PDF '{filename}' contained no readable text."
            if payload.get("truncated"):
                return (
                    f"[ERROR] PDF '{filename}' exceeds the AgentFlow extraction safety limit of "
                    f"{settings.agent_flow_document_max_chars:,} characters. Select a smaller page range "
                    "or increase SHOGUN_AGENT_FLOW_DOCUMENT_MAX_CHARS."
                )
            metadata = payload.get("metadata") or {}
            warnings = payload.get("warnings") or []
            page_range = (
                f"pages {metadata.get('start_page', start_page)}-"
                f"{metadata.get('end_page', end_page or 'end')}"
            )
            suffix = f"\n\n[Read warnings: {'; '.join(str(item) for item in warnings)}]" if warnings else ""
            log.info("[Flow/Files] pdf_read: %s, %s, chars=%d", input_path, page_range, len(content))
            return f"[PDF: {filename}; {page_range}]\n{content}{suffix}"

        # Office formats still require Office App Mode.
        office_cfg = load_office_config()
        if not office_cfg.enabled:
            return "[BLOCKED] Office App Mode is disabled. Enable it in the Katana settings."

        # ── Excel Operations ──
        if action == "excel_read":
            from shogun.office.adapters.excel_adapter import (
                open_workbook,
                read_used_range,
                list_sheets,
                close_workbook,
            )

            abs_path = _resolve(input_path)
            handle = open_workbook(abs_path)
            try:
                target_sheet = sheet_name or list_sheets(handle)[0]
                data = read_used_range(handle, target_sheet)
                # Format as CSV-like text
                lines = []
                for row in data:
                    lines.append("\t".join(str(c) if c is not None else "" for c in row))
                result = f"[Sheet: {target_sheet}] ({len(data)} rows)\n" + "\n".join(lines)
                log.info("[Flow/Office] excel_read: %s, sheet=%s, rows=%d", input_path, target_sheet, len(data))
                return result
            finally:
                close_workbook(handle)

        elif action == "excel_create":
            abs_out = _resolve_output(output_path, ".xlsx", "output.xlsx")
            template = _create_template("xlsx")
            if template and mapping_payload:
                import shutil

                import openpyxl

                from shogun.services.file_template import resolve_workspace_template

                source = resolve_workspace_template(str(template.get("template_path") or ""), root)
                Path(abs_out).parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(source, abs_out)
                wb = openpyxl.load_workbook(abs_out)
                try:
                    target_sheet = str(mapping_payload.get("sheet") or sheet_name or wb.sheetnames[0])
                    if target_sheet not in wb.sheetnames:
                        raise ValueError(f'Excel template does not contain sheet "{target_sheet}"')
                    written = _write_openpyxl_mapping(
                        wb[target_sheet], mapping_payload, str(config.get("start_range") or "A1")
                    )
                    wb.save(abs_out)
                finally:
                    wb.close()
                row_summary = f"{written} mapped item(s) populated from template"
            elif template:
                from shogun.services.file_template import render_excel_template, resolve_workspace_template

                source = resolve_workspace_template(str(template.get("template_path") or ""), root)
                rows_written = render_excel_template(
                    source,
                    Path(abs_out),
                    context_str,
                    str(template.get("example_handling") or "replace"),
                    sheet_name or None,
                    str(template.get("data_start_cell") or "") or None,
                    str(template.get("render_mode") or "adaptive"),
                    str(template.get("guidance_mode") or "structure_only"),
                    template.get("merge_key_columns"),
                    template.get("merge_preserve_columns"),
                )
                row_summary = f"{rows_written} cells populated from template"
            else:
                import openpyxl

                from shogun.office.adapters.excel_adapter import log_excel_payload_shape

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = str((mapping_payload or {}).get("sheet") or sheet_name or "Sheet1")
                rows = (mapping_payload or {}).get("rows") or _excel_rows_from_context(context_str)
                if rows:
                    from openpyxl.utils import get_column_letter

                    width = max(len(row) for row in rows)
                    destination = f"A1:{get_column_letter(width)}{len(rows)}"
                    log_excel_payload_shape("excel_create", ws.title, destination, rows)
                if mapping_payload:
                    written = _write_openpyxl_mapping(
                        ws, mapping_payload, str(config.get("start_range") or "A1")
                    )
                else:
                    for r_idx, cells in enumerate(rows, 1):
                        for c_idx, val in enumerate(cells, 1):
                            ws.cell(row=r_idx, column=c_idx, value=_safe_spreadsheet_value(val))
                    written = len(rows)
                Path(abs_out).parent.mkdir(parents=True, exist_ok=True)
                wb.save(abs_out)
                wb.close()
                unit = "cells" if mapping_payload and mapping_payload.get("type") == "cells" else "rows"
                row_summary = f"{written} {unit}"
            saved_path = await _record_output(abs_out)
            log.info("[Flow/Office] excel_create: %s (%s)", saved_path, row_summary)
            return f"Excel workbook created: {saved_path} ({row_summary})"

        elif action == "excel_write":
            from shogun.office.adapters.excel_adapter import (
                open_workbook,
                write_range,
                save_as,
                close_workbook,
            )

            abs_in = _resolve(input_path)
            abs_out = (
                _resolve_output(output_path, ".xlsx", Path(abs_in).name)
                if output_path
                else str(_scheduled_output_path(Path(abs_in), trigger_type, run_id))
            )
            handle = open_workbook(abs_in)
            try:
                data = (
                    _mapping_table_rows(mapping_payload)
                    if mapping_payload and mapping_payload.get("type") != "cells"
                    else _excel_rows_from_context(context_str)
                )
                target_sheet = str((mapping_payload or {}).get("sheet") or sheet_name or handle.workbook.sheetnames[0])
                if mapping_payload and mapping_payload.get("type") == "cells":
                    for target_cell, value in (mapping_payload.get("cells") or {}).items():
                        write_range(
                            handle,
                            target_sheet,
                            str(target_cell).upper(),
                            [[_safe_spreadsheet_value(value)]],
                        )
                    written = len(mapping_payload.get("cells") or {})
                elif data:
                    start_range = str(
                        (mapping_payload or {}).get("start_cell")
                        or config.get("start_range")
                        or "A1"
                    ).strip() or "A1"
                    safe_data = [
                        [_safe_spreadsheet_value(value) for value in row]
                        for row in data
                    ]
                    write_range(handle, target_sheet, start_range, safe_data)
                    written = (
                        len(mapping_payload.get("rows") or [])
                        if mapping_payload
                        else len(data)
                    )
                else:
                    written = 0
                Path(abs_out).parent.mkdir(parents=True, exist_ok=True)
                save_as(handle, abs_out)
                saved_path = await _record_output(abs_out)
                log.info(
                    "[Flow/Office] excel_write: %s → %s (%d mapped items)", input_path, saved_path, written
                )
                return f"Excel updated: {saved_path} ({written} mapped item(s) written)"
            finally:
                close_workbook(handle)

        # ── Word Operations ──
        elif action == "word_read":
            from shogun.office.adapters.word_adapter import (
                open_document,
                read_text,
                close_document,
            )

            abs_path = _resolve(input_path)
            handle = open_document(abs_path)
            try:
                text = read_text(handle)
                log.info("[Flow/Office] word_read: %s (%d chars)", input_path, len(text))
                return text
            finally:
                close_document(handle)

        elif action == "word_create":
            abs_out = _resolve_output(output_path, ".docx", "output.docx")
            template = _create_template("docx")
            if template:
                from shogun.services.file_template import render_word_template, resolve_workspace_template

                source = resolve_workspace_template(str(template.get("template_path") or ""), root)
                populated = render_word_template(
                    source,
                    Path(abs_out),
                    context_str,
                    str(template.get("example_handling") or "replace"),
                    str(template.get("render_mode") or "adaptive"),
                )
                template_summary = f"{populated} template item(s) populated"
            else:
                from docx import Document

                doc = Document()
                # Use the inline text template or predecessor context as content.
                content = config.get("content_template", "") or context_str or ""
                if config.get("content_template"):
                    content = content.replace("{{context}}", context_str)
                for para in content.split("\n"):
                    doc.add_paragraph(para)
                Path(abs_out).parent.mkdir(parents=True, exist_ok=True)
                doc.save(abs_out)
                template_summary = "standard document"
            saved_path = await _record_output(abs_out)
            log.info("[Flow/Office] word_create: %s (%s)", saved_path, template_summary)
            return f"Word document created: {saved_path} ({template_summary})"

        elif action == "word_replace":
            from shogun.office.adapters.word_adapter import (
                open_document,
                replace_placeholders,
                save_as,
                close_document,
            )

            abs_in = _resolve(input_path)
            abs_out = (
                _resolve_output(output_path, ".docx", Path(abs_in).name)
                if output_path
                else str(_scheduled_output_path(Path(abs_in), trigger_type, run_id))
            )
            handle = open_document(abs_in)
            try:
                replacements = config.get("replacements", {})
                if not replacements:
                    return "[ERROR] No replacements specified for word_replace"
                count = replace_placeholders(handle, replacements)
                Path(abs_out).parent.mkdir(parents=True, exist_ok=True)
                save_as(handle, abs_out)
                saved_path = await _record_output(abs_out)
                log.info(
                    "[Flow/Office] word_replace: %s → %s (%d replacements)",
                    input_path,
                    saved_path,
                    count,
                )
                return f"Word document updated: {saved_path} ({count} replacements)"
            finally:
                close_document(handle)

        # ── PowerPoint Operations ──
        elif action == "pptx_read":
            from shogun.office.adapters.pptx_adapter import (
                open_presentation,
                list_slides,
                read_slide_text,
                close_presentation,
            )

            abs_path = _resolve(input_path)
            handle = open_presentation(abs_path)
            try:
                slides = list_slides(handle)
                texts = []
                for i, s in enumerate(slides):
                    text = read_slide_text(handle, i)
                    texts.append(f"[Slide {i + 1}: {s.get('title', 'Untitled')}]\n{text}")
                result = "\n\n".join(texts)
                log.info("[Flow/Office] pptx_read: %s (%d slides)", input_path, len(slides))
                return result
            finally:
                close_presentation(handle)

        elif action == "pptx_replace":
            from shogun.office.adapters.pptx_adapter import (
                open_presentation,
                replace_placeholders,
                save_as,
                close_presentation,
            )

            abs_in = _resolve(input_path)
            abs_out = (
                _resolve_output(output_path, ".pptx", Path(abs_in).name)
                if output_path
                else str(_scheduled_output_path(Path(abs_in), trigger_type, run_id))
            )
            handle = open_presentation(abs_in)
            try:
                replacements = config.get("replacements", {})
                if not replacements:
                    return "[ERROR] No replacements specified for pptx_replace"
                count = replace_placeholders(handle, replacements)
                Path(abs_out).parent.mkdir(parents=True, exist_ok=True)
                save_as(handle, abs_out)
                saved_path = await _record_output(abs_out)
                log.info("[Flow/Office] pptx_replace: %s → %s", input_path, saved_path)
                return f"Presentation updated: {saved_path} ({count} replacements)"
            finally:
                close_presentation(handle)

        else:
            return f"[ERROR] Unknown files action: {action}"

    except ImportError as exc:
        return f"[ERROR] Missing dependency for '{action}': {exc}"
    except Exception as exc:
        return f"[ERROR] Files '{action}' failed: {str(exc)[:500]}"


# ═══════════════════════════════════════════════════════════════
# LLM RESOLUTION & CALLING
# ═══════════════════════════════════════════════════════════════


PROVIDER_URLS = {
    "ollama": "http://127.0.0.1:11434",
    "lmstudio": "http://localhost:1234/v1",
    "local": "http://localhost:1234/v1",
    "openai": "https://api.openai.com/v1",
    "openrouter": "https://openrouter.ai/api/v1",
    "anthropic": "https://api.anthropic.com/v1",
    "google": "https://generativelanguage.googleapis.com/v1beta/openai",
    "custom": "https://api.openai.com/v1",
}


async def _resolve_llm(
    session: AsyncSession,
    routing_profile_id: str | None = None,
) -> tuple[ModelProvider | None, str, str, dict]:
    """Resolve the LLM provider, model name, base URL, and headers.

    Returns (provider, model_name, base_url, headers) or (None, "", "", {}).
    """
    chain = await _resolve_llm_chain(session, routing_profile_id)
    if not chain:
        return None, "", "", {}
    return chain[0]


def _provider_connection(
    provider: ModelProvider,
    model_name: str | None = None,
) -> tuple[ModelProvider, str, str, dict]:
    """Build an API target tuple for one provider/model."""
    resolved_model = (
        model_name
        or provider.config.get("model_id")
        or provider.config.get("model")
        or (provider.config.get("models") or [None])[0]
        or provider.name
    )
    # Resolve base URL
    base_url = provider.base_url or PROVIDER_URLS.get(provider.provider_type, "https://api.openai.com/v1")
    if provider.provider_type == "ollama" and not base_url.rstrip("/").endswith("/v1"):
        base_url = base_url.rstrip("/") + "/v1"

    # Build headers
    api_key = provider_api_key(provider.config)
    headers: dict[str, str] = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    if provider.provider_type == "openrouter":
        headers["HTTP-Referer"] = "https://shogun.ai"
        headers["X-Title"] = "Shogun AgentFlow"
    if provider.provider_type == "google" and provider.config.get("oauth_project_id"):
        headers["x-goog-user-project"] = str(provider.config["oauth_project_id"])

    return provider, resolved_model, base_url, headers


async def _resolve_model_target(
    session: AsyncSession, target_id: str | uuid.UUID
) -> tuple[ModelProvider, str, str, dict] | None:
    """Resolve routing IDs for registry entries, providers, or model definitions.

    Katana's governed router stores model-registry IDs in custom profiles.  Those
    IDs must resolve through the registry entry's exact ``provider_id`` so the
    selected provider's protected credential and endpoint travel with the model.
    """
    try:
        parsed_id = target_id if isinstance(target_id, uuid.UUID) else uuid.UUID(str(target_id))
    except (TypeError, ValueError):
        return None

    provider = await session.scalar(
        select(ModelProvider).where(
            ModelProvider.id == parsed_id,
            ModelProvider.status == "connected",
        )
    )
    if provider:
        return _provider_connection(provider)

    registry_entry = await session.get(ModelRegistryEntry, parsed_id)
    if registry_entry and registry_entry.enabled and registry_entry.provider_id:
        provider = await session.scalar(
            select(ModelProvider).where(
                ModelProvider.id == registry_entry.provider_id,
                ModelProvider.status == "connected",
            )
        )
        if provider:
            return _provider_connection(provider, registry_entry.model_id)

    definition = await session.scalar(select(ModelDefinition).where(ModelDefinition.id == parsed_id))
    if definition and definition.provider and definition.provider.status == "connected":
        return _provider_connection(definition.provider, definition.model_key)
    return None


async def _resolve_llm_chain(
    session: AsyncSession,
    routing_profile_id: str | None = None,
) -> list[tuple[ModelProvider, str, str, dict]]:
    """Resolve a primary model and every configured fallback in exact order."""
    profile: ModelRoutingProfile | None = None
    try:
        if routing_profile_id:
            try:
                profile_id = uuid.UUID(routing_profile_id)
            except (TypeError, ValueError):
                profile_id = None
            profile = await session.scalar(
                select(ModelRoutingProfile).where(
                    ModelRoutingProfile.id == profile_id
                    if profile_id
                    else ModelRoutingProfile.name == routing_profile_id
                )
            )
        else:
            profile = await session.scalar(
                select(ModelRoutingProfile)
                .where(ModelRoutingProfile.is_default.is_(True))
                .order_by(ModelRoutingProfile.updated_at.desc())
                .limit(1)
            )
    except Exception as exc:
        log.warning("Failed to resolve routing profile %s: %s", routing_profile_id, exc)

    chain: list[tuple[ModelProvider, str, str, dict]] = []
    seen: set[tuple[str, str]] = set()
    if profile:
        rule = next((r for r in profile.rules if r.get("task_type") == "*"), None)
        if not rule and profile.rules:
            rule = profile.rules[0]
        if rule:
            target_ids = [rule.get("primary_model_id"), *(rule.get("fallback_model_ids") or [])]
            for target_id in target_ids:
                if not target_id:
                    continue
                target = await _resolve_model_target(session, target_id)
                if target:
                    key = (str(target[0].id), target[1])
                    if key not in seen:
                        seen.add(key)
                        chain.append(target)
                else:
                    log.warning("Configured routing target %s is unavailable; skipping it", target_id)

    if not chain:
        provider = await session.scalar(
            select(ModelProvider).where(ModelProvider.status == "connected").order_by(ModelProvider.created_at).limit(1)
        )
        if provider:
            chain.append(_provider_connection(provider))
    refreshed: list[tuple[ModelProvider, str, str, dict]] = []
    for provider, model_name, _base_url, _headers in chain:
        if provider.auth_type == "oauth":
            await ensure_provider_access_token(session, provider)
        refreshed.append(_provider_connection(provider, model_name))
    await session.flush()
    return refreshed


_VISION_MODEL_MARKERS = (
    "gemma3",
    "llava",
    "bakllava",
    "minicpm-v",
    "moondream",
    "qwen-vl",
    "qwen2-vl",
    "qwen2.5-vl",
    "qwen3-vl",
    "pixtral",
    "llama3.2-vision",
    "llama-3.2-vision",
    "gemini",
    "gpt-4o",
    "gpt-4.1",
    "gpt-5",
    "claude-3",
    "claude-sonnet",
)


def _model_supports_vision(model_name: str, provider: ModelProvider | None = None) -> bool:
    """Conservatively identify provider models known to accept image input."""
    config = (provider.config or {}) if provider else {}
    if config.get("supports_vision") is True:
        return True
    normalized = model_name.lower().replace("_", "-")
    return any(marker in normalized for marker in _VISION_MODEL_MARKERS)


async def _resolve_vision_chain(
    session: AsyncSession,
    routing_profile_id: str | None = None,
) -> list[tuple[ModelProvider, str, str, dict]]:
    """Resolve only endpoints that are explicitly or reliably vision-capable.

    The normal routing profile may intentionally point at a fast text model.
    Images must never be sent to that endpoint unless it supports vision.
    """
    chain: list[tuple[ModelProvider, str, str, dict]] = []
    seen: set[tuple[str, str]] = set()

    def add(target: tuple[ModelProvider, str, str, dict]) -> None:
        key = (str(target[0].id), target[1])
        if key not in seen:
            seen.add(key)
            chain.append(target)

    # Preserve the configured routing order for any vision-capable entries.
    for target in await _resolve_llm_chain(session, routing_profile_id):
        if _model_supports_vision(target[1], target[0]):
            add(target)

    # Model definitions are authoritative capability declarations.
    definitions = (
        (
            await session.execute(
                select(ModelDefinition)
                .where(
                    ModelDefinition.status == "available",
                    ModelDefinition.supports_vision.is_(True),
                )
                .order_by(ModelDefinition.created_at)
            )
        )
        .scalars()
        .unique()
        .all()
    )
    for definition in definitions:
        if definition.provider and definition.provider.status == "connected":
            add(_provider_connection(definition.provider, definition.model_key))

    # Older installations store their model directly on the provider.
    providers = (
        (
            await session.execute(
                select(ModelProvider).where(ModelProvider.status == "connected").order_by(ModelProvider.created_at)
            )
        )
        .scalars()
        .all()
    )
    for provider in providers:
        vision_model = (provider.config or {}).get("vision_model")
        if vision_model:
            add(_provider_connection(provider, str(vision_model)))
        default_target = _provider_connection(provider)
        if _model_supports_vision(default_target[1], provider):
            add(default_target)

    return chain


async def _resolve_task_llm_chain(
    session: AsyncSession,
    *,
    prompt: str,
    task_type: str | None = None,
    required_capabilities: list[str] | None = None,
    routing_profile_id: str | None = None,
    run_id: str | uuid.UUID | None = None,
    stack_run_id: str | uuid.UUID | None = None,
    step_id: str | None = None,
    retry_count: int = 0,
    verification_status: str | None = None,
    risk_level: str = "low",
    context_size_estimate: int = 0,
    escalation_level: int = 0,
    exclude_model_ids: list[str] | None = None,
    local_only: bool = False,
) -> tuple[list[tuple[ModelProvider, str, str, dict]], dict | None]:
    """Select a governed task-aware model chain with legacy fallback."""
    from shogun.services.model_router import NoEligibleModelError

    try:
        from shogun.schemas.model_router import ModelRouteRequest
        from shogun.services.model_router import ModelRoutingService

        def parsed(value):
            try:
                return value if isinstance(value, uuid.UUID) else uuid.UUID(str(value)) if value else None
            except (TypeError, ValueError):
                return None

        result = await ModelRoutingService(session).route(
            ModelRouteRequest(
                prompt=prompt,
                task_type=task_type,
                required_capabilities=required_capabilities or ["chat"],
                profile_override=routing_profile_id,
                run_id=parsed(run_id),
                stack_run_id=parsed(stack_run_id),
                step_id=step_id,
                retry_count=retry_count,
                verification_status=verification_status,
                risk_level=risk_level if risk_level in {"low", "medium", "high", "critical"} else "low",
                context_size_estimate=context_size_estimate,
                escalation_level=escalation_level,
                exclude_model_ids=exclude_model_ids or [],
                local_only=local_only,
            )
        )
        chain: list[tuple[ModelProvider, str, str, dict]] = []
        for entry in [result.selected, *result.fallbacks]:
            provider = await session.get(ModelProvider, entry.provider_id) if entry.provider_id else None
            if provider and provider.status == "connected":
                if getattr(provider, "auth_type", None) == "oauth":
                    await ensure_provider_access_token(session, provider)
                chain.append(_provider_connection(provider, entry.model_id))
        if chain:
            flush = getattr(session, "flush", None)
            if flush is not None:
                await flush()
            return chain, result.payload
    except NoEligibleModelError as exc:
        # Compatibility path for upgraded desktop installations: Comms and
        # pre-router AgentFlows resolve directly from connected providers. If
        # the governed registry has no ordinary chat candidate, keep AgentFlow
        # operational through that same provider chain. Policy-sensitive
        # failures (budget, premium approval, context limits, custom profiles,
        # vision/tool requirements) do not opt into this fallback and remain
        # authoritative hard failures.
        if exc.allow_connected_fallback and set(required_capabilities or ["chat"]) == {"chat"}:
            legacy_chain = await _resolve_llm_chain(session, routing_profile_id)
            if legacy_chain:
                provider, model_name, *_ = legacy_chain[0]
                log.warning(
                    "Task-aware routing found no chat candidate; using connected-provider "
                    "compatibility route %s/%s",
                    provider.name,
                    model_name,
                )
                return legacy_chain, {
                    "active_profile": "connected_provider_compatibility",
                    "selected_model": model_name,
                    "selected_provider": provider.provider_type,
                    "selected_max_output_tokens": None,
                    "fallback_reason": str(exc),
                }
        raise
    except Exception as exc:
        log.info("Task-aware routing unavailable; using legacy model chain: %s", exc)
    return await _resolve_llm_chain(session, routing_profile_id), None


async def _call_llm(
    messages: list[dict],
    model_name: str,
    base_url: str,
    headers: dict,
    timeout: int = 120,
    max_tokens: int | None = None,
    temperature: float = 0.3,
    seed: int | None = None,
    provider_type: str = "",
    provider_auth_type: str = "",
    provider_config: dict[str, Any] | None = None,
    reasoning_effort: str | None = None,
) -> str:
    """Make a non-streaming chat completion call and return the response text."""
    payload = {
        "model": model_name,
        "messages": messages,
        "stream": False,
        "temperature": temperature,
    }
    if max_tokens is not None:
        payload["max_tokens"] = max_tokens
    if seed is not None:
        payload["seed"] = seed
    if provider_type:
        apply_chat_reasoning(
            payload,
            provider_type=provider_type,
            model_id=model_name,
            provider_config=provider_config,
            explicit_effort=reasoning_effort,
        )

    resp = await model_chat_completion(
        auth_type=provider_auth_type,
        base_url=base_url,
        headers=headers,
        payload=payload,
        timeout=float(timeout),
    )
    if resp.status_code >= 400:
        body = resp.text[:500]
        raise ValueError(f"LLM API error {resp.status_code}: {body}")

    data = resp.json()
    choices = data.get("choices", [])
    if not choices:
        raise ValueError("LLM returned no choices")

    content = choices[0].get("message", {}).get("content", "")
    if not content:
        raise ValueError("LLM returned empty content")

    return content


_AGENTFLOW_ROWS_TOOL_NAME = "agentflow_submit_rows"


def _agentflow_rows_tool(expected_width: int | None) -> dict[str, Any]:
    width_text = (
        f" Each row must contain exactly {expected_width} cell values."
        if expected_width is not None
        else " Each row must contain one value per destination column."
    )
    return {
        "type": "function",
        "function": {
            "name": _AGENTFLOW_ROWS_TOOL_NAME,
            "description": (
                "Submit every extracted spreadsheet row from the current document chunk."
                + width_text
                + " This operation cannot access files or choose an output path."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "rows": {
                        "type": "array",
                        "description": "Two-dimensional array: outer items are rows, inner items are cell values.",
                        "items": {"type": "array", "items": {}},
                    }
                },
                "required": ["rows"],
                "additionalProperties": False,
            },
        },
    }


def _validate_agentflow_rows(arguments: dict[str, Any], expected_width: int | None) -> list[list[Any]]:
    rows = arguments.get("rows")
    if not isinstance(rows, list):
        raise ValueError("agentflow_submit_rows requires a rows array")
    validated: list[list[Any]] = []
    invalid: list[tuple[int, int]] = []
    for index, row in enumerate(rows, start=1):
        if not isinstance(row, list):
            raise ValueError(f"agentflow_submit_rows row {index} is not an array")
        if expected_width is not None and len(row) != expected_width:
            invalid.append((index, len(row)))
        validated.append(row)
    if invalid:
        detail = ", ".join(f"row {row} has {width}" for row, width in invalid[:6])
        raise ValueError(
            f"agentflow_submit_rows requires exactly {expected_width} values per row; {detail}"
        )
    return validated


def _text_tool_messages(messages: list[dict], expected_width: int | None) -> list[dict]:
    width_rule = (
        f"Every inner array MUST contain exactly {expected_width} values."
        if expected_width is not None
        else "Every inner array must contain one value per destination column."
    )
    protocol = (
        "\n\n--- SHOGUN STRUCTURED TOOL PROTOCOL ---\n"
        "Do not answer with prose or Markdown. Submit the extracted rows by returning exactly:\n"
        '<tool_call>{"tool":"agentflow_submit_rows","arguments":{"rows":[["cell 1","cell 2"]]}}</tool_call>\n'
        f"{width_rule} Include every relevant record in this chunk. Use an empty rows array only when "
        "the chunk contains no relevant records."
    )
    prepared = [dict(message) for message in messages]
    if prepared and prepared[-1].get("role") == "user":
        prepared[-1]["content"] = str(prepared[-1].get("content") or "") + protocol
    else:
        prepared.append({"role": "user", "content": protocol})
    return prepared


async def _call_llm_rows(
    messages: list[dict],
    model_name: str,
    base_url: str,
    headers: dict,
    *,
    profile: dict[str, Any],
    expected_width: int | None,
    timeout: int,
    max_tokens: int | None,
    temperature: float,
    seed: int | None,
    provider_auth_type: str = "",
    row_validator: Callable[[list[list[Any]]], None] | None = None,
) -> tuple[list[list[Any]], str, str]:
    """Request rows through the model's persisted tool transport.

    The operation is deliberately an internal structured-output sink. It does
    not grant filesystem or Office authority; those remain separate governed
    AgentFlow nodes.
    """
    mode = str(profile.get("mode") or "text")
    request_messages = messages if mode == "native" else _text_tool_messages(messages, expected_width)
    payload: dict[str, Any] = {
        "model": model_name,
        "messages": request_messages,
        "stream": False,
        "temperature": temperature,
    }
    if max_tokens is not None:
        payload["max_tokens"] = max_tokens
    if seed is not None:
        payload["seed"] = seed
    if mode == "native":
        payload["tools"] = [_agentflow_rows_tool(expected_width)]
        payload["tool_choice"] = {
            "type": "function",
            "function": {"name": _AGENTFLOW_ROWS_TOOL_NAME},
        }

    response = await model_chat_completion(
        auth_type=provider_auth_type,
        base_url=base_url,
        headers=headers,
        payload=payload,
        timeout=float(timeout),
    )
    if response.status_code >= 400 and mode == "native" and "tool_choice" in payload:
        # Some OpenAI-compatible servers implement native tools but not forced
        # function choice. A second native attempt distinguishes that case
        # from models that do not support the tools field at all.
        payload.pop("tool_choice", None)
        response = await model_chat_completion(
            auth_type=provider_auth_type,
            base_url=base_url,
            headers=headers,
            payload=payload,
            timeout=float(timeout),
        )
    if response.status_code >= 400:
        # A model/provider may claim native support while rejecting the tools
        # field. Retry the same bounded operation through Shogun's text adapter.
        if mode == "native" and bool(profile.get("fallback_enabled", True)):
            fallback = dict(profile)
            fallback["mode"] = "text"
            fallback["adapter_id"] = "shogun_text_v1"
            return await _call_llm_rows(
                messages,
                model_name,
                base_url,
                headers,
                profile=fallback,
                expected_width=expected_width,
                timeout=timeout,
                max_tokens=max_tokens,
                temperature=temperature,
                seed=seed,
                provider_auth_type=provider_auth_type,
                row_validator=row_validator,
            )
        raise ValueError(f"LLM API error {response.status_code}: {response.text[:500]}")

    data = response.json()
    calls = normalize_native_tool_calls(data) if mode == "native" else []
    choices = data.get("choices") or []
    message = choices[0].get("message", {}) if choices else data.get("message", {})
    content = message.get("content", "") if isinstance(message, dict) else ""
    if isinstance(content, list):
        content = "\n".join(
            str(part.get("text") or "") for part in content if isinstance(part, dict)
        )
    try:
        if not calls:
            calls = normalize_text_tool_calls(str(content or ""), {_AGENTFLOW_ROWS_TOOL_NAME})
        call = next((item for item in calls if item.get("tool") == _AGENTFLOW_ROWS_TOOL_NAME), None)
        if not call:
            from shogun.services.file_template import parse_excel_rows

            try:
                parsed_rows = parse_excel_rows(str(content or ""), require_structured_json=True)
                rows = _validate_agentflow_rows({"rows": parsed_rows}, expected_width)
            except Exception:
                if mode == "native" and bool(profile.get("fallback_enabled", True)):
                    raise
                # Text adapters may legitimately return an exact-width Markdown
                # table or TSV matrix. Accept those structured representations,
                # but never turn arbitrary prose into a successful spreadsheet.
                parsed_rows = parse_excel_rows(str(content or ""), require_structured_json=False)
                rows = _validate_agentflow_rows({"rows": parsed_rows}, expected_width)
        else:
            rows = _validate_agentflow_rows(call.get("arguments") or {}, expected_width)
        if row_validator:
            row_validator(rows)
        return rows, json.dumps(rows, ensure_ascii=False, default=str), mode
    except Exception as parse_error:
        if mode == "native" and bool(profile.get("fallback_enabled", True)):
            compact_content = re.sub(r"\s+", " ", str(content or "")).strip()
            structural_preview = re.sub(r"[A-Za-z0-9]", "x", compact_content)[:160]
            log.warning(
                "Native AgentFlow row response from %s was structurally invalid; "
                "retrying with Shogun text adapter (content_length=%d, structural_preview=%r): %s",
                model_name,
                len(compact_content),
                structural_preview,
                parse_error,
            )
            fallback = dict(profile)
            fallback["mode"] = "text"
            fallback["adapter_id"] = "shogun_text_v1"
            return await _call_llm_rows(
                messages,
                model_name,
                base_url,
                headers,
                profile=fallback,
                expected_width=expected_width,
                timeout=timeout,
                max_tokens=max_tokens,
                temperature=temperature,
                seed=seed,
                provider_auth_type=provider_auth_type,
                row_validator=row_validator,
            )
        if isinstance(parse_error, IncompleteMatrixOutputError):
            raise
        raise ValueError(
            "The model did not submit rows through agentflow_submit_rows or valid JSON rows"
        ) from parse_error


async def _call_llm_chain_rows(
    messages: list[dict],
    model_chain: list[tuple[ModelProvider, str, str, dict]],
    *,
    timeout: int,
    retry_count: int,
    context: str,
    expected_width: int | None,
    max_tokens: int | None,
    routing_context: dict[str, Any] | None,
    row_validator: Callable[[list[list[Any]]], None] | None = None,
) -> list[list[Any]]:
    """Call routed models until one returns a validated canonical row matrix."""
    last_error: Exception | None = None
    last_provider: ModelProvider | None = None
    last_model = "unknown-model"
    input_characters = len(json.dumps(messages, ensure_ascii=False, default=str))
    profiles = (routing_context or {}).get("tool_calling_profiles") or {}
    registry_ids = (routing_context or {}).get("tool_calling_registry_ids") or {}
    for provider, model_name, base_url, headers in model_chain:
        last_provider = provider
        last_model = model_name
        route_key = f"{getattr(provider, 'id', '')}:{model_name}"
        profile = profiles.get(route_key) or infer_tool_calling_profile(
            model_name,
            str(getattr(provider, "provider_type", "")),
        )
        parameters = ((routing_context or {}).get("request_parameters") or {}).get(route_key) or {}
        temperature = max(0.0, min(2.0, float(parameters.get("temperature", 0.1))))
        configured_seed = (routing_context or {}).get("flow_seed")
        seed = int(configured_seed) if configured_seed is not None else None
        for attempt in range(1 + retry_count):
            started = time.perf_counter()
            try:
                rows, usage_output, used_mode = await _call_llm_rows(
                    messages,
                    model_name,
                    base_url,
                    headers,
                    profile=profile,
                    expected_width=expected_width,
                    timeout=timeout,
                    max_tokens=max_tokens,
                    temperature=temperature,
                    seed=seed,
                    provider_auth_type=str(getattr(provider, "auth_type", "") or ""),
                    row_validator=row_validator,
                )
                if row_validator:
                    row_validator(rows)
                await _record_model_usage(
                    provider,
                    model_name,
                    messages,
                    usage_output,
                    time.perf_counter() - started,
                    routing_context,
                    None,
                    success=True,
                )
                if profile.get("mode") == "native" and used_mode == "text" and registry_ids.get(route_key):
                    try:
                        from shogun.services.model_router import ModelRegistryService

                        async with async_session_factory() as repair_session:
                            item = await repair_session.get(
                                ModelRegistryEntry,
                                uuid.UUID(str(registry_ids[route_key])),
                            )
                            if item:
                                await ModelRegistryService(repair_session).mark_tool_calling_fallback(
                                    item,
                                    "Native AgentFlow tool request failed or returned an invalid structure; "
                                    "Shogun text adapter succeeded.",
                                )
                                await repair_session.commit()
                    except Exception as repair_error:
                        log.warning("Could not persist AgentFlow tool-profile fallback: %s", repair_error)
                return rows
            except Exception as exc:
                last_error = exc
                await _record_model_usage(
                    provider,
                    model_name,
                    messages,
                    "",
                    time.perf_counter() - started,
                    routing_context,
                    None,
                    success=False,
                    error=str(exc),
                )
                log.warning(
                    "%s structured tool call failed for %s (attempt %d/%d): %s",
                    context,
                    model_name,
                    attempt + 1,
                    1 + retry_count,
                    exc,
                )
                # Repeating a syntactically valid but incomplete/malformed row
                # response with the same prompt and model is rarely useful,
                # especially for deterministic local routes. The chunk
                # executor already has targeted subdivision and corrective
                # retries, and the outer loop can still try the next model.
                if isinstance(exc, IncompleteMatrixOutputError) or _is_structured_rows_failure(exc):
                    break
                if attempt < retry_count:
                    await asyncio.sleep(2**attempt)
    if last_error is None:
        raise ValueError(f"{context} failed without a model response")
    provider_name = getattr(last_provider, "name", None) or "unknown-provider"
    raise ModelCallError(
        context=context,
        provider=str(provider_name),
        model=last_model,
        timeout=timeout,
        cause=last_error,
        input_characters=input_characters,
    ) from last_error


async def _call_llm_chain(
    messages: list[dict],
    model_chain: list[tuple[ModelProvider, str, str, dict]],
    *,
    timeout: int,
    retry_count: int,
    context: str,
    max_tokens: int | None = None,
    routing_context: dict[str, Any] | None = None,
    usage_session: AsyncSession | None = None,
) -> str:
    """Call each model in order, transparently notifying on every transition."""
    last_error: Exception | None = None
    last_provider: ModelProvider | None = None
    last_model = "unknown-model"
    input_characters = len(json.dumps(messages, ensure_ascii=False, default=str))
    for model_index, (_provider, model_name, base_url, headers) in enumerate(model_chain):
        last_provider = _provider
        last_model = model_name
        for attempt in range(1 + retry_count):
            started = time.perf_counter()
            try:
                route_key = f"{getattr(_provider, 'id', '')}:{model_name}"
                model_parameters = ((routing_context or {}).get("request_parameters") or {}).get(route_key) or {}
                temperature = max(0.0, min(2.0, float(model_parameters.get("temperature", 0.3))))
                reasoning_effort = model_parameters.get("reasoning_effort")
                provider_config = getattr(_provider, "config", {}) or {}
                provider_type = getattr(_provider, "provider_type", None)
                provider_kwargs = (
                    {
                        "provider_type": provider_type,
                        "provider_auth_type": str(getattr(_provider, "auth_type", "") or ""),
                        "provider_config": provider_config,
                        "reasoning_effort": reasoning_effort,
                    }
                    if provider_type
                    else {}
                )
                configured_seed = (routing_context or {}).get("flow_seed")
                seed_match = str((routing_context or {}).get("flow_seed_model_id") or "").strip()
                physical_model = f"{getattr(_provider, 'id', '')}:{model_name}"
                provider_model = f"{getattr(_provider, 'provider_type', '')}/{model_name}"
                seed = (
                    int(configured_seed)
                    if configured_seed is not None
                    and (not seed_match or seed_match in {model_name, physical_model, provider_model})
                    else None
                )
                default_controls = temperature == 0.3 and seed is None
                if max_tokens is None and default_controls:
                    result = await _call_llm(
                        messages,
                        model_name,
                        base_url,
                        headers,
                        timeout,
                        **provider_kwargs,
                    )
                elif default_controls:
                    result = await _call_llm(
                        messages,
                        model_name,
                        base_url,
                        headers,
                        timeout,
                        max_tokens=max_tokens,
                        **provider_kwargs,
                    )
                else:
                    result = await _call_llm(
                        messages,
                        model_name,
                        base_url,
                        headers,
                        timeout,
                        max_tokens=max_tokens,
                        temperature=temperature,
                        seed=seed,
                        **provider_kwargs,
                    )
                await _record_model_usage(
                    _provider,
                    model_name,
                    messages,
                    result,
                    time.perf_counter() - started,
                    routing_context,
                    usage_session,
                    success=True,
                )
                return result
            except Exception as exc:
                last_error = exc
                await _record_model_usage(
                    _provider,
                    model_name,
                    messages,
                    "",
                    time.perf_counter() - started,
                    routing_context,
                    usage_session,
                    success=False,
                    error=str(exc),
                )
                log.warning(
                    "%s model '%s' failed (attempt %d/%d, timeout=%ss): %s",
                    context,
                    model_name,
                    attempt + 1,
                    1 + retry_count,
                    timeout,
                    exc,
                )
                if attempt < retry_count:
                    await asyncio.sleep(2**attempt)

        if model_index + 1 < len(model_chain):
            next_model = model_chain[model_index + 1][1]
            reason = (
                f"timeout after {timeout}s"
                if isinstance(last_error, (httpx.TimeoutException, asyncio.TimeoutError))
                else str(last_error)[:300]
            )
            from shogun.services.notification_service import notify_model_fallback

            await notify_model_fallback(
                from_model=model_name,
                to_model=next_model,
                reason=reason,
                context=context,
                timeout_seconds=timeout,
            )

    if last_error is None:
        raise ValueError(f"{context} failed without a model response")
    provider_name = (
        getattr(last_provider, "name", None)
        or getattr(last_provider, "provider_type", None)
        or "unknown-provider"
    )
    raise ModelCallError(
        context=context,
        provider=str(provider_name),
        model=last_model,
        timeout=timeout,
        cause=last_error,
        input_characters=input_characters,
    ) from last_error


async def _resolve_samurai_tools(governance_context: dict[str, Any] | None) -> tuple[list[dict[str, Any]], Callable]:
    """Load the governed, read-only artifact tools for Samurai inference.

    A Samurai transforms inputs into a typed result.  It may inspect approved
    artifacts, but the explicit downstream node remains the sole owner of
    writes, sends, saves, and other side effects.
    """
    from shogun.services.native_skills import NATIVE_TOOLS, execute_native_tool
    from shogun.services.posture_guard import filter_tools_by_posture, get_posture_permissions

    posture_permissions = (governance_context or {}).get("permissions")
    if not posture_permissions:
        try:
            posture_permissions = await get_posture_permissions()
        except Exception:
            posture_permissions = {}

    artifact_read_tools = {
        "workspace_info",
        "workspace_list",
        "workspace_read",
        "workspace_read_image",
        "workspace_read_pdf",
        "office_excel_open",
        "office_excel_open_attachment",
        "office_excel_read_range",
        "office_excel_list_sheets",
        "office_excel_get_metadata",
        "office_word_open",
        "office_word_get_metadata",
        "office_word_read_text",
        "office_word_read_page",
        "office_word_read_pages",
        "office_word_read_headings",
        "office_pptx_open",
        "office_pptx_get_metadata",
        "file_detect_type",
        "file_inspect",
        "file_read",
        "file_preview",
        "file_schema",
        "file_query",
        "file_extract",
        "file_compare",
        "file_validate",
        "file_list_formats",
    }

    allowed_tools, _ = filter_tools_by_posture(NATIVE_TOOLS, posture_permissions)
    filtered_tools = [
        tool
        for tool in allowed_tools
        if tool.get("function", {}).get("name") in artifact_read_tools
    ]

    async def _executor(tool_name: str, args: dict[str, Any], session: AsyncSession) -> str:
        return await execute_native_tool(tool_name, args, session)

    return filtered_tools, _executor


async def _call_llm_with_tools(
    messages: list[dict],
    model_name: str,
    base_url: str,
    headers: dict,
    timeout: int = 120,
    max_tokens: int | None = None,
    temperature: float = 0.3,
    seed: int | None = None,
    provider_auth_type: str = "",
    tools: list[dict] | None = None,
    tool_executor: Callable | None = None,
    max_tool_rounds: int = 6,
    max_tool_calls: int | None = None,
    governance_context: dict[str, Any] | None = None,
    tool_profile: dict[str, Any] | None = None,
) -> str:
    """Run a bounded, governed tool loop through a model-specific adapter."""
    if not tools or not tool_executor:
        return await _call_llm(
            messages,
            model_name,
            base_url,
            headers,
            timeout=timeout,
            max_tokens=max_tokens,
            temperature=temperature,
            seed=seed,
            provider_auth_type=provider_auth_type,
        )

    profile = dict(tool_profile or {})
    mode = str(profile.get("mode") or "native")
    if mode == "unsupported":
        return await _call_llm(
            messages,
            model_name,
            base_url,
            headers,
            timeout=timeout,
            max_tokens=max_tokens,
            temperature=temperature,
            seed=seed,
            provider_auth_type=provider_auth_type,
        )
    formatted_tools = [
        {"type": t.get("type", "function"), "function": t["function"]}
        for t in tools
    ]
    allowed_tool_names = {
        str(tool.get("function", {}).get("name") or "")
        for tool in tools
    }

    current_messages = [dict(m) for m in messages]
    if mode == "text":
        from shogun.services.native_skills import generate_tool_prompt

        tool_prompt = generate_tool_prompt(tools)
        if current_messages and current_messages[0].get("role") == "system":
            current_messages[0]["content"] = (
                str(current_messages[0].get("content") or "") + "\n\n" + tool_prompt
            )
        else:
            current_messages.insert(0, {"role": "system", "content": tool_prompt})
    # Reserve a final tool-free response after the last permitted tool round.
    # The old loop raised immediately after the last tool result, discarded all
    # gathered evidence, and made callers repeat the entire research run.
    max_tool_rounds = max(1, int(max_tool_rounds))
    max_tool_calls = max(1, int(max_tool_calls)) if max_tool_calls is not None else None
    tool_calls_used = 0
    force_finalize = False
    tool_result_cache: dict[str, str] = {}

    for _round_num in range(max_tool_rounds + 1):
        tools_enabled = _round_num < max_tool_rounds and not force_finalize
        request_messages = current_messages
        if not tools_enabled:
            request_messages = [
                *current_messages,
                {
                    "role": "system",
                    "content": (
                        "The governed tool budget is complete. Do not request more tools. "
                        "Use the evidence already gathered and return the final answer now."
                    ),
                },
            ]
        payload: dict[str, Any] = {
            "model": model_name,
            "messages": request_messages,
            "stream": False,
            "temperature": temperature,
        }
        if mode == "native" and tools_enabled:
            payload["tools"] = formatted_tools
        if max_tokens is not None:
            payload["max_tokens"] = max_tokens
        if seed is not None:
            payload["seed"] = seed

        resp = await model_chat_completion(
            auth_type=provider_auth_type,
            base_url=base_url,
            headers=headers,
            payload=payload,
            timeout=float(timeout),
        )
        if resp.status_code >= 400:
            body = resp.text[:500]
            err_lower = body.lower()
            if any(k in err_lower for k in ("invalid tool", "does not support tool", "tool use is not supported")):
                if mode == "native" and bool(profile.get("fallback_enabled", True)):
                    log.warning(
                        "Model %s rejected native tools; using its Shogun text adapter",
                        model_name,
                    )
                    fallback_profile = dict(profile)
                    fallback_profile.update({"mode": "text", "adapter_id": "shogun_text_v1"})
                    return await _call_llm_with_tools(
                        messages,
                        model_name,
                        base_url,
                        headers,
                        timeout=timeout,
                        max_tokens=max_tokens,
                        temperature=temperature,
                        seed=seed,
                        provider_auth_type=provider_auth_type,
                        tools=tools,
                        tool_executor=tool_executor,
                        max_tool_rounds=max_tool_rounds,
                        max_tool_calls=max_tool_calls,
                        governance_context=governance_context,
                        tool_profile=fallback_profile,
                    )
            raise ValueError(f"LLM API error {resp.status_code}: {body}")

        data = resp.json()
        choices = data.get("choices", [])
        if not choices:
            raise ValueError("LLM returned no choices")

        message_obj = choices[0].get("message", {})
        content = message_obj.get("content", "")
        if isinstance(content, list):
            content = "\n".join(
                str(part.get("text") or "") for part in content if isinstance(part, dict)
            )
        if not tools_enabled:
            canonical_calls = []
        elif mode == "native":
            canonical_calls = normalize_native_tool_calls(data)
        else:
            canonical_calls = normalize_text_tool_calls(str(content or ""), allowed_tool_names)

        if not canonical_calls:
            if not content:
                raise ValueError("LLM returned empty content and no tool calls")
            return str(content)

        if mode == "native":
            current_messages.append({
                "role": "assistant",
                "content": content,
                "tool_calls": message_obj.get("tool_calls", []),
            })
        else:
            current_messages.append({"role": "assistant", "content": str(content or "")})

        if canonical_calls:
            from shogun.services.content_wrapper import wrap_tool_result
            from shogun.services.tool_gate import GateAction, check_tool_access

            posture_permissions = dict((governance_context or {}).get("permissions") or {})
            posture_tier = (
                (governance_context or {}).get("posture_level")
                or posture_permissions.get("active_tier")
                or (governance_context or {}).get("active_tier")
                or "standard"
            )
            gate_mode = "ronin_desktop" if posture_tier == "ronin" else str(posture_tier)
            if gate_mode not in {"standard", "campaign", "ronin_browser", "ronin_desktop"}:
                gate_mode = "standard"
            campaign_preset = posture_permissions.get("active_campaign_preset")
            if isinstance(campaign_preset, str) and campaign_preset:
                try:
                    from shogun.services.campaign_presets import get_preset

                    campaign_preset = get_preset(campaign_preset)
                except Exception as preset_error:
                    log.warning("Could not resolve Campaign preset for Samurai tool call: %s", preset_error)
                    campaign_preset = None

            for call in canonical_calls:
                call_id = str(call.get("id") or f"call-{uuid.uuid4().hex[:10]}")
                func_name = str(call.get("tool") or "")
                args = call.get("arguments") or {}
                cache_key = json.dumps(
                    {"tool": func_name, "arguments": args},
                    ensure_ascii=False,
                    sort_keys=True,
                    default=str,
                )

                if cache_key in tool_result_cache:
                    res_str = tool_result_cache[cache_key]
                elif max_tool_calls is not None and tool_calls_used >= max_tool_calls:
                    force_finalize = True
                    res_str = json.dumps({
                        "status": "budget_complete",
                        "message": (
                            f"The governed tool-call budget ({max_tool_calls}) is complete. "
                            "Use the evidence already gathered and return the final answer."
                        ),
                    })
                elif func_name not in allowed_tool_names:
                    tool_calls_used += 1
                    res_str = json.dumps({
                        "status": "blocked",
                        "message": f"Tool '{func_name}' is not in this Samurai node's scoped tool set.",
                    })
                else:
                    tool_calls_used += 1
                    try:
                        gate_decision = await check_tool_access(
                            mode=gate_mode,
                            tool_name=func_name,
                            args=args,
                            campaign_preset=campaign_preset if isinstance(campaign_preset, dict) else None,
                        )
                    except Exception as gate_error:
                        # Governance failures are never permission grants.
                        log.warning("ToolGate evaluation error for %s: %s", func_name, gate_error)
                        res_str = json.dumps({
                            "status": "blocked",
                            "message": f"ToolGate could not safely evaluate '{func_name}'.",
                        })
                    else:
                        if gate_decision.action == GateAction.BLOCK:
                            res_str = json.dumps({
                                "status": "blocked",
                                "message": f"ToolGate blocked '{func_name}': {gate_decision.reason}",
                            })
                        elif gate_decision.action == GateAction.CONFIRM:
                            # AgentFlow runs have no interactive confirmation
                            # channel. Never interpret CONFIRM as ALLOW.
                            res_str = json.dumps({
                                "status": "permission_required",
                                "message": (
                                    f"Tool '{func_name}' requires interactive approval and was not executed."
                                ),
                                "reason": gate_decision.reason,
                            })
                        else:
                            try:
                                async with async_session_factory() as session:
                                    res_str = await tool_executor(func_name, args, session)
                            except Exception as exec_error:
                                res_str = json.dumps({
                                    "status": "error",
                                    "message": f"Tool execution failed for '{func_name}': {exec_error}",
                                })

                tool_result_cache.setdefault(cache_key, str(res_str))
                if max_tool_calls is not None and tool_calls_used >= max_tool_calls:
                    force_finalize = True

                res_str = wrap_tool_result(func_name, str(res_str))
                if mode == "native":
                    current_messages.append({
                        "role": "tool",
                        "tool_call_id": call_id,
                        "name": func_name,
                        "content": res_str,
                    })
                else:
                    current_messages.append({
                        "role": "user",
                        "content": (
                            "<tool_result>"
                            + json.dumps(
                                {"tool": func_name, "result": res_str},
                                ensure_ascii=False,
                            )
                            + "</tool_result>\nContinue the task using this real result."
                        ),
                    })

    raise RuntimeError(f"Samurai node could not finalize after {max_tool_rounds} governed tool rounds")


async def _call_llm_chain_with_tools(
    messages: list[dict],
    model_chain: list[tuple[ModelProvider, str, str, dict]],
    *,
    timeout: int,
    retry_count: int,
    context: str,
    max_tokens: int | None = None,
    routing_context: dict[str, Any] | None = None,
    usage_session: AsyncSession | None = None,
    tools: list[dict] | None = None,
    tool_executor: Callable | None = None,
    max_tool_rounds: int = 6,
    max_tool_calls: int | None = None,
    governance_context: dict[str, Any] | None = None,
) -> str:
    """Call each model in order with tool execution support."""
    if not tools or not tool_executor or getattr(_call_llm_chain, "__name__", "") != "_call_llm_chain":
        return await _call_llm_chain(
            messages,
            model_chain,
            timeout=timeout,
            retry_count=retry_count,
            context=context,
            max_tokens=max_tokens,
            routing_context=routing_context,
            usage_session=usage_session,
        )

    last_error: Exception | None = None
    last_provider: ModelProvider | None = None
    last_model = "unknown-model"
    input_characters = len(json.dumps(messages, ensure_ascii=False, default=str))
    tool_profiles = (routing_context or {}).get("tool_calling_profiles") or {}
    for model_index, chain_item in enumerate(model_chain):
        if isinstance(chain_item, (list, tuple)) and len(chain_item) >= 4:
            _provider, model_name, base_url, headers = chain_item[:4]
        else:
            _provider = chain_item
            model_name = getattr(_provider, "model_name", "unknown-model")
            base_url = getattr(_provider, "base_url", "http://localhost:8000")
            headers = getattr(_provider, "headers", {})

        last_provider = _provider
        last_model = model_name
        for attempt in range(1 + retry_count):
            started = time.perf_counter()
            try:
                route_key = f"{getattr(_provider, 'id', '')}:{model_name}"
                tool_profile = tool_profiles.get(route_key) or infer_tool_calling_profile(
                    model_name,
                    str(getattr(_provider, "provider_type", "") or ""),
                    tool_capability=True,
                )
                model_parameters = ((routing_context or {}).get("request_parameters") or {}).get(route_key) or {}
                temperature = max(0.0, min(2.0, float(model_parameters.get("temperature", 0.3))))
                configured_seed = (routing_context or {}).get("flow_seed")
                seed_match = str((routing_context or {}).get("flow_seed_model_id") or "").strip()
                physical_model = f"{getattr(_provider, 'id', '')}:{model_name}"
                provider_model = f"{getattr(_provider, 'provider_type', '')}/{model_name}"
                seed = (
                    int(configured_seed)
                    if configured_seed is not None
                    and (not seed_match or seed_match in {model_name, physical_model, provider_model})
                    else None
                )
                result = await _call_llm_with_tools(
                    messages,
                    model_name,
                    base_url,
                    headers,
                    timeout=timeout,
                    max_tokens=max_tokens,
                    temperature=temperature,
                    seed=seed,
                    provider_auth_type=str(getattr(_provider, "auth_type", "") or ""),
                    tools=tools,
                    tool_executor=tool_executor,
                    max_tool_rounds=max_tool_rounds,
                    max_tool_calls=max_tool_calls,
                    governance_context=governance_context,
                    tool_profile=tool_profile,
                )
                await _record_model_usage(
                    _provider,
                    model_name,
                    messages,
                    result,
                    time.perf_counter() - started,
                    routing_context,
                    usage_session,
                    success=True,
                )
                return result
            except Exception as exc:
                last_error = exc
                await _record_model_usage(
                    _provider,
                    model_name,
                    messages,
                    "",
                    time.perf_counter() - started,
                    routing_context,
                    usage_session,
                    success=False,
                    error=str(exc),
                )
                log.warning(
                    "%s model '%s' failed (attempt %d/%d, timeout=%ss): %s",
                    context,
                    model_name,
                    attempt + 1,
                    1 + retry_count,
                    timeout,
                    exc,
                )
                if attempt < retry_count:
                    await asyncio.sleep(2**attempt)

        if model_index + 1 < len(model_chain):
            next_item = model_chain[model_index + 1]
            next_model = (
                next_item[1]
                if isinstance(next_item, (list, tuple)) and len(next_item) > 1
                else "fallback-model"
            )
            reason = (
                f"timeout after {timeout}s"
                if isinstance(last_error, (httpx.TimeoutException, asyncio.TimeoutError))
                else str(last_error)[:300]
            )
            from shogun.services.notification_service import notify_model_fallback

            await notify_model_fallback(
                from_model=model_name,
                to_model=next_model,
                reason=reason,
                context=context,
                timeout_seconds=timeout,
            )

    if last_error is None:
        raise ValueError(f"{context} failed without a model response")
    provider_name = (
        getattr(last_provider, "name", None)
        or getattr(last_provider, "provider_type", None)
        or "unknown-provider"
    )
    raise ModelCallError(
        context=context,
        provider=str(provider_name),
        model=last_model,
        timeout=timeout,
        cause=last_error,
        input_characters=input_characters,
    ) from last_error


def _is_structured_rows_failure(error: Exception) -> bool:
    """Return whether a model responded but violated the row contract."""
    cause = error.__cause__ if isinstance(error, ModelCallError) else error
    if isinstance(cause, json.JSONDecodeError) or (
        isinstance(error, ModelCallError) and error.cause_type == "JSONDecodeError"
    ):
        return True
    text = f"{error} {cause or ''}".lower()
    return any(
        marker in text
        for marker in (
            "did not submit rows",
            "valid json rows",
            "must contain exactly",
            "requires exactly",
            "row matrix",
            "structured json",
            "expecting value",
            "unterminated string",
            "expecting ',' delimiter",
            "extra data",
        )
    )


async def _call_llm_chain_rows_with_fallback(
    messages: list[dict],
    model_chain: list[tuple[ModelProvider, str, str, dict]],
    *,
    timeout: int,
    retry_count: int,
    context: str,
    expected_width: int | None,
    max_tokens: int | None,
    routing_context: dict[str, Any] | None,
    governance_context: dict[str, Any] | None,
    row_validator: Callable[[list[list[Any]]], None] | None = None,
) -> list[list[Any]]:
    """Return an exact row matrix or a typed error suitable for chunk subdivision.

    ``_call_llm_chain_rows`` already tries the model's native row-submission
    tool and Shogun's text adapter. General Samurai tools cannot repair an
    in-memory response and previously turned a recoverable malformed chunk
    into a terminal prose response. Preserve infrastructure errors, but tag a
    structural response failure so the caller can retry only that source chunk
    at a smaller size.
    """
    try:
        return await _call_llm_chain_rows(
            messages,
            model_chain,
            timeout=timeout,
            retry_count=retry_count,
            context=context,
            expected_width=expected_width,
            max_tokens=max_tokens,
            routing_context=routing_context,
            row_validator=row_validator,
        )
    except Exception as original_error:
        underlying_error = (
            original_error.__cause__
            if isinstance(original_error, ModelCallError)
            else original_error
        )
        if isinstance(underlying_error, IncompleteMatrixOutputError):
            raise underlying_error
        if not _is_structured_rows_failure(original_error):
            raise
        raise MalformedMatrixOutputError(
            f"{context} returned malformed structured rows after both its native row protocol "
            "and Shogun's structured text adapter were attempted."
        ) from original_error


async def _record_model_usage(
    provider: ModelProvider,
    model_name: str,
    messages: list[dict],
    output: str,
    elapsed_seconds: float,
    routing_context: dict[str, Any] | None,
    session: AsyncSession | None,
    *,
    success: bool,
    error: str | None = None,
) -> None:
    """Record best-effort usage without making telemetry a runtime dependency."""
    if not routing_context:
        return
    try:
        from shogun.schemas.model_router import ModelUsageCreate
        from shogun.services.model_router import ModelUsageLogger

        input_tokens = max(1, len(json.dumps(messages, ensure_ascii=False, default=str)) // 4)
        body = ModelUsageCreate(
            routing_decision_id=routing_context.get("id"),
            stack_run_id=routing_context.get("stack_run_id"),
            model_id=model_name,
            provider=provider.provider_type,
            input_tokens=input_tokens,
            output_tokens=len(output) // 4,
            latency_ms=max(0, int(elapsed_seconds * 1000)),
            success=success,
            error_json={"message": error[:500]} if error else {},
        )
        if session is not None:
            await ModelUsageLogger(session).log(body)
            return
        async with async_session_factory() as telemetry_session:
            await ModelUsageLogger(telemetry_session).log(body)
            await telemetry_session.commit()
    except Exception as exc:
        log.debug("Model usage telemetry could not be recorded: %s", exc)


# ═══════════════════════════════════════════════════════════════
# GRAPH UTILITIES
# ═══════════════════════════════════════════════════════════════


def _topological_sort(
    nodes: list[AgentFlowNode],
    edges: list[AgentFlowEdge],
) -> list[list[str]]:
    """Sort nodes into execution layers (Kahn's algorithm).

    Returns a list of layers, where each layer contains node IDs that can
    be executed in parallel.

    Raises ValueError if the graph contains a cycle.
    """
    node_ids = {str(n.id) for n in nodes}

    # Build adjacency and in-degree
    in_degree: dict[str, int] = {nid: 0 for nid in node_ids}
    adjacency: dict[str, list[str]] = {nid: [] for nid in node_ids}

    for edge in edges:
        src = str(edge.source_node_id)
        tgt = str(edge.target_node_id)
        if src in node_ids and tgt in node_ids:
            adjacency[src].append(tgt)
            in_degree[tgt] += 1

    # Kahn's algorithm with layer tracking
    queue: deque[str] = deque()
    for nid, deg in in_degree.items():
        if deg == 0:
            queue.append(nid)

    layers: list[list[str]] = []
    visited = 0

    while queue:
        # All nodes currently in the queue form one parallel layer
        layer = list(queue)
        queue.clear()
        layers.append(layer)
        visited += len(layer)

        for nid in layer:
            for neighbor in adjacency[nid]:
                in_degree[neighbor] -= 1
                if in_degree[neighbor] == 0:
                    queue.append(neighbor)

    if visited != len(node_ids):
        raise ValueError("Agent Flow contains a cycle — cannot execute. Remove circular dependencies and try again.")

    return layers


def _mark_downstream_skipped(
    node_id: str,
    edge_by_source: dict[str, list[tuple[str, str | None]]],
    skipped: set[str],
) -> None:
    """Recursively mark all downstream nodes as skipped."""
    for target_id, _ in edge_by_source.get(node_id, []):
        if target_id not in skipped:
            skipped.add(target_id)
            _mark_downstream_skipped(target_id, edge_by_source, skipped)


# ═══════════════════════════════════════════════════════════════
# STATE MANAGEMENT HELPERS
# ═══════════════════════════════════════════════════════════════


async def _record_node_artifact(
    run_id: uuid.UUID,
    node_id: str,
    artifact_path: str,
) -> None:
    """Associate a generated workspace artifact with its run node."""
    async with _run_state_lock(run_id):
        async with async_session_factory() as session:
            result = await session.execute(select(AgentFlowRun).where(AgentFlowRun.id == run_id))
            run = result.scalar_one_or_none()
            if not run:
                return

            states = dict(run.node_states or {})
            node_state = dict(states.get(node_id, {}))
            node_state["artifact_path"] = artifact_path
            states[node_id] = node_state
            run.node_states = states
            artifacts = list(run.artifacts or [])
            if not any(item.get("path_or_ref") == artifact_path for item in artifacts if isinstance(item, dict)):
                artifacts.append(
                    {
                        "artifact_type": "file",
                        "path_or_ref": artifact_path,
                        "created_by_run_id": str(run_id),
                        "created_by_node_id": node_id,
                    }
                )
            run.artifacts = artifacts
            from sqlalchemy.orm.attributes import flag_modified

            flag_modified(run, "node_states")
            flag_modified(run, "artifacts")
            await session.commit()


async def _update_node_state(
    run_id: uuid.UUID,
    node_id: str,
    status: str,
    output: Any = None,
    error: str | None = None,
    failure_event_id: str | None = None,
) -> None:
    """Update a single node's execution state in the run record."""
    async with _run_state_lock(run_id):
        async with async_session_factory() as session:
            result = await session.execute(select(AgentFlowRun).where(AgentFlowRun.id == run_id))
            run = result.scalar_one_or_none()
            if not run:
                return

            states = dict(run.node_states or {})
            now = datetime.now(timezone.utc).isoformat()

            node_state = dict(states.get(node_id, {}))
            node_state["status"] = status

            if status == "running":
                node_state["started_at"] = now
            elif status in ("completed", "failed", "skipped"):
                node_state["completed_at"] = now

            if output is not None:
                # Preserve the complete result so View Result and run history match
                # the generated artifact exactly.
                node_state["output"] = str(output)
            if error:
                node_state["error"] = error[:2000]
            if failure_event_id:
                node_state["failure_event_id"] = failure_event_id

            states[node_id] = node_state
            run.node_states = states
            from sqlalchemy.orm.attributes import flag_modified

            flag_modified(run, "node_states")
            await session.commit()


async def _update_node_progress(
    run_id: uuid.UUID,
    node_id: str,
    completed: int,
    total: int,
) -> None:
    """Persist measurable node progress without changing its execution status."""
    async with _run_state_lock(run_id):
        async with async_session_factory() as session:
            run = await session.get(AgentFlowRun, run_id)
            if not run:
                return

            states = dict(run.node_states or {})
            node_state = dict(states.get(node_id) or {})
            if node_state.get("status") != "running":
                return

            safe_total = max(1, int(total))
            safe_completed = max(0, min(int(completed), safe_total))
            node_state["progress_percent"] = round((safe_completed / safe_total) * 100)
            node_state["progress_completed"] = safe_completed
            node_state["progress_total"] = safe_total
            states[node_id] = node_state
            run.node_states = states
            from sqlalchemy.orm.attributes import flag_modified

            flag_modified(run, "node_states")
            await session.commit()


async def _record_node_failure_event(
    run_id: uuid.UUID,
    node: AgentFlowNode,
    error: Exception,
) -> str | None:
    """Write a deep-linkable audit event for an AgentFlow node failure."""
    try:
        from shogun.services.event_logger import EventLogger

        error_message = _error_message(error)
        detail = {
            "flow_id": str(node.flow_id),
            "flow_run_id": str(run_id),
            "node_id": str(node.id),
            "node_label": node.label,
            "node_type": node.node_type,
            "error_type": getattr(error, "cause_type", type(error).__name__),
            "error": error_message,
        }
        for attribute, key in (
            ("provider", "provider"),
            ("model", "model"),
            ("timeout_seconds", "timeout_seconds"),
            ("input_characters", "input_characters"),
            ("estimated_input_tokens", "estimated_input_tokens"),
        ):
            value = getattr(error, attribute, None)
            if value is not None:
                detail[key] = value
        source_intelligence = getattr(error, "source_intelligence", None)
        if isinstance(source_intelligence, dict):
            detail["source_intelligence"] = deepcopy(source_intelligence)

        return await EventLogger.emit_incident_event(
            "agent_flow.node.failed",
            f"AgentFlow node '{node.label}' failed",
            result="error",
            severity="error",
            risk_score="medium",
            trace_id=str(run_id),
            detail=detail,
        )
    except Exception:
        log.exception("Failed to write audit event for AgentFlow node %s", node.id)
        return None


async def _fail_run(
    run_id: uuid.UUID,
    error_message: str,
    node_states_override: bool = False,
) -> None:
    """Mark a run as failed."""
    async with async_session_factory() as session:
        result = await session.execute(select(AgentFlowRun).where(AgentFlowRun.id == run_id))
        run = result.scalar_one_or_none()
        if not run:
            return

        run.status = "failed"
        run.completed_at = datetime.now(timezone.utc)
        run.error_message = error_message[:2000]
        await session.commit()

    await _sync_run_edge_status(run_id, "failed")
    log.error("Flow run %s FAILED: %s", run_id, error_message)


async def _complete_run(
    run_id: uuid.UUID,
    result_summary: dict[str, Any],
) -> None:
    """Mark a run as completed with results."""
    async with async_session_factory() as session:
        result = await session.execute(select(AgentFlowRun).where(AgentFlowRun.id == run_id))
        run = result.scalar_one_or_none()
        if not run:
            return

        run.status = "completed"
        run.completed_at = datetime.now(timezone.utc)

        # Preserve structured child handoffs alongside the legacy string
        # summary used by the existing run-history UI.
        run.output_payload = _json_object(result_summary, "Flow output")

        # Preserve final reports in full. This is the authoritative content
        # shown by View Result and must not differ from the workspace artifact.
        complete_summary = {}
        for k, v in result_summary.items():
            complete_summary[k] = str(v)
        run.result_summary = complete_summary
        from sqlalchemy.orm.attributes import flag_modified

        flag_modified(run, "result_summary")
        await session.commit()

    await _sync_run_edge_status(run_id, "completed")
    log.info("Flow run %s COMPLETED", run_id)


async def _cancel_run_record(run_id: uuid.UUID, reason: str) -> None:
    async with async_session_factory() as session:
        run = await session.get(AgentFlowRun, run_id)
        if run and run.status not in {"completed", "failed", "cancelled"}:
            run.status = "cancelled"
            run.completed_at = datetime.now(timezone.utc)
            run.error_message = reason[:2000]
            await session.commit()
    await _sync_run_edge_status(run_id, "cancelled")


async def _sync_run_edge_status(run_id: uuid.UUID, status: str) -> None:
    async with async_session_factory() as session:
        result = await session.execute(select(AgentFlowRunEdge).where(AgentFlowRunEdge.child_run_id == run_id))
        edge = result.scalar_one_or_none()
        if edge:
            edge.status = status
            if status in {"completed", "failed", "cancelled"}:
                edge.completed_at = datetime.now(timezone.utc)
            await session.commit()


def _truncate(text: str, max_len: int) -> str:
    """Truncate text with an ellipsis marker."""
    if len(text) <= max_len:
        return text
    return text[: max_len - 20] + "\n\n[...truncated...]"
