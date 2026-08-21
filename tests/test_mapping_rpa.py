"""Deterministic Mapping / RPA engine and AgentFlow handoff tests."""

from __future__ import annotations

from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from shogun.api.deps import get_db
from shogun.config import settings
from shogun.db.base import Base
from shogun.db.models.mapping_template import MappingTemplate
from shogun.engine import flow_engine
from shogun.engine.flow_engine import (
    _exec_mapping_rpa,
    _exec_office,
    _mapping_payload_from_predecessors,
    _write_openpyxl_mapping,
)
from shogun.mapping.engine import execute_mapping
from shogun.mapping.errors import MappingFieldMissing, MappingSchemaError, MappingTypeError
from shogun.mapping.schema import MappingConfig
from shogun.schemas.agent_flow import AgentFlowNodeCreate
from shogun.services.transformation_profile_registry import profile_content_hash


def _config(**overrides):
    config = {
        "version": 1,
        "name": "Supplier PDF to Excel",
        "mode": "strict",
        "output": {"type": "table", "start_cell": "A1"},
        "mappings": [
            {"source": "article_number", "target": "A", "type": "string", "required": True, "transform": ["trim"]},
            {"source": "description", "target": "B", "type": "string", "required": True},
            {"source": "quantity", "target": "C", "type": "integer", "required": True, "transform": ["convert"]},
            {"source": "unit_price", "target": "D", "type": "decimal", "transform": ["decimal_normalize"]},
            {"source": "currency", "target": "E", "type": "string", "default": "EUR", "transform": ["uppercase"]},
        ],
    }
    config.update(overrides)
    return config


def test_canonical_pdf_to_excel_mapping_is_typed_and_ordered():
    result = execute_mapping(
        {
            "article_number": " 68947124 ",
            "description": "Pump Housing",
            "quantity": "4",
            "unit_price": "129,95",
            "currency": "eur",
        },
        _config(),
    )

    assert result["status"] == "SUCCESS"
    assert result["rows"] == [["68947124", "Pump Housing", 4, 129.95, "EUR"]]
    assert result["records_written"] == 1


def test_array_mapping_nested_path_and_lineage():
    config = _config(
        input_path="items",
        output={"type": "range", "start_cell": "A12", "sheet": "Products"},
    )
    result = execute_mapping(
        {
            "items": [
                {"article_number": "A1", "description": "Pump", "quantity": "2", "unit_price": "10", "source_page": 4},
                {
                    "article_number": "A2",
                    "description": "Valve",
                    "quantity": "5",
                    "unit_price": "4,50",
                    "source_page": 5,
                },
            ]
        },
        config,
    )

    assert result["rows"] == [["A1", "Pump", 2, 10, "EUR"], ["A2", "Valve", 5, 4.5, "EUR"]]
    assert result["start_cell"] == "A12"
    assert result["sheet"] == "Products"
    assert [item["page"] for item in result["lineage"]] == [4, 5]


def test_aliases_are_explicit_and_deterministic():
    result = execute_mapping(
        {"Artikel-Nr": "68947124", "description": "Pump", "quantity": 4, "unit_price": 1},
        _config(aliases={"article_number": ["Artikel-Nr", "SKU"]}),
    )
    assert result["rows"][0][0] == "68947124"


def test_missing_required_field_fails_before_output():
    with pytest.raises(MappingFieldMissing) as error:
        execute_mapping({"description": "Pump", "quantity": 4, "unit_price": 1}, _config())
    assert error.value.code == "VALIDATION_FAILED"
    assert error.value.field == "article_number"


def test_strict_mode_rejects_unknown_fields():
    payload = {
        "article_number": "A1",
        "description": "Pump",
        "quantity": 1,
        "unit_price": 2,
        "unexpected": "do not silently discard",
    }
    with pytest.raises(MappingSchemaError, match="Unknown fields"):
        execute_mapping(payload, _config())


def test_lenient_mode_converts_values_but_strict_requires_explicit_conversion():
    config = {
        "mode": "strict",
        "output": {"type": "table"},
        "mappings": [{"source": "quantity", "target": "A", "type": "integer", "required": True}],
    }
    with pytest.raises(MappingTypeError):
        execute_mapping({"quantity": "5"}, config)
    config["mode"] = "lenient"
    assert execute_mapping({"quantity": "5"}, config)["rows"] == [[5]]


def test_serialized_optional_rule_does_not_gain_an_implicit_default():
    config = MappingConfig.model_validate(
        {
            "mode": "lenient",
            "output": {"type": "table"},
            "mappings": [{"source": "quantity", "target": "A", "type": "integer"}],
        }
    )
    restored = MappingConfig.model_validate(config.model_dump(mode="json"))
    assert restored.mappings[0].has_default is False


def test_cell_mapping_and_openpyxl_handoff_preserve_types():
    config = {
        "mode": "lenient",
        "output": {"type": "cells", "sheet": "Invoice"},
        "mappings": [
            {"source": "customer", "target": "B4", "type": "string", "required": True},
            {"source": "quantity", "target": "E3", "type": "integer", "required": True},
        ],
    }
    result = execute_mapping({"customer": "Customer A", "quantity": "4"}, config)
    assert result["cells"] == {"B4": "Customer A", "E3": 4}

    workbook = Workbook()
    sheet = workbook.active
    _write_openpyxl_mapping(sheet, result)
    assert sheet["B4"].value == "Customer A"
    assert sheet["E3"].value == 4


def test_openpyxl_handoff_writes_safe_headers_and_neutralizes_formulas():
    workbook = Workbook()
    sheet = workbook.active
    payload = {
        "type": "table",
        "start_cell": "B2",
        "include_headers": True,
        "headers": ["=malicious header", "name"],
        "rows": [["@SUM(A1:A2)", "Pump"]],
    }

    written = _write_openpyxl_mapping(sheet, payload)

    assert written == 1
    assert sheet["B2"].value == "'=malicious header"
    assert sheet["C2"].value == "name"
    assert sheet["B3"].value == "'@SUM(A1:A2)"
    assert sheet["C3"].value == "Pump"


def test_openpyxl_handoff_rejects_nonscalar_cells():
    workbook = Workbook()
    sheet = workbook.active
    with pytest.raises(ValueError, match="must be scalar"):
        _write_openpyxl_mapping(
            sheet,
            {"type": "table", "rows": [[{"nested": "not a cell"}]]},
        )


def test_safe_formula_and_duplicate_replace():
    config = {
        "mode": "lenient",
        "duplicate_key": "sku",
        "duplicate_policy": "replace",
        "output": {"type": "table"},
        "mappings": [
            {"source": "sku", "target": "A", "type": "string", "required": True},
            {"source": "quantity", "target": "B", "type": "integer", "required": True},
            {"source": "unit_price", "target": "C", "type": "decimal", "required": True},
            {"expression": "quantity * unit_price", "target": "D", "type": "decimal"},
        ],
    }
    result = execute_mapping(
        [
            {"sku": "A1", "quantity": 1, "unit_price": 10},
            {"sku": "A1", "quantity": 2, "unit_price": 10},
        ],
        config,
    )
    assert result["rows"] == [["A1", 2, 10, 20]]


def test_record_skip_produces_partial_without_corrupting_valid_rows():
    config = {
        "mode": "lenient",
        "on_record_error": "skip",
        "output": {"type": "table"},
        "mappings": [{"source": "quantity", "target": "A", "type": "integer", "required": True}],
    }
    result = execute_mapping([{"quantity": "5"}, {"quantity": "four"}], config)
    assert result["status"] == "PARTIAL"
    assert result["rows"] == [[5]]
    assert result["records_failed"] == 1


def test_office_handoff_rejects_validation_failed_payload():
    with pytest.raises(ValueError, match="not writable"):
        _mapping_payload_from_predecessors(
            {
                "mapping": {
                    "__shogun_mapping_output__": True,
                    "status": "VALIDATION_FAILED",
                    "errors": [{"message": "article_number is missing"}],
                }
            }
        )


@pytest.mark.anyio
async def test_agentflow_executor_routes_structured_mapping_failures():
    result = await _exec_mapping_rpa(
        _config(),
        {"samurai": {"description": "Pump"}},
        flow_id="flow-1",
        node_id="mapping-1",
    )
    assert result["status"] == "VALIDATION_FAILED"
    assert result["errors"][0]["field"] == "article_number"


def test_agentflow_node_schema_normalizes_mapping_configuration():
    node = AgentFlowNodeCreate(node_type="mapping_rpa", label="Map", config=_config())
    assert node.config["mappings"][4]["has_default"] is True
    assert node.config["output"]["type"] == "table"


def test_agentflow_node_schema_preserves_explicit_transformation_profile():
    profile = {
        "id": "supplier_report_v1",
        "adapter": "sectioned_record_matrix_v1",
        "parameters": {"section_pattern": r"(?m)^Record: (?P<section_id>\S+)"},
    }
    node = AgentFlowNodeCreate(
        node_type="mapping_rpa",
        label="Map",
        config=_config(transformation_profile=profile),
    )

    assert node.config["transformation_profile"] == {
        **profile,
        "model_fallback": False,
        "registry_version": None,
        "content_hash": None,
    }


def test_contract_mapping_requires_a_profile_but_not_mapping_rules():
    profile = {
        "id": "supplier_report_v2",
        "adapter": "sectioned_record_matrix_v1",
        "parameters": {"section_pattern": r"(?m)^Record: (?P<section_id>\S+)"},
    }

    config = MappingConfig.model_validate(
        {
            "name": "Supplier extraction contract",
            "execution_mode": "contract",
            "transformation_profile": profile,
        }
    )

    assert config.mappings == []
    assert config.transformation_profile is not None
    assert config.transformation_profile.model_fallback is False
    with pytest.raises(ValueError, match="contract execution requires a transformation_profile"):
        MappingConfig.model_validate({"execution_mode": "contract"})
    with pytest.raises(ValueError, match="transform execution requires at least one mapping rule"):
        MappingConfig.model_validate({})


def test_transformation_profile_registry_pin_is_all_or_nothing_and_normalized():
    digest = "AB" * 32
    profile = MappingConfig.model_validate(
        {
            "execution_mode": "contract",
            "transformation_profile": {
                "id": "private_sectioned_report_v2",
                "adapter": "sectioned_record_matrix_v1",
                "registry_version": 2,
                "content_hash": digest,
            },
        }
    ).transformation_profile

    assert profile is not None
    assert profile.is_registry_pinned is True
    assert profile.content_hash == digest.lower()
    with pytest.raises(ValueError, match="registry_version and content_hash must be supplied together"):
        MappingConfig.model_validate(
            {
                "execution_mode": "contract",
                "transformation_profile": {
                    "id": "private_sectioned_report_v2",
                    "adapter": "sectioned_record_matrix_v1",
                    "registry_version": 2,
                },
            }
        )


@pytest.mark.anyio
async def test_contract_mapping_resolves_minimal_pinned_registry_reference(monkeypatch):
    definition = {
        "id": "private_sectioned_report_v2",
        "adapter": "sectioned_record_matrix_v1",
        "parameters": {"section_pattern": r"(?m)^Record: (?P<section_id>\S+)"},
    }
    digest = profile_content_hash(definition)
    evidence = {
        "profile_id": definition["id"],
        "version": 2,
        "content_hash": digest,
        "status": "active",
        "adapter_id": definition["adapter"],
        "adapter_status": "available",
        "version_id": "version-2",
    }

    async def resolve(profile):
        assert profile.id == definition["id"]
        assert profile.adapter == definition["adapter"]
        assert profile.registry_version == 2
        assert profile.content_hash == digest
        assert profile.parameters == {}
        return definition, evidence

    monkeypatch.setattr(flow_engine, "_resolve_registered_enterprise_profile", resolve)
    result = await flow_engine._exec_mapping_rpa(
        {
            "name": "SAP registry contract",
            "execution_mode": "contract",
            "transformation_profile": {
                "id": definition["id"],
                "adapter": "sectioned_record_matrix_v1",
                "registry_version": 2,
                "content_hash": digest,
            },
        },
        {},
        flow_id="flow-1",
        node_id="contract-1",
    )

    assert result == {
        "__shogun_mapping_profile_contract__": True,
        "status": "SUCCESS",
        "type": "transformation_profile",
        "profile_id": definition["id"],
        "adapter": "sectioned_record_matrix_v1",
        "registry_version": 2,
        "content_hash": digest,
        "resolved_definition": definition,
        "registry_evidence": evidence,
    }


@pytest.mark.anyio
async def test_unpinned_contract_only_accepts_exact_active_registry_snapshot(monkeypatch):
    definition = {
        "id": "private_sectioned_report_v2",
        "adapter": "sectioned_record_matrix_v1",
        "parameters": {"section_pattern": r"(?m)^Record: (?P<section_id>\S+)"},
    }
    evidence = {
        "profile_id": definition["id"],
        "version": 2,
        "content_hash": profile_content_hash(definition),
        "status": "active",
        "adapter_id": definition["adapter"],
        "adapter_status": "available",
        "version_id": "version-2",
    }

    async def resolve(_profile):
        return definition, evidence

    monkeypatch.setattr(flow_engine, "_resolve_registered_enterprise_profile", resolve)
    base_config = {
        "execution_mode": "contract",
        "transformation_profile": {**definition, "model_fallback": False},
    }
    result = await flow_engine._exec_mapping_rpa(
        base_config,
        {},
        flow_id="flow-1",
        node_id="contract-1",
    )
    assert result["resolved_definition"] == definition
    assert result["registry_evidence"] == evidence

    mismatched = {
        **base_config,
        "transformation_profile": {
            **base_config["transformation_profile"],
            "parameters": {"section_pattern": "caller-controlled"},
        },
    }
    with pytest.raises(MappingSchemaError, match="does not exactly match the active registry definition"):
        await flow_engine._exec_mapping_rpa(
            mismatched,
            {},
            flow_id="flow-1",
            node_id="contract-1",
        )


def test_samurai_contract_carrier_rejects_forged_resolved_definition():
    definition = {
        "id": "private_sectioned_report_v2",
        "adapter": "sectioned_record_matrix_v1",
        "parameters": {"section_pattern": "trusted"},
    }
    digest = profile_content_hash(definition)
    profile = MappingConfig.model_validate(
        {
            "execution_mode": "contract",
            "transformation_profile": {
                "id": definition["id"],
                "adapter": definition["adapter"],
                "registry_version": 2,
                "content_hash": digest,
            },
        }
    ).transformation_profile
    assert profile is not None
    evidence = {
        "profile_id": definition["id"],
        "version": 2,
        "content_hash": digest,
        "status": "active",
        "adapter_id": definition["adapter"],
        "adapter_status": "available",
        "version_id": "version-2",
    }
    carrier = {
        "__shogun_mapping_profile_contract__": True,
        "status": "SUCCESS",
        "type": "transformation_profile",
        "profile_id": definition["id"],
        "adapter": definition["adapter"],
        "registry_version": 2,
        "content_hash": digest,
        "resolved_definition": {
            **definition,
            "parameters": {"section_pattern": "forged after resolution"},
        },
        "registry_evidence": evidence,
    }

    with pytest.raises(MappingSchemaError, match="failed its content-hash check"):
        flow_engine._trusted_contract_profile_from_carrier(
            profile,
            carrier,
            carrier_label="SAP registry contract",
        )


@pytest.mark.anyio
async def test_canonical_mapping_to_excel_file_end_to_end(tmp_path, monkeypatch):
    mapped = execute_mapping(
        {
            "article_number": "68947124",
            "description": "Pump Housing",
            "quantity": "4",
            "unit_price": "129,95",
        },
        _config(output={"type": "range", "start_cell": "A2", "sheet": "Products"}),
    )
    monkeypatch.setattr(settings, "workspace_path", tmp_path)
    monkeypatch.setattr("shogun.office.config.load_office_config", lambda: SimpleNamespace(enabled=True))

    result = await _exec_office(
        {
            "action": "excel_create",
            "output_path": "Output",
            "output_filename": "canonical.xlsx",
            "sheet_name": "Products",
        },
        "ignored textual context",
        predecessor_outputs={"mapping": mapped},
    )

    assert "canonical.xlsx" in result
    workbook = load_workbook(tmp_path / "Output" / "canonical.xlsx", data_only=True)
    try:
        sheet = workbook["Products"]
        assert [sheet.cell(2, column).value for column in range(1, 6)] == [
            "68947124",
            "Pump Housing",
            4,
            129.95,
            "EUR",
        ]
    finally:
        workbook.close()


@pytest.mark.anyio
async def test_preview_endpoint_returns_structured_validation_failure(client):
    response = await client.post(
        "/api/v1/mapping-rpa/preview",
        json={"config": _config(), "input": {"description": "Pump"}},
        headers={"X-Shogun-Infrastructure-Token": str(settings.infrastructure_admin_token or "")},
    )
    assert response.status_code == 200
    body = response.json()["data"]
    assert body["status"] == "VALIDATION_FAILED"
    assert body["errors"][0]["field"] == "article_number"


@pytest.mark.anyio
async def test_mapping_template_crud_is_versioned_and_soft_deleted(api_app, client, tmp_path):
    database_url = f"sqlite+aiosqlite:///{(tmp_path / 'mapping-templates.db').as_posix()}"
    engine = create_async_engine(database_url)
    async with engine.begin() as connection:
        await connection.run_sync(
            lambda sync_connection: Base.metadata.create_all(
                sync_connection,
                tables=[MappingTemplate.__table__],
            )
        )
    sessions = async_sessionmaker(engine, expire_on_commit=False)

    async def override_db():
        async with sessions() as session:
            yield session

    api_app.dependency_overrides[get_db] = override_db
    headers = {"X-Shogun-Infrastructure-Token": str(settings.infrastructure_admin_token or "")}
    try:
        created = await client.post(
            "/api/v1/mapping-rpa/templates",
            headers=headers,
            json={
                "name": "Supplier Product PDF",
                "scope": "private",
                "owner_id": "system",
                "config": _config(),
            },
        )
        assert created.status_code == 201
        template_id = created.json()["data"]["id"]

        updated = await client.put(
            f"/api/v1/mapping-rpa/templates/{template_id}",
            headers=headers,
            json={"description": "Version two"},
        )
        assert updated.status_code == 200
        assert updated.json()["data"]["version"] == 2

        listed = await client.get("/api/v1/mapping-rpa/templates", headers=headers)
        assert [item["id"] for item in listed.json()["data"]] == [template_id]

        deleted = await client.delete(f"/api/v1/mapping-rpa/templates/{template_id}", headers=headers)
        assert deleted.json()["data"]["deleted"] is True
        listed = await client.get("/api/v1/mapping-rpa/templates", headers=headers)
        assert listed.json()["data"] == []
    finally:
        api_app.dependency_overrides.pop(get_db, None)
        await engine.dispose()
