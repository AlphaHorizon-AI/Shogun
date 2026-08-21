import uuid
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from fastapi import UploadFile

from shogun.config import settings
from shogun.engine import flow_engine


@pytest.mark.asyncio
async def test_agent_flow_upload_uses_configured_upload_directory(tmp_path, monkeypatch):
    from shogun.api.agent_flow import upload_flow_document

    flow_id = uuid.uuid4()
    upload_root = tmp_path / "uploads"
    monkeypatch.setattr(settings, "uploads_path", upload_root)

    class FakeFlowService:
        async def get_by_id(self, requested_flow_id):
            assert requested_flow_id == flow_id
            return SimpleNamespace(id=flow_id)

    response = await upload_flow_document(
        flow_id,
        UploadFile(filename="sample-report.pdf", file=BytesIO(b"%PDF-1.4 test")),
        FakeFlowService(),
    )

    upload_dir = upload_root / "agent_flows" / flow_id.hex
    stored = list(upload_dir.glob("*.pdf"))
    assert len(stored) == 1
    assert stored[0].read_bytes() == b"%PDF-1.4 test"
    assert response.data["filename"] == "sample-report.pdf"
    assert response.data["stored_filename"] == stored[0].name
    assert response.data["path"] == str(stored[0])


@pytest.mark.asyncio
async def test_document_input_requires_an_uploaded_file():
    with pytest.raises(ValueError, match="No document was uploaded"):
        await flow_engine._exec_input({"input_type": "document"}, "")


@pytest.mark.asyncio
async def test_document_input_reports_incomplete_upload():
    with pytest.raises(ValueError, match="did not complete successfully"):
        await flow_engine._exec_input(
            {
                "input_type": "document",
                "uploaded_file": {
                    "filename": "sample-report.pdf",
                    "size": 1234,
                    "path": "",
                    "error": "Upload failed",
                },
            },
            "",
        )


@pytest.mark.asyncio
async def test_document_input_uses_bounded_format_reader(tmp_path):
    document = tmp_path / "input.txt"
    document.write_text("mapped source content", encoding="utf-8")

    result = await flow_engine._exec_input(
        {
            "input_type": "document",
            "uploaded_file": {"path": str(document), "filename": document.name},
        },
        "",
    )

    assert "[Document: input.txt]" in result
    assert "mapped source content" in result


@pytest.mark.asyncio
async def test_agent_flow_document_is_not_clipped_at_legacy_chat_limit(tmp_path):
    content = "row-data\n" * 20_000
    document = tmp_path / "large-input.txt"
    document.write_text(content, encoding="utf-8")

    result = await flow_engine._exec_input(
        {
            "input_type": "document",
            "uploaded_file": {"path": str(document), "filename": document.name},
        },
        "",
    )

    assert len(content) > 100_000
    assert content in result.replace("\r\n", "\n")


@pytest.mark.asyncio
async def test_document_input_reads_workspace_file(tmp_path, monkeypatch):
    workspace = tmp_path / "workspace"
    document = workspace / "Input" / "source.txt"
    document.parent.mkdir(parents=True)
    document.write_text("workspace source content", encoding="utf-8")
    monkeypatch.setattr(settings, "workspace_path", workspace)

    result = await flow_engine._exec_input(
        {
            "input_type": "document",
            "document_source": "workspace",
            "workspace_path": "Input/source.txt",
        },
        "",
    )

    assert "[Document: source.txt]" in result
    assert "workspace source content" in result


@pytest.mark.asyncio
async def test_document_input_blocks_workspace_escape(tmp_path, monkeypatch):
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    outside = tmp_path / "outside.txt"
    outside.write_text("must not be read", encoding="utf-8")
    monkeypatch.setattr(settings, "workspace_path", workspace)

    with pytest.raises(ValueError, match="must remain inside the configured workspace"):
        await flow_engine._exec_input(
            {
                "input_type": "document",
                "document_source": "workspace",
                "workspace_path": str(outside),
            },
            "",
        )


@pytest.mark.asyncio
async def test_document_input_reads_bound_chat_attachment(monkeypatch):
    from shogun.services import file_formats

    file_id = uuid.uuid4()

    class FakeSessionContext:
        async def __aenter__(self):
            return object()

        async def __aexit__(self, exc_type, exc, traceback):
            return False

    class FakeFileFormatService:
        def __init__(self, session=None, allowed_roots=None):
            assert session is not None

        async def read(self, *, file_id, max_chars):
            assert max_chars == settings.agent_flow_document_max_chars
            return {"filename": "attached.pdf", "content": "attachment source content"}

    monkeypatch.setattr(flow_engine, "async_session_factory", lambda: FakeSessionContext())
    monkeypatch.setattr(file_formats, "FileFormatService", FakeFileFormatService)

    result = await flow_engine._exec_input(
        {
            "input_type": "document",
            "document_source": "attachment",
            "attachment_file_id": str(file_id),
        },
        "",
    )

    assert "[Document: attached.pdf]" in result
    assert "attachment source content" in result


@pytest.mark.asyncio
async def test_samurai_node_receives_complete_predecessor_document(monkeypatch):
    predecessor = "page data\n" * 10_000
    captured: dict[str, str] = {}

    async def update_state(*_args, **_kwargs):
        return None

    async def execute_samurai(_config, context, _governance, **_kwargs):
        captured["context"] = context
        return "done"

    monkeypatch.setattr(flow_engine, "_update_node_state", update_state)
    monkeypatch.setattr(flow_engine, "_exec_samurai", execute_samurai)
    monkeypatch.setattr(flow_engine, "_finalize_node_skills", update_state)
    node = SimpleNamespace(
        id=uuid.uuid4(),
        flow_id=uuid.uuid4(),
        node_type="samurai",
        label="Extract",
        config={"task_description": "Extract all rows"},
    )
    predecessor_id = str(uuid.uuid4())
    predecessor_node = SimpleNamespace(label="Input PDF")

    result = await flow_engine._execute_single_node(
        uuid.uuid4(),
        node,
        {predecessor_id: predecessor},
        {predecessor_id: predecessor_node},
    )

    assert result == "done"
    assert predecessor in captured["context"]
    assert "[...truncated...]" not in captured["context"]


@pytest.mark.asyncio
async def test_samurai_receives_template_as_fixed_context(monkeypatch):
    captured: dict[str, str] = {}

    async def update_state(*_args, **_kwargs):
        return None

    async def execute_samurai(_config, context, _governance, **kwargs):
        captured["context"] = context
        captured["fixed"] = kwargs.get("fixed_context_str", "")
        return "done"

    monkeypatch.setattr(flow_engine, "_update_node_state", update_state)
    monkeypatch.setattr(flow_engine, "_exec_samurai", execute_samurai)
    monkeypatch.setattr(flow_engine, "_finalize_node_skills", update_state)
    monkeypatch.setattr(flow_engine, "_node_uses_active_skill_context", lambda *_args: False)
    template_id = str(uuid.uuid4())
    document_id = str(uuid.uuid4())
    node = SimpleNamespace(
        id=uuid.uuid4(),
        flow_id=uuid.uuid4(),
        node_type="samurai",
        label="Extract",
        config={"task_description": "Extract"},
    )
    template = {
        "__shogun_file_template__": True,
        "template_path": "Templates/output.xlsx",
        "format": "xlsx",
        "contract": "22 columns",
        "manifest": {"logical_columns": 22},
    }

    await flow_engine._execute_single_node(
        uuid.uuid4(),
        node,
        {document_id: "PDF DATA", template_id: template},
        {
            document_id: SimpleNamespace(label="Input PDF"),
            template_id: SimpleNamespace(label="Template"),
        },
    )

    assert "PDF DATA" in captured["context"]
    assert "FILE TEMPLATE CONTRACT" not in captured["context"]
    assert "FILE TEMPLATE CONTRACT" in captured["fixed"]
    assert '"logical_columns": 22' in captured["fixed"]


@pytest.mark.asyncio
async def test_one_samurai_receives_one_contract_and_all_other_predecessors(monkeypatch):
    from shogun.services.transformation_profile_registry import profile_content_hash

    captured: dict[str, object] = {}
    resolved_definition = {
        "id": "private_sectioned_report_v2",
        "adapter": "sectioned_record_matrix_v1",
        "parameters": {"section_pattern": r"(?m)^Record: (?P<section_id>\S+)"},
    }
    digest = profile_content_hash(resolved_definition)
    profile = {
        "id": resolved_definition["id"],
        "adapter": resolved_definition["adapter"],
        "parameters": {"section_pattern": "CALLER CONTROLLED AND MUST BE IGNORED"},
        "model_fallback": True,
        "registry_version": 2,
        "content_hash": digest,
        "lifecycle": "active",
    }
    registry_evidence = {
        "profile_id": resolved_definition["id"],
        "version": 2,
        "content_hash": digest,
        "status": "active",
        "adapter_id": resolved_definition["adapter"],
        "adapter_status": "available",
        "version_id": "version-2",
    }

    async def update_state(*_args, **_kwargs):
        return None

    async def execute_samurai(config, context, _governance, **kwargs):
        captured["profiles"] = config.get("_transformation_profiles")
        captured["artifacts"] = config.get("_input_artifacts")
        captured["context"] = context
        captured["fixed"] = kwargs.get("fixed_context_str")
        return "done"

    monkeypatch.setattr(flow_engine, "_update_node_state", update_state)
    monkeypatch.setattr(flow_engine, "_exec_samurai", execute_samurai)
    monkeypatch.setattr(flow_engine, "_finalize_node_skills", update_state)
    monkeypatch.setattr(flow_engine, "_node_uses_active_skill_context", lambda *_args: False)

    mapping_id = str(uuid.uuid4())
    pdf_ids = [str(uuid.uuid4()) for _ in range(3)]
    template_id = str(uuid.uuid4())
    node = SimpleNamespace(
        id=uuid.uuid4(),
        flow_id=uuid.uuid4(),
        node_type="samurai",
        label="Transform",
        config={"task_description": "Transform records"},
    )
    node_map = {
        mapping_id: SimpleNamespace(
            label="Mapping contract",
            node_type="mapping_rpa",
            config={
                "execution_mode": "contract",
                "transformation_profile": profile,
            },
        ),
        **{
            pdf_id: SimpleNamespace(
                id=pdf_id,
                label=f"PDF {index}",
                node_type="office",
                config={"action": "pdf_read"},
            )
            for index, pdf_id in enumerate(pdf_ids, start=1)
        },
        template_id: SimpleNamespace(
            id=template_id,
            label="Output template",
            node_type="file_template",
            config={},
        ),
    }
    carrier_output = {
        "__shogun_mapping_profile_contract__": True,
        "status": "SUCCESS",
        "type": "transformation_profile",
        "profile_id": profile["id"],
        "adapter": profile["adapter"],
        "registry_version": 2,
        "content_hash": digest,
        "resolved_definition": resolved_definition,
        "registry_evidence": registry_evidence,
    }
    template = {
        "__shogun_file_template__": True,
        "template_path": "Templates/output.xlsx",
        "format": "xlsx",
        "contract": "24 columns",
        "manifest": {"logical_columns": 24},
    }

    result = await flow_engine._execute_single_node(
        uuid.uuid4(),
        node,
        {
            mapping_id: carrier_output,
            pdf_ids[0]: "PDF ONE SOURCE",
            pdf_ids[1]: "PDF TWO SOURCE",
            pdf_ids[2]: "PDF THREE SOURCE",
            template_id: template,
        },
        node_map,
    )

    assert result == "done"
    assert captured["profiles"] == [resolved_definition]
    assert "CALLER CONTROLLED" not in str(captured["profiles"])
    assert "lifecycle" not in captured["profiles"][0]
    assert all(f"PDF {word} SOURCE" in captured["context"] for word in ("ONE", "TWO", "THREE"))
    assert "__shogun_mapping_profile_contract__" not in captured["context"]
    assert "__shogun_mapping_profile_contract__" not in str(captured["artifacts"])
    assert "FILE TEMPLATE CONTRACT" in captured["fixed"]


@pytest.mark.asyncio
async def test_domain_shaped_input_does_not_implicitly_activate_a_profile(monkeypatch):
    captured: dict[str, object] = {}

    async def update_state(*_args, **_kwargs):
        return None

    async def execute_samurai(config, context, _governance, **_kwargs):
        captured["profiles"] = config.get("_transformation_profiles")
        captured["context"] = context
        return "model path"

    monkeypatch.setattr(flow_engine, "_update_node_state", update_state)
    monkeypatch.setattr(flow_engine, "_exec_samurai", execute_samurai)
    monkeypatch.setattr(flow_engine, "_finalize_node_skills", update_state)
    monkeypatch.setattr(flow_engine, "_node_uses_active_skill_context", lambda *_args: False)

    source_id = str(uuid.uuid4())
    node = SimpleNamespace(
        id=uuid.uuid4(),
        flow_id=uuid.uuid4(),
        node_type="samurai",
        label="Generic extraction",
        config={"task_description": "Extract every private report record into rows"},
    )

    result = await flow_engine._execute_single_node(
        uuid.uuid4(),
        node,
        {source_id: "Object: ITEM-A\nQuantity: 12\nLine: ITEM-A ORDER-1"},
        {
            source_id: SimpleNamespace(
                id=source_id,
                label="Private report PDF",
                node_type="office",
                config={"action": "pdf_read"},
            )
        },
    )

    assert result == "model path"
    assert captured["profiles"] == []
    assert "Object: ITEM-A" in captured["context"]


@pytest.mark.asyncio
async def test_samurai_instruction_file_replaces_typed_prompt(tmp_path, monkeypatch):
    upload_root = tmp_path / "uploads"
    instruction = upload_root / "agent_flows" / uuid.uuid4().hex / "instruction.md"
    instruction.parent.mkdir(parents=True)
    instruction.write_text("Follow only these attached instructions.", encoding="utf-8")
    monkeypatch.setattr(settings, "uploads_path", upload_root)

    resolved = await flow_engine._resolve_samurai_task_description(
        {
            "task_description": "This typed prompt must not be used.",
            "instruction_file": {
                "filename": "operator-instructions.md",
                "path": str(instruction),
            },
        }
    )

    assert resolved == "Follow only these attached instructions."
    assert "typed prompt" not in resolved


@pytest.mark.asyncio
async def test_samurai_execution_resolves_instruction_before_dispatch(tmp_path, monkeypatch):
    upload_root = tmp_path / "uploads"
    instruction = upload_root / "agent_flows" / uuid.uuid4().hex / "instruction.md"
    instruction.parent.mkdir(parents=True)
    instruction.write_text("Attached Samurai task", encoding="utf-8")
    monkeypatch.setattr(settings, "uploads_path", upload_root)

    captured: dict[str, object] = {}

    async def update_state(*_args, **_kwargs):
        return None

    async def execute_samurai(config, _context, _governance, **_kwargs):
        captured["config"] = config
        return "done"

    monkeypatch.setattr(flow_engine, "_update_node_state", update_state)
    monkeypatch.setattr(flow_engine, "_exec_samurai", execute_samurai)
    monkeypatch.setattr(flow_engine, "_finalize_node_skills", update_state)
    monkeypatch.setattr(flow_engine, "_node_uses_active_skill_context", lambda *_args: False)

    node = SimpleNamespace(
        id=uuid.uuid4(),
        flow_id=uuid.uuid4(),
        node_type="samurai",
        label="Attached instructions",
        config={
            "task_description": "Old typed task",
            "instruction_file": {
                "filename": "instruction.md",
                "path": str(instruction),
            },
        },
    )

    result = await flow_engine._execute_single_node(uuid.uuid4(), node, {}, {})

    assert result == "done"
    dispatched = captured["config"]
    assert isinstance(dispatched, dict)
    assert dispatched["task_description"] == "Attached Samurai task"
    assert dispatched["_instruction_file_resolved"] is True


@pytest.mark.asyncio
async def test_samurai_instruction_file_cannot_read_outside_flow_uploads(tmp_path, monkeypatch):
    upload_root = tmp_path / "uploads"
    outside = tmp_path / "outside.md"
    outside.write_text("must not be read", encoding="utf-8")
    monkeypatch.setattr(settings, "uploads_path", upload_root)

    with pytest.raises(ValueError, match="must be an AgentFlow upload"):
        await flow_engine._resolve_samurai_task_description(
            {
                "task_description": "fallback",
                "instruction_file": {"filename": outside.name, "path": str(outside)},
            }
        )


def test_legacy_failure_sentinels_are_real_failures():
    with pytest.raises(RuntimeError, match="Office App Mode is disabled"):
        flow_engine._validated_node_result("[BLOCKED] Office App Mode is disabled")
    with pytest.raises(RuntimeError, match="Permission denied"):
        flow_engine._validated_node_result("[ERROR] Permission denied")


@pytest.mark.asyncio
async def test_pdf_read_uses_workspace_file_without_office_app_mode(tmp_path, monkeypatch):
    from shogun.office import config as office_config
    from shogun.services import file_formats

    pdf_path = tmp_path / "Input" / "scheduled.pdf"
    pdf_path.parent.mkdir(parents=True)
    pdf_path.write_bytes(b"%PDF-1.4 test fixture")
    monkeypatch.setattr(settings, "workspace_path", tmp_path)
    monkeypatch.setattr(office_config, "load_office_config", lambda: SimpleNamespace(enabled=False))

    class FakeFileFormatService:
        def __init__(self, session=None, allowed_roots=None):
            assert session is None
            assert allowed_roots == [tmp_path.resolve()]

        async def read(self, *, path, start, end, max_chars):
            assert Path(path) == pdf_path.resolve()
            assert (start, end) == (3, 8)
            assert max_chars == settings.agent_flow_document_max_chars
            return {
                "filename": "scheduled.pdf",
                "content": "scheduled PDF content",
                "truncated": False,
                "metadata": {"start_page": 3, "end_page": 8},
                "warnings": [],
            }

    monkeypatch.setattr(file_formats, "FileFormatService", FakeFileFormatService)

    result = await flow_engine._exec_office(
        {
            "action": "pdf_read",
            "input_path": "Input/scheduled.pdf",
            "start_page": 3,
            "end_page": 8,
        },
        "",
    )

    assert "[PDF: scheduled.pdf; pages 3-8]" in result
    assert "scheduled PDF content" in result


@pytest.mark.asyncio
async def test_excel_create_combines_destination_folder_and_filename(tmp_path, monkeypatch):
    from shogun.office import config as office_config

    monkeypatch.setattr(settings, "workspace_path", tmp_path)
    monkeypatch.setattr(office_config, "load_office_config", lambda: SimpleNamespace(enabled=True))

    result = await flow_engine._exec_office(
        {
            "action": "excel_create",
            "output_path": "Output",
            "sheet_name": "Mapped Data",
        },
        "Column A\tColumn B\nOne\tTwo",
    )

    assert (tmp_path / "Output" / "output.xlsx").is_file()
    assert "Output/output.xlsx" in result


@pytest.mark.asyncio
async def test_scheduled_excel_create_versions_each_run_output(tmp_path, monkeypatch):
    from shogun.office import config as office_config

    monkeypatch.setattr(settings, "workspace_path", tmp_path)
    monkeypatch.setattr(office_config, "load_office_config", lambda: SimpleNamespace(enabled=True))
    run_ids = [
        uuid.UUID("12345678-0000-0000-0000-000000000001"),
        uuid.UUID("87654321-0000-0000-0000-000000000002"),
    ]

    results = [
        await flow_engine._exec_office(
            {
                "action": "excel_create",
                "output_path": "Output/result.xlsx",
                "sheet_name": "Mapped Data",
            },
            f"Run\tValue\n{index}\tSaved",
            run_id=run_id,
            trigger_type="scheduled",
        )
        for index, run_id in enumerate(run_ids, 1)
    ]

    outputs = sorted((tmp_path / "Output").glob("result_v*.xlsx"))
    assert len(outputs) == 2
    assert any(path.name.endswith("_12345678.xlsx") for path in outputs)
    assert any(path.name.endswith("_87654321.xlsx") for path in outputs)
    assert not (tmp_path / "Output" / "result.xlsx").exists()
    assert all(str(path.relative_to(tmp_path)).replace("\\", "/") in "\n".join(results) for path in outputs)


@pytest.mark.asyncio
async def test_scheduled_workspace_write_versions_each_run_output(tmp_path, monkeypatch):
    from shogun.services import posture_guard

    async def allow_workspace():
        return {"workspace_enabled": True}

    monkeypatch.setattr(settings, "workspace_path", tmp_path)
    monkeypatch.setattr(posture_guard, "get_posture_permissions", allow_workspace)

    result = await flow_engine._exec_workspace(
        {"action": "write_file", "path": "Output/summary.txt"},
        "scheduled content",
        run_id=uuid.UUID("abcdef12-0000-0000-0000-000000000003"),
        trigger_type="scheduled",
    )

    outputs = list((tmp_path / "Output").glob("summary_v*_abcdef12.txt"))
    assert len(outputs) == 1
    assert outputs[0].read_text(encoding="utf-8") == "scheduled content"
    assert "Output/summary_v" in result
    assert not (tmp_path / "Output" / "summary.txt").exists()


@pytest.mark.asyncio
async def test_excel_create_converts_markdown_table_to_columns(tmp_path, monkeypatch):
    import openpyxl

    from shogun.office import config as office_config

    monkeypatch.setattr(settings, "workspace_path", tmp_path)
    monkeypatch.setattr(office_config, "load_office_config", lambda: SimpleNamespace(enabled=True))
    markdown = """Narrative that should not become a worksheet row.

| Item | Quantity | Date |
| --- | ---: | --- |
| 140000 | 26 | 21.07.2026 |
| Item | Quantity | Date |
| 140006 | 3 | 21.07.2026 |
"""

    await flow_engine._exec_office(
        {"action": "excel_create", "output_path": "Output/result.xlsx", "sheet_name": "Mapped Data"},
        markdown,
    )

    workbook = openpyxl.load_workbook(tmp_path / "Output" / "result.xlsx", read_only=True)
    try:
        rows = list(workbook["Mapped Data"].iter_rows(values_only=True))
    finally:
        workbook.close()
    assert rows == [
        ("Item", "Quantity", "Date"),
        ("140000", "26", "21.07.2026"),
        ("140006", "3", "21.07.2026"),
    ]


@pytest.mark.asyncio
async def test_excel_write_places_two_dimensional_json_in_a2_c3(tmp_path, monkeypatch, caplog):
    import logging

    import openpyxl

    from shogun.office import config as office_config

    source = tmp_path / "Input" / "target.xlsx"
    source.parent.mkdir(parents=True)
    workbook = openpyxl.Workbook()
    workbook.active.title = "Sheet1"
    workbook.save(source)
    workbook.close()

    monkeypatch.setattr(settings, "workspace_path", tmp_path)
    monkeypatch.setattr(office_config, "load_office_config", lambda: SimpleNamespace(enabled=True))
    caplog.set_level(logging.INFO, logger="shogun.office.adapters.excel")

    result = await flow_engine._exec_office(
        {
            "action": "excel_write",
            "input_path": "Input/target.xlsx",
            "output_path": "Output/result.xlsx",
            "sheet_name": "Sheet1",
            "start_range": "A2",
        },
        '```json\n[["TEST-A","TEST-B","TEST-C"],["TEST-D","TEST-E","TEST-F"]]\n```',
    )

    written = openpyxl.load_workbook(tmp_path / "Output" / "result.xlsx", read_only=True)
    try:
        assert [[written["Sheet1"].cell(row, col).value for col in range(1, 4)] for row in range(2, 4)] == [
            ["TEST-A", "TEST-B", "TEST-C"],
            ["TEST-D", "TEST-E", "TEST-F"],
        ]
    finally:
        written.close()
    assert "Output/result.xlsx" in result
    assert "runtime_type=list" in caplog.text
    assert "first_is_array=True" in caplog.text
    assert "rows=2" in caplog.text
    assert "range=A2" in caplog.text


def test_structured_chunk_matrices_are_validated_and_deduplicated():
    merged = flow_engine._merge_structured_chunk_matrices(
        ['[["A", 1, ""]]', '[["A", 1, ""], ["B", 2, ""]]'],
        "Return only one valid two-dimensional array. Do not create duplicate rows.",
        '[MACHINE-READABLE TEMPLATE MANIFEST]\n{"logical_columns": 3}',
        {"deduplicate_rows": True},
    )

    assert merged is not None
    assert __import__("json").loads(merged) == [["A", 1, ""], ["B", 2, ""]]


def test_structured_chunk_matrices_preserve_identical_source_occurrences_by_default():
    merged = flow_engine._merge_structured_chunk_matrices(
        ['[["A", 100], ["A", 100]]'],
        "Return valid rows and do not create parser-overlap duplicates.",
        '[MACHINE-READABLE TEMPLATE MANIFEST]\n{"logical_columns": 2}',
        {},
        force_matrix_output=True,
    )

    assert merged is not None
    assert __import__("json").loads(merged) == [["A", 100], ["A", 100]]


def test_structured_chunk_matrix_rejects_markdown_summary():
    with pytest.raises(ValueError, match="two-dimensional JSON array"):
        flow_engine._merge_structured_chunk_matrices(
            ["| Item | Quantity |\n| --- | --- |\n| A | 1 |"],
            "Return only one valid two-dimensional array.",
            '[MACHINE-READABLE TEMPLATE MANIFEST]\n{"logical_columns": 2}',
            {},
        )


def test_excel_template_contract_enforces_matrix_without_prompt_wording():
    fixed_context = """[FILE TEMPLATE CONTRACT]
Format: xlsx
[MACHINE-READABLE TEMPLATE MANIFEST]
{"kind": "excel", "logical_columns": 2}
"""

    merged = flow_engine._merge_structured_chunk_matrices(
        ['[["A", 1], ["B", 2]]'],
        "Extract all relevant records from the source document.",
        fixed_context,
        {},
    )

    assert __import__("json").loads(merged) == [["A", 1], ["B", 2]]


def test_downstream_excel_contract_forces_chunk_matrix_merge_without_template():
    merged = flow_engine._merge_structured_chunk_matrices(
        ['[["A", 1]]', '[["B", 2]]'],
        "Extract the records.",
        "",
        {},
        force_matrix_output=True,
    )

    assert __import__("json").loads(merged) == [["A", 1], ["B", 2]]


def test_excel_template_contract_rejects_summary_without_prompt_wording():
    fixed_context = """[FILE TEMPLATE CONTRACT]
Format: xlsx
[MACHINE-READABLE TEMPLATE MANIFEST]
{"kind": "excel", "logical_columns": 2}
"""

    with pytest.raises(ValueError, match="two-dimensional JSON array"):
        flow_engine._merge_structured_chunk_matrices(
            ["| Item | Quantity |\n| --- | --- |\n| A | 1 |"],
            "Extract all relevant records from the source document.",
            fixed_context,
            {},
        )


def test_model_context_splitting_preserves_source_units():
    text = "prefix\n" + "".join(
        f"--- Page {index} ---\n" + (str(index) * 700) + "\n"
        for index in range(1, 5)
    )

    chunks = flow_engine._split_model_context(text, 1200)

    assert "".join(chunks) == text
    assert all(chunk.count("--- Page") <= 2 for chunk in chunks)


def test_profile_context_keeps_section_continuation_pages_together():
    first_section = (
        "--- Page 1 ---\nRecord: 140000\n"
        + ("A" * 450)
        + "\n--- Page 2 ---\ncontinuation without a new material\n"
        + ("B" * 350)
    )
    second_section = "\n--- Page 3 ---\nRecord: 140006\n" + ("C" * 700)
    text = "[Document]\n" + first_section + second_section
    profile = {
        "id": "record_sections_v1",
        "adapter": "sectioned_record_matrix_v1",
        "parameters": {"section_pattern": r"(?m)^Record: (?P<section_id>\S+)"},
    }

    chunks = flow_engine._split_model_context(text, 1000, profile=profile)

    assert "".join(chunks) == text
    first_chunk = next(chunk for chunk in chunks if "Record: 140000" in chunk)
    assert "continuation without a new material" in first_chunk
    assert "Record: 140006" not in first_chunk


def test_default_source_units_use_generic_page_boundaries():
    text = (
        "--- Page 1 ---\nRecord: A1\nmaster data\n"
        "--- Page 2 ---\ncontinuation data\n"
    )

    assert flow_engine._model_source_units(text) == [
        "--- Page 1 ---\nRecord: A1\nmaster data\n",
        "--- Page 2 ---\ncontinuation data\n",
    ]
